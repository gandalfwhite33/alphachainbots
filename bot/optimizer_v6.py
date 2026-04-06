#!/usr/bin/env python3
"""
optimizer_v6.py — Optimizador de parámetros v6 para AlphaChainBots.
Extiende v5 con 30+ nuevos filtros: MACD, ADX, Supertrend, Ichimoku, Stoch RSI,
CCI, Williams %R, Momentum, Bollinger Bands, Keltner, OBV, VWAP, CVD, Market
Structure, Order Blocks, Pivots, Fib Retracement, RSI Divergence, BTC Correlation,
Fear & Greed, Funding Rate, Kelly sizing, Chandelier exit, min_confluences.

Uso:
    python optimizer_v6.py [--coin BTC] [--direction both] [--days 365]
                           [--samples 500000] [--out ./resultados_btc_v6]
                           [--workers N] [--resume]
"""

import os, sys, json, time, random, argparse, hashlib, pickle
import platform, traceback, multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Optional, List, Dict, Tuple
from pathlib import Path

import numpy as np
import requests

# ─── VERSIÓN Y CONSTANTES ─────────────────────────────────────────────────────
VERSION        = "6.0.0"
INITIAL_EQUITY = 10_000.0
HL_URL         = "https://api.hyperliquid.xyz/info"
BINANCE_FUND   = "https://fapi.binance.com/fapi/v1/fundingRate"
FNG_URL        = "https://api.alternative.me/fng/?limit=1"

COINS      = ["BTC"]
COIN_ID    = "btc"
DIRECTION  = "both"   # sobrescrito por --direction
TIMEFRAMES = ["15m", "30m", "1h", "2h", "4h"]

INTERVAL_MS = {
    "15m": 900_000, "30m": 1_800_000, "1h": 3_600_000,
    "2h": 7_200_000, "4h": 14_400_000,
}

CHECKPOINT_EVERY = 20_000
PRINT_TOP_EVERY  = 25_000

# ─── OPCIONES DE PARÁMETROS ───────────────────────────────────────────────────
# ── v5 (mantenidos) ──────────────────────────────────────────────────────────
MA_PAIRS      = [("ema",8,21),("ema",13,34),("ema",21,55),("ema",20,50),
                 ("sma",50,100),("sma",100,200)]
LEVERAGES     = [2, 3, 5, 10, 15]
TRAILING_PCTS = [0.003,0.005,0.008,0.010,0.015,0.020,0.030]
SL_TYPES      = ["fixed","trailing","atr"]
FIB_MODES     = ["required","optional","disabled"]
RSI_FILTERS   = ["none","rsi50","rsi55"]
EMA200_FILT   = ["none","soft","strict"]
ATR_FILTERS   = ["none","min","max"]
COMPOUNDS     = [True, False]
TP_FIBS       = [True, False]
TIME_FILTERS  = ["none","london_ny","asia"]
VOL_PROFILES  = ["disabled","strict","relaxed"]
LIQ_CONFIRMS  = [True, False]
RISK_PCTS     = [0.01,0.02,0.03,0.05]

# ── v6 nuevos ─────────────────────────────────────────────────────────────────
MACD_FILTERS     = ["none","signal_cross","histogram","divergence"]
ADX_FILTERS      = ["none","20","25","30"]
ST_FILTERS       = ["none","required"]
ICHI_FILTERS     = ["none","above_cloud","tk_cross"]
STOCH_RSI        = ["none","oversold_ob","cross"]
CCI_FILTERS      = ["none","100","200"]
WILLIAMS_R       = ["none","required"]
MOM_FILTERS      = ["none","positive","accelerating"]
BB_FILTERS       = ["none","breakout","squeeze","mean_reversion"]
ATR_VOL          = ["none","low","medium","high"]
KELT_FILTERS     = ["none","breakout","squeeze"]
OBV_FILTERS      = ["none","trend","divergence"]
VWAP_FILTERS     = ["none","required"]
VOL_DELTA        = ["none","confirm","divergence"]
CVD_FILTERS      = ["none","confirm","divergence"]
MKT_STRUCT       = ["none","required"]
BREAKOUT_RANGE   = ["none","10","20","50"]
CANDLE_PAT       = ["none","hammer","engulfing","doji"]
ORDER_BLOCK      = ["none","required"]
PIVOT_FILTERS    = ["none","daily","weekly"]
SR_BREAKOUT      = ["none","20","50","100"]
FIB_RET          = ["none","38","50","61"]
RSI_DIV          = ["none","required","bonus"]
BTC_CORR         = ["none","required"]
FUNDING_FILT     = ["none","extreme","moderate"]
FNG_FILTERS      = ["none","extreme_fear","extreme_greed","both"]
SESSION_FILT     = ["none","asia","london","ny","london_ny_overlap","all_sessions"]
POS_SIZING       = ["fixed","kelly","atr_based","volatility_adjusted"]
MAX_TRADES_DAY   = [0, 1, 2, 5, 10]   # 0 = unlimited
TRAIL_TYPES      = ["fixed","atr_dynamic","chandelier"]
MIN_CONF         = [0, 1, 2, 3, 4, 5]

# Globales cargados en main() y compartidos por fork
_FEAR_GREED_IDX: Optional[float] = None  # 0–100
_FUNDING_RATE:   Optional[float] = None  # fracción decimal

_PARAM_FIELDS_SET = {
    "interval","ma_type","ma_fast","ma_slow","leverage","trailing_pct",
    "sl_type","fib_mode","rsi_filter","ema200_filter","atr_filter",
    "compound","tp_fib","time_filter","vol_profile","liq_confirm","risk_pct",
    # v6
    "macd_filter","adx_filter","supertrend_filter","ichimoku_filter",
    "stoch_rsi","cci_filter","williams_r","momentum_filter",
    "bb_filter","atr_volatility","keltner_filter",
    "obv_filter","vwap_filter","volume_delta","cvd_filter",
    "market_structure","breakout_range","candle_pattern","order_block",
    "pivot_filter","sr_breakout","fib_retracement","rsi_divergence",
    "btc_correlation","funding_filter","fear_greed_filter","session_filter",
    "position_sizing","max_trades_day","trailing_type","min_confluences",
}
_METRIC_FIELDS = {
    "total_pnl","total_pnl_pct","win_rate","total_trades",
    "max_drawdown","sharpe","profit_factor","best_trade","worst_trade",
}


def ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


# ─── DATACLASS DE PARÁMETROS ──────────────────────────────────────────────────
@dataclass
class OptParams:
    # v5
    interval:      str
    ma_type:       str
    ma_fast:       int
    ma_slow:       int
    leverage:      float
    trailing_pct:  float
    sl_type:       str
    fib_mode:      str
    rsi_filter:    str
    ema200_filter: str
    atr_filter:    str
    compound:      bool
    tp_fib:        bool
    time_filter:   str
    vol_profile:   str
    liq_confirm:   bool
    risk_pct:      float
    # v6 – tendencia
    macd_filter:        str
    adx_filter:         str
    supertrend_filter:  str
    ichimoku_filter:    str
    # v6 – momentum
    stoch_rsi:          str
    cci_filter:         str
    williams_r:         str
    momentum_filter:    str
    # v6 – volatilidad
    bb_filter:          str
    atr_volatility:     str
    keltner_filter:     str
    # v6 – volumen
    obv_filter:         str
    vwap_filter:        str
    volume_delta:       str
    cvd_filter:         str
    # v6 – estructura de precio
    market_structure:   str
    breakout_range:     str
    candle_pattern:     str
    order_block:        str
    pivot_filter:       str
    sr_breakout:        str
    fib_retracement:    str
    rsi_divergence:     str
    # v6 – filtros de mercado
    btc_correlation:    str
    funding_filter:     str
    fear_greed_filter:  str
    session_filter:     str
    # v6 – gestión de capital
    position_sizing:    str
    max_trades_day:     int
    trailing_type:      str
    min_confluences:    int

    def param_key(self) -> str:
        d = {k: v for k, v in asdict(self).items()}
        return hashlib.md5(json.dumps(d, sort_keys=True).encode()).hexdigest()


def random_params(rng: random.Random) -> OptParams:
    ma = rng.choice(MA_PAIRS)
    return OptParams(
        interval=rng.choice(TIMEFRAMES), ma_type=ma[0], ma_fast=ma[1], ma_slow=ma[2],
        leverage=rng.choice(LEVERAGES), trailing_pct=rng.choice(TRAILING_PCTS),
        sl_type=rng.choice(SL_TYPES), fib_mode=rng.choice(FIB_MODES),
        rsi_filter=rng.choice(RSI_FILTERS), ema200_filter=rng.choice(EMA200_FILT),
        atr_filter=rng.choice(ATR_FILTERS), compound=rng.choice(COMPOUNDS),
        tp_fib=rng.choice(TP_FIBS), time_filter=rng.choice(TIME_FILTERS),
        vol_profile=rng.choice(VOL_PROFILES), liq_confirm=rng.choice(LIQ_CONFIRMS),
        risk_pct=rng.choice(RISK_PCTS),
        macd_filter=rng.choice(MACD_FILTERS), adx_filter=rng.choice(ADX_FILTERS),
        supertrend_filter=rng.choice(ST_FILTERS), ichimoku_filter=rng.choice(ICHI_FILTERS),
        stoch_rsi=rng.choice(STOCH_RSI), cci_filter=rng.choice(CCI_FILTERS),
        williams_r=rng.choice(WILLIAMS_R), momentum_filter=rng.choice(MOM_FILTERS),
        bb_filter=rng.choice(BB_FILTERS), atr_volatility=rng.choice(ATR_VOL),
        keltner_filter=rng.choice(KELT_FILTERS), obv_filter=rng.choice(OBV_FILTERS),
        vwap_filter=rng.choice(VWAP_FILTERS), volume_delta=rng.choice(VOL_DELTA),
        cvd_filter=rng.choice(CVD_FILTERS), market_structure=rng.choice(MKT_STRUCT),
        breakout_range=rng.choice(BREAKOUT_RANGE), candle_pattern=rng.choice(CANDLE_PAT),
        order_block=rng.choice(ORDER_BLOCK), pivot_filter=rng.choice(PIVOT_FILTERS),
        sr_breakout=rng.choice(SR_BREAKOUT), fib_retracement=rng.choice(FIB_RET),
        rsi_divergence=rng.choice(RSI_DIV), btc_correlation=rng.choice(BTC_CORR),
        funding_filter=rng.choice(FUNDING_FILT), fear_greed_filter=rng.choice(FNG_FILTERS),
        session_filter=rng.choice(SESSION_FILT), position_sizing=rng.choice(POS_SIZING),
        max_trades_day=rng.choice(MAX_TRADES_DAY), trailing_type=rng.choice(TRAIL_TYPES),
        min_confluences=rng.choice(MIN_CONF),
    )


def perturb(p: OptParams, rng: random.Random) -> OptParams:
    d = asdict(p)
    all_keys = list(_PARAM_FIELDS_SET - {"ma_type","ma_fast","ma_slow"}) + ["ma_pair"]
    key = rng.choice(all_keys)
    opts = {
        "interval": TIMEFRAMES, "leverage": LEVERAGES, "trailing_pct": TRAILING_PCTS,
        "sl_type": SL_TYPES, "fib_mode": FIB_MODES, "rsi_filter": RSI_FILTERS,
        "ema200_filter": EMA200_FILT, "atr_filter": ATR_FILTERS,
        "compound": COMPOUNDS, "tp_fib": TP_FIBS, "time_filter": TIME_FILTERS,
        "vol_profile": VOL_PROFILES, "liq_confirm": LIQ_CONFIRMS, "risk_pct": RISK_PCTS,
        "macd_filter": MACD_FILTERS, "adx_filter": ADX_FILTERS,
        "supertrend_filter": ST_FILTERS, "ichimoku_filter": ICHI_FILTERS,
        "stoch_rsi": STOCH_RSI, "cci_filter": CCI_FILTERS,
        "williams_r": WILLIAMS_R, "momentum_filter": MOM_FILTERS,
        "bb_filter": BB_FILTERS, "atr_volatility": ATR_VOL,
        "keltner_filter": KELT_FILTERS, "obv_filter": OBV_FILTERS,
        "vwap_filter": VWAP_FILTERS, "volume_delta": VOL_DELTA,
        "cvd_filter": CVD_FILTERS, "market_structure": MKT_STRUCT,
        "breakout_range": BREAKOUT_RANGE, "candle_pattern": CANDLE_PAT,
        "order_block": ORDER_BLOCK, "pivot_filter": PIVOT_FILTERS,
        "sr_breakout": SR_BREAKOUT, "fib_retracement": FIB_RET,
        "rsi_divergence": RSI_DIV, "btc_correlation": BTC_CORR,
        "funding_filter": FUNDING_FILT, "fear_greed_filter": FNG_FILTERS,
        "session_filter": SESSION_FILT, "position_sizing": POS_SIZING,
        "max_trades_day": MAX_TRADES_DAY, "trailing_type": TRAIL_TYPES,
        "min_confluences": MIN_CONF,
    }
    if key == "ma_pair":
        ma = rng.choice(MA_PAIRS)
        d["ma_type"] = ma[0]; d["ma_fast"] = ma[1]; d["ma_slow"] = ma[2]
    elif key in opts:
        d[key] = rng.choice(opts[key])
    return OptParams(**d)


# ─── CANDLE FETCHING ──────────────────────────────────────────────────────────
def _fetch_hl(coin: str, interval: str, days: int) -> Optional[np.ndarray]:
    end_ms   = int(time.time() * 1000)
    start_ms = end_ms - days * 86_400_000
    try:
        r = requests.post(HL_URL, json={
            "type": "candleSnapshot",
            "req":  {"coin": coin, "interval": interval,
                     "startTime": start_ms, "endTime": end_ms},
        }, timeout=30)
        if not r.ok: return None
        candles = r.json()
        if not candles: return None
        return np.array([
            [c["T"], float(c["o"]), float(c["h"]), float(c["l"]),
             float(c["c"]), float(c.get("v", 0))]
            for c in candles
        ], dtype=np.float64)
    except Exception as e:
        print(f"[{ts()}]   ! Error {coin}/{interval}: {e}")
        return None


def load_or_fetch(cache_dir: Path, coin: str, interval: str, days: int) -> Optional[np.ndarray]:
    cache_file = cache_dir / f"{coin}_{interval}_{days}d.pkl"
    if cache_file.exists():
        try:
            with open(cache_file, "rb") as f:
                data = pickle.load(f)
            if data is not None and len(data) > 50:
                return data
        except Exception:
            pass
    time.sleep(random.uniform(0.05, 0.25))
    arr = _fetch_hl(coin, interval, days)
    if arr is not None and len(arr) > 50:
        try:
            with open(cache_file, "wb") as f:
                pickle.dump(arr, f)
        except Exception:
            pass
    return arr


# ─── INDICADORES VECTORIZADOS ─────────────────────────────────────────────────
def _ema(arr: np.ndarray, n: int) -> np.ndarray:
    if len(arr) < n:
        return np.full_like(arr, np.nan)
    out = np.empty_like(arr)
    out[:n-1] = np.nan
    out[n-1]  = float(np.nanmean(arr[:n]))
    k = 2.0 / (n + 1); k1 = 1.0 - k
    prev = out[n-1]
    for i in range(n, len(arr)):
        if arr[i] == arr[i]:   # not NaN
            prev = arr[i] * k + prev * k1
        out[i] = prev
    return out


def _sma(arr: np.ndarray, n: int) -> np.ndarray:
    if len(arr) < n:
        return np.full_like(arr, np.nan)
    out = np.full_like(arr, np.nan)
    cs = np.cumsum(np.where(np.isnan(arr), 0, arr))
    out[n-1] = cs[n-1] / n
    out[n:]  = (cs[n:] - cs[:len(arr)-n]) / n
    return out


def _rsi(closes: np.ndarray, n: int = 14) -> np.ndarray:
    out = np.full_like(closes, np.nan)
    if len(closes) < n + 1: return out
    d = np.diff(closes)
    gains = np.maximum(d, 0.0); losses = np.maximum(-d, 0.0)
    ag = float(np.mean(gains[:n])); al = float(np.mean(losses[:n]))
    for i in range(n, len(d)):
        ag = (ag*(n-1) + gains[i]) / n
        al = (al*(n-1) + losses[i]) / n
        rs = ag / al if al > 0 else 100.0
        out[i+1] = 100.0 - 100.0 / (1.0 + rs)
    return out


def _atr(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 14) -> np.ndarray:
    out = np.full_like(close, np.nan)
    if len(close) < n + 1: return out
    tr = np.maximum(high[1:]-low[1:],
         np.maximum(np.abs(high[1:]-close[:-1]), np.abs(low[1:]-close[:-1])))
    out[n] = float(np.mean(tr[:n]))
    for i in range(n, len(tr)):
        out[i+1] = (out[i]*(n-1) + tr[i]) / n
    return out


def _roll_max(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w: return out
    shape   = arr.shape[:-1] + (arr.shape[-1]-w+1, w)
    strides = arr.strides + (arr.strides[-1],)
    wins = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w-1:] = wins.max(axis=1)
    return out


def _roll_min(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w: return out
    shape   = arr.shape[:-1] + (arr.shape[-1]-w+1, w)
    strides = arr.strides + (arr.strides[-1],)
    wins = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w-1:] = wins.min(axis=1)
    return out


def _roll_std(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w: return out
    shape   = arr.shape[:-1] + (arr.shape[-1]-w+1, w)
    strides = arr.strides + (arr.strides[-1],)
    wins = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w-1:] = wins.std(axis=1)
    return out


def _macd(closes: np.ndarray, fast: int = 12, slow: int = 26,
          sig: int = 9) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    L = len(closes)
    ef = _ema(closes, fast); es = _ema(closes, slow)
    line   = ef - es                          # shape (L,)
    signal = np.full(L, np.nan)
    # Find first non-NaN index in line, compute EMA of signal from there
    valid_idx = np.where(~np.isnan(line))[0]
    if len(valid_idx) >= sig:
        start = valid_idx[0]
        sig_sub = _ema(line[start:], sig)     # shape (L - start,)
        signal[start:start + len(sig_sub)] = sig_sub
    hist = line - signal
    return line, signal, hist                 # all shape (L,)


def _adx(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 14) -> np.ndarray:
    L   = len(close)
    adx = np.full(L, np.nan)
    if L < n * 2 + 2: return adx
    # TR and DM arrays have length L-1 (diff over consecutive bars)
    tr  = np.maximum(high[1:]-low[1:],
          np.maximum(np.abs(high[1:]-close[:-1]), np.abs(low[1:]-close[:-1])))
    dmp = np.where((high[1:]-high[:-1]) > (low[:-1]-low[1:]),
                   np.maximum(high[1:]-high[:-1], 0.0), 0.0)
    dmm = np.where((low[:-1]-low[1:]) > (high[1:]-high[:-1]),
                   np.maximum(low[:-1]-low[1:], 0.0), 0.0)
    N = len(tr)  # N = L - 1
    atr_w = np.zeros(N); dmp_w = np.zeros(N); dmm_w = np.zeros(N)
    if n - 1 < N:
        atr_w[n-1] = float(np.sum(tr[:n]))
        dmp_w[n-1] = float(np.sum(dmp[:n]))
        dmm_w[n-1] = float(np.sum(dmm[:n]))
    for i in range(n, N):
        atr_w[i] = atr_w[i-1] - atr_w[i-1]/n + tr[i]
        dmp_w[i] = dmp_w[i-1] - dmp_w[i-1]/n + dmp[i]
        dmm_w[i] = dmm_w[i-1] - dmm_w[i-1]/n + dmm[i]
    with np.errstate(divide="ignore", invalid="ignore"):
        dip = np.where(atr_w > 0, dmp_w/atr_w*100, 0.0)
        dim = np.where(atr_w > 0, dmm_w/atr_w*100, 0.0)
        dx  = np.where(dip+dim > 0, np.abs(dip-dim)/(dip+dim)*100, 0.0)
    # adx_arr is indexed over the diff space (length N = L-1)
    # We map it back to the close space with a +1 offset
    adx_arr = np.full(N, np.nan)
    first = n * 2 - 2
    if first < N:
        adx_arr[first] = float(np.mean(dx[n-1:n*2-1]))
        for i in range(first + 1, N):
            adx_arr[i] = (adx_arr[i-1]*(n-1) + dx[i]) / n
    # adx_arr[i] corresponds to close[i+1] — shift by 1 to align
    adx[1:] = adx_arr
    return adx   # shape (L,)


def _supertrend(high: np.ndarray, low: np.ndarray, close: np.ndarray,
                n: int = 10, m: float = 3.0) -> np.ndarray:
    """Returns direction array: 1=bullish, -1=bearish, 0=undefined."""
    direction = np.zeros(len(close))
    atr = _atr(high, low, close, n)
    hl2 = (high + low) / 2.0
    ub_basic = hl2 + m * atr
    lb_basic = hl2 - m * atr
    ub = np.copy(ub_basic); lb = np.copy(lb_basic)
    for i in range(1, len(close)):
        if np.isnan(atr[i]): continue
        ub[i] = ub_basic[i] if ub_basic[i] < ub[i-1] or close[i-1] > ub[i-1] else ub[i-1]
        lb[i] = lb_basic[i] if lb_basic[i] > lb[i-1] or close[i-1] < lb[i-1] else lb[i-1]
        if close[i] > ub[i]:   direction[i] = 1
        elif close[i] < lb[i]: direction[i] = -1
        else:                  direction[i] = direction[i-1]
    return direction


def _ichimoku(high: np.ndarray, low: np.ndarray, close: np.ndarray):
    """Returns tenkan, kijun, span_a (unshifted), span_b (unshifted)."""
    tenkan = (_roll_max(high, 9)  + _roll_min(low, 9))  / 2.0
    kijun  = (_roll_max(high, 26) + _roll_min(low, 26)) / 2.0
    span_a = (tenkan + kijun) / 2.0
    span_b = (_roll_max(high, 52) + _roll_min(low, 52)) / 2.0
    return tenkan, kijun, span_a, span_b


def _stoch_rsi(closes: np.ndarray, rsi_n: int = 14, stoch_n: int = 14,
               k_s: int = 3, d_s: int = 3) -> Tuple[np.ndarray, np.ndarray]:
    rsi = _rsi(closes, rsi_n)
    rsi_max = _roll_max(rsi, stoch_n)
    rsi_min = _roll_min(rsi, stoch_n)
    with np.errstate(invalid="ignore", divide="ignore"):
        raw_k = np.where(rsi_max - rsi_min > 0,
                         (rsi - rsi_min) / (rsi_max - rsi_min) * 100.0, 50.0)
    raw_k[np.isnan(rsi)] = np.nan
    k = _sma(raw_k, k_s)
    d = _sma(k, d_s)
    return k, d


def _cci(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 20) -> np.ndarray:
    tp = (high + low + close) / 3.0
    tp_sma = _sma(tp, n)
    out = np.full_like(close, np.nan)
    for i in range(n-1, len(close)):
        mean_dev = float(np.mean(np.abs(tp[i-n+1:i+1] - tp_sma[i])))
        if mean_dev > 0:
            out[i] = (tp[i] - tp_sma[i]) / (0.015 * mean_dev)
    return out


def _williams_r_arr(high: np.ndarray, low: np.ndarray,
                    close: np.ndarray, n: int = 14) -> np.ndarray:
    hh = _roll_max(high, n); ll = _roll_min(low, n)
    with np.errstate(invalid="ignore", divide="ignore"):
        wr = -100.0 * (hh - close) / np.where(hh - ll > 0, hh - ll, np.nan)
    return wr


def _bbands(closes: np.ndarray, n: int = 20,
            k: float = 2.0) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    sma = _sma(closes, n); std = _roll_std(closes, n)
    upper = sma + k * std; lower = sma - k * std
    with np.errstate(invalid="ignore", divide="ignore"):
        width = np.where(sma > 0, (upper - lower) / sma, np.nan)
    return upper, lower, width


def _keltner(high: np.ndarray, low: np.ndarray, close: np.ndarray,
             n: int = 20, m: float = 2.0) -> Tuple[np.ndarray, np.ndarray]:
    ema = _ema(close, n); atr = _atr(high, low, close, n)
    return ema + m * atr, ema - m * atr


def _obv(close: np.ndarray, volume: np.ndarray) -> np.ndarray:
    obv = np.zeros(len(close))
    for i in range(1, len(close)):
        if close[i] > close[i-1]:   obv[i] = obv[i-1] + volume[i]
        elif close[i] < close[i-1]: obv[i] = obv[i-1] - volume[i]
        else:                        obv[i] = obv[i-1]
    return obv


def _vwap_daily(timestamps: np.ndarray, high: np.ndarray,
                low: np.ndarray, close: np.ndarray,
                volume: np.ndarray) -> np.ndarray:
    tp = (high + low + close) / 3.0
    vwap = np.full_like(close, np.nan)
    cum_tv = 0.0; cum_v = 0.0; prev_day = -1
    days = (timestamps // 86_400_000).astype(np.int64)
    for i in range(len(close)):
        if days[i] != prev_day:
            cum_tv = 0.0; cum_v = 0.0; prev_day = int(days[i])
        cum_tv += tp[i] * volume[i]; cum_v += volume[i]
        vwap[i] = cum_tv / cum_v if cum_v > 0 else close[i]
    return vwap


def _cvd(opens: np.ndarray, closes: np.ndarray,
         volumes: np.ndarray) -> np.ndarray:
    """Cumulative volume delta approximation using candle body direction."""
    delta = np.where(closes >= opens, volumes, -volumes)
    return np.cumsum(delta)


def _market_structure(high: np.ndarray, low: np.ndarray, n: int = 10) -> np.ndarray:
    """1=bullish (HH+HL), -1=bearish (LH+LL), 0=neutral."""
    struct = np.zeros(len(high))
    for i in range(n, len(high)):
        ph = high[i-n:i]; pl = low[i-n:i]
        is_hh = high[i] > np.max(ph); is_hl = low[i] > np.min(pl)
        is_lh = high[i] < np.max(ph); is_ll = low[i] < np.min(pl)
        if is_hh and is_hl:   struct[i] = 1
        elif is_lh and is_ll: struct[i] = -1
    return struct


def _order_blocks(closes: np.ndarray, atr: np.ndarray,
                  multiplier: float = 2.0) -> Tuple[np.ndarray, np.ndarray]:
    """Bullish OB: large bearish candle before a bullish move; vice versa."""
    ob_bull = np.zeros(len(closes), dtype=bool)
    ob_bear = np.zeros(len(closes), dtype=bool)
    for i in range(1, len(closes)):
        av = atr[i] if atr[i] == atr[i] else 0.0
        if av <= 0: continue
        body = abs(closes[i] - closes[i-1])
        if body > multiplier * av:
            if closes[i] > closes[i-1]: ob_bear[i] = True   # big bull → bearish OB above
            else:                        ob_bull[i] = True   # big bear → bullish OB below
    return ob_bull, ob_bear


def _pivot_levels(timestamps: np.ndarray, high: np.ndarray,
                  low: np.ndarray, close: np.ndarray,
                  weekly: bool = False) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Daily (or weekly) pivot P, R1, S1."""
    period_ms = 604_800_000 if weekly else 86_400_000
    periods = (timestamps // period_ms).astype(np.int64)
    P  = np.full_like(close, np.nan)
    R1 = np.full_like(close, np.nan)
    S1 = np.full_like(close, np.nan)
    prev_p = -1; ph = pl = pc = 0.0
    for i in range(len(close)):
        p = int(periods[i])
        if p != prev_p and prev_p >= 0:
            pivot = (ph + pl + pc) / 3.0
            R1[i] = 2 * pivot - pl
            S1[i] = 2 * pivot - ph
            P[i]  = pivot
            prev_p = p
        elif p == prev_p and i > 0:
            P[i] = P[i-1]; R1[i] = R1[i-1]; S1[i] = S1[i-1]
        if p != prev_p:
            prev_p = p; ph = high[i]; pl = low[i]; pc = close[i]
        else:
            ph = max(ph, high[i]); pl = min(pl, low[i]); pc = close[i]
    return P, R1, S1


def _rsi_divergence_arr(close: np.ndarray, rsi: np.ndarray,
                        lb: int = 14) -> Tuple[np.ndarray, np.ndarray]:
    """Bull div: price LL, RSI HL. Bear div: price HH, RSI LH."""
    bull = np.zeros(len(close), dtype=bool)
    bear = np.zeros(len(close), dtype=bool)
    for i in range(lb*2, len(close)):
        if close[i] < close[i-lb] and rsi[i] > rsi[i-lb] and rsi[i] == rsi[i]:
            bull[i] = True
        if close[i] > close[i-lb] and rsi[i] < rsi[i-lb] and rsi[i] == rsi[i]:
            bear[i] = True
    return bull, bear


def _hour_mask(ts_arr: np.ndarray, allowed: set) -> np.ndarray:
    hours = (ts_arr // 3_600_000 % 24).astype(np.int32)
    mask  = np.zeros(len(ts_arr), dtype=bool)
    for h in allowed:
        mask |= (hours == h)
    return mask


_LONDON_NY = set(range(7, 22))
_ASIA      = set(range(0, 9))
_NY        = set(range(13, 22))
_LDN_NY_OV = set(range(13, 18))

_FIBO_L = np.array([0.236, 0.382, 0.500, 0.618, 0.786])


def _near_fibo_fast(price: float, hi: float, lo: float) -> bool:
    rng = hi - lo
    if rng <= 0 or price <= 0: return False
    levels = lo + rng * _FIBO_L
    return bool(np.any(np.abs(levels - price) / price <= 0.015))


# ─── WORKER INIT ──────────────────────────────────────────────────────────────
_worker_ind: Dict[str, Dict[str, dict]] = {}


def _worker_init(cache_pkl: str):
    global _worker_ind, _FEAR_GREED_IDX, _FUNDING_RATE, DIRECTION, COIN_ID, COINS
    try:
        with open(cache_pkl, "rb") as f:
            raw: dict = pickle.load(f)
        _FEAR_GREED_IDX = raw.get("_fng")
        _FUNDING_RATE   = raw.get("_funding")
        # Restore globals overridden by main() — critical on Windows (spawn)
        if raw.get("_direction"): DIRECTION = raw["_direction"]
        if raw.get("_coin_id"):   COIN_ID   = raw["_coin_id"]
        if raw.get("_coins"):     COINS     = raw["_coins"]
        candles         = raw.get("candles", {})
        btc_candles     = raw.get("btc_candles", {})
    except Exception:
        _worker_ind = {}
        return

    # Pre-compute BTC trend direction per TF (for correlation filter)
    btc_trend: Dict[str, np.ndarray] = {}
    for tf, arr in btc_candles.items():
        if arr is None or len(arr) < 60: continue
        c = arr[:, 4]
        e20 = _ema(c, 20); e50 = _ema(c, 50)
        btc_trend[tf] = np.where(e20 > e50, 1.0, -1.0)

    _worker_ind = {}
    for tf, coins_dict in candles.items():
        _worker_ind[tf] = {}
        for coin, arr in coins_dict.items():
            if arr is None or len(arr) < 250: continue
            ts_col  = arr[:, 0]
            opens   = arr[:, 1]
            highs   = arr[:, 2]
            lows    = arr[:, 3]
            closes  = arr[:, 4]
            volumes = arr[:, 5]

            atr14 = _atr(highs, lows, closes, 14)
            rsi14 = _rsi(closes, 14)
            ema200 = _ema(closes, 200)
            vol_ma = _sma(volumes, 20)
            roll_hi = _roll_max(highs, 60)
            roll_lo = _roll_min(lows, 60)

            # MACD
            macd_l, macd_s, macd_h = _macd(closes)
            # Divergences: price lower + MACD hist higher → bull; vice versa
            lb = 14
            macd_bull_div = np.zeros(len(closes), dtype=bool)
            macd_bear_div = np.zeros(len(closes), dtype=bool)
            for i in range(lb*2, len(closes)):
                if macd_h[i] != macd_h[i]: continue
                lo_c = np.nanmin(closes[i-lb:i]); lo_mh = np.nanmin(macd_h[i-lb:i])
                hi_c = np.nanmax(closes[i-lb:i]); hi_mh = np.nanmax(macd_h[i-lb:i])
                if closes[i] <= lo_c and macd_h[i] > lo_mh: macd_bull_div[i] = True
                if closes[i] >= hi_c and macd_h[i] < hi_mh: macd_bear_div[i] = True

            adx   = _adx(highs, lows, closes)
            st    = _supertrend(highs, lows, closes)
            tenkan, kijun, span_a, span_b = _ichimoku(highs, lows, closes)
            cloud_top = np.maximum(span_a, span_b)
            cloud_bot = np.minimum(span_a, span_b)

            sk, sd = _stoch_rsi(closes)
            cci     = _cci(highs, lows, closes)
            wr      = _williams_r_arr(highs, lows, closes)

            mom10   = closes - np.roll(closes, 10)
            mom10[:10] = np.nan

            bb_u, bb_l, bb_w = _bbands(closes)
            kelt_u, kelt_l   = _keltner(highs, lows, closes)
            bb_w_sma         = _sma(bb_w, 20)   # for squeeze detection

            obv_arr  = _obv(closes, volumes)
            obv_ema  = _ema(obv_arr, 20)
            obv_bull_div = np.zeros(len(closes), dtype=bool)
            obv_bear_div = np.zeros(len(closes), dtype=bool)
            for i in range(lb*2, len(closes)):
                lo_c = np.nanmin(closes[i-lb:i]); lo_o = np.nanmin(obv_arr[i-lb:i])
                hi_c = np.nanmax(closes[i-lb:i]); hi_o = np.nanmax(obv_arr[i-lb:i])
                if closes[i] <= lo_c and obv_arr[i] > lo_o: obv_bull_div[i] = True
                if closes[i] >= hi_c and obv_arr[i] < hi_o: obv_bear_div[i] = True

            vwap_arr = _vwap_daily(ts_col, highs, lows, closes, volumes)
            cvd_arr  = _cvd(opens, closes, volumes)
            cvd_ema  = _ema(cvd_arr.astype(float), 14)

            ms_arr   = _market_structure(highs, lows)
            ob_bull, ob_bear = _order_blocks(closes, atr14)

            piv_p,  piv_r1,  piv_s1  = _pivot_levels(ts_col, highs, lows, closes, weekly=False)
            wpiv_p, wpiv_r1, wpiv_s1 = _pivot_levels(ts_col, highs, lows, closes, weekly=True)

            rsi_bull_div, rsi_bear_div = _rsi_divergence_arr(closes, rsi14)

            price_mean  = float(np.nanmean(closes))
            atr_valid   = atr14[~np.isnan(atr14)]
            atr_pct_ref = (float(np.nanmean(atr_valid)) / price_mean
                           if len(atr_valid) > 0 and price_mean > 0 else 0.0)
            atr_std     = float(np.nanstd(atr_valid)) if len(atr_valid) > 1 else 0.0
            atr_mean    = float(np.nanmean(atr_valid)) if len(atr_valid) > 0 else 0.0

            # Rolling window breakout arrays (for sr_breakout / breakout_range)
            roll_hi10  = _roll_max(highs, 10);  roll_lo10  = _roll_min(lows, 10)
            roll_hi20  = _roll_max(highs, 20);  roll_lo20  = _roll_min(lows, 20)
            roll_hi50  = _roll_max(highs, 50);  roll_lo50  = _roll_min(lows, 50)
            roll_hi100 = _roll_max(highs, 100); roll_lo100 = _roll_min(lows, 100)

            ind: dict = {
                "ts": ts_col, "opens": opens, "highs": highs, "lows": lows,
                "closes": closes, "volumes": volumes,
                "ema200": ema200, "rsi14": rsi14, "atr14": atr14,
                "vol_ma": vol_ma, "roll_hi": roll_hi, "roll_lo": roll_lo,
                "h_london": _hour_mask(ts_col, _LONDON_NY),
                "h_asia":   _hour_mask(ts_col, _ASIA),
                "h_ny":     _hour_mask(ts_col, _NY),
                "h_lno":    _hour_mask(ts_col, _LDN_NY_OV),
                "atr_pct_ref": atr_pct_ref, "price_mean": price_mean,
                "atr_mean": atr_mean, "atr_std": atr_std,
                # MACD
                "macd_l": macd_l, "macd_s": macd_s, "macd_h": macd_h,
                "macd_bull_div": macd_bull_div, "macd_bear_div": macd_bear_div,
                # ADX
                "adx": adx,
                # Supertrend
                "st": st,
                # Ichimoku
                "cloud_top": cloud_top, "cloud_bot": cloud_bot,
                "tenkan": tenkan, "kijun": kijun,
                # Stoch RSI
                "sk": sk, "sd": sd,
                # CCI
                "cci": cci,
                # Williams %R
                "wr": wr,
                # Momentum
                "mom10": mom10,
                # Bollinger
                "bb_u": bb_u, "bb_l": bb_l, "bb_w": bb_w, "bb_w_sma": bb_w_sma,
                # Keltner
                "kelt_u": kelt_u, "kelt_l": kelt_l,
                # OBV
                "obv": obv_arr, "obv_ema": obv_ema,
                "obv_bull_div": obv_bull_div, "obv_bear_div": obv_bear_div,
                # VWAP
                "vwap": vwap_arr,
                # CVD
                "cvd": cvd_arr.astype(float), "cvd_ema": cvd_ema,
                # Market structure
                "ms": ms_arr,
                # Order blocks
                "ob_bull": ob_bull, "ob_bear": ob_bear,
                # Pivots
                "piv_p": piv_p, "piv_r1": piv_r1, "piv_s1": piv_s1,
                "wpiv_p": wpiv_p, "wpiv_r1": wpiv_r1, "wpiv_s1": wpiv_s1,
                # RSI divergence
                "rsi_bull_div": rsi_bull_div, "rsi_bear_div": rsi_bear_div,
                # Breakout range arrays
                "roll_hi10": roll_hi10, "roll_lo10": roll_lo10,
                "roll_hi20": roll_hi20, "roll_lo20": roll_lo20,
                "roll_hi50": roll_hi50, "roll_lo50": roll_lo50,
                "roll_hi100": roll_hi100, "roll_lo100": roll_lo100,
                # BTC correlation
                "btc_trend": btc_trend.get(tf),
            }

            # Pre-compute MA variants
            for (ma_type, fast, slow) in MA_PAIRS:
                k = f"{ma_type}_{fast}_{slow}"
                fn = _ema if ma_type == "ema" else _sma
                ind[f"maf_{k}"] = fn(closes, fast)
                ind[f"mas_{k}"] = fn(closes, slow)

            _worker_ind[tf][coin] = ind


# ─── SIMULACIÓN FAST ──────────────────────────────────────────────────────────
def _simulate_fast(params: OptParams, coins_ind: Dict[str, dict]) -> dict:
    all_pnl = 0.0; all_trades = 0; all_wins = 0
    gross_win = 0.0; gross_loss = 0.0
    all_best = 0.0; all_worst = 0.0
    equity = INITIAL_EQUITY; peak_equity = INITIAL_EQUITY; max_dd = 0.0
    eq_curve = [INITIAL_EQUITY]

    min_bars  = max(params.ma_slow + 50, 260)
    ma_key    = f"{params.ma_type}_{params.ma_fast}_{params.ma_slow}"
    sl_type   = params.sl_type
    fib_mode  = params.fib_mode
    rsi_filt  = params.rsi_filter
    e200_filt = params.ema200_filter
    atr_filt  = params.atr_filter
    vol_prof  = params.vol_profile
    time_filt = params.time_filter
    trailing  = params.trailing_pct
    compound  = params.compound
    tp_fib    = params.tp_fib
    risk_lev  = params.risk_pct * params.leverage

    # v6 params
    macd_filt   = params.macd_filter
    adx_filt    = params.adx_filter
    st_filt     = params.supertrend_filter
    ichi_filt   = params.ichimoku_filter
    stoch_filt  = params.stoch_rsi
    cci_filt    = params.cci_filter
    wr_filt     = params.williams_r
    mom_filt    = params.momentum_filter
    bb_filt     = params.bb_filter
    atr_vol_f   = params.atr_volatility
    kelt_filt   = params.keltner_filter
    obv_filt    = params.obv_filter
    vwap_filt   = params.vwap_filter
    vdelta_filt = params.volume_delta
    cvd_filt    = params.cvd_filter
    ms_filt     = params.market_structure
    bo_range    = params.breakout_range
    cpat_filt   = params.candle_pattern
    ob_filt     = params.order_block
    piv_filt    = params.pivot_filter
    sr_bo_filt  = params.sr_breakout
    fib_ret_f   = params.fib_retracement
    rsi_div_f   = params.rsi_divergence
    btc_corr    = params.btc_correlation
    fund_filt   = params.funding_filter
    fng_filt    = params.fear_greed_filter
    sess_filt   = params.session_filter
    pos_siz     = params.position_sizing
    max_td      = params.max_trades_day
    trail_type  = params.trailing_type
    min_conf    = params.min_confluences

    # --- Static (global) filters: Fear & Greed and Funding rate ---
    fng_val  = _FEAR_GREED_IDX  # may be None
    fund_val = _FUNDING_RATE    # may be None

    for coin, ind in coins_ind.items():
        closes = ind.get("closes")
        if closes is None or len(closes) < min_bars + 10: continue

        maf = ind.get(f"maf_{ma_key}"); mas = ind.get(f"mas_{ma_key}")
        if maf is None or mas is None: continue

        # Resolve session hour mask
        if time_filt == "london_ny":     hour_ok = ind["h_london"]
        elif time_filt == "asia":        hour_ok = ind["h_asia"]
        elif sess_filt == "asia":        hour_ok = ind["h_asia"]
        elif sess_filt == "london":      hour_ok = ind["h_london"]
        elif sess_filt == "ny":          hour_ok = ind["h_ny"]
        elif sess_filt == "london_ny_overlap": hour_ok = ind["h_lno"]
        else:                            hour_ok = None

        highs   = ind["highs"]; lows  = ind["lows"]
        opens   = ind["opens"]; vols  = ind["volumes"]
        atr14   = ind["atr14"]; ema200 = ind["ema200"]
        rsi14   = ind["rsi14"]; vol_ma = ind["vol_ma"]
        roll_hi = ind["roll_hi"]; roll_lo = ind["roll_lo"]
        atr_pct_ref = ind["atr_pct_ref"]; price_mean = ind["price_mean"]
        atr_mean    = ind["atr_mean"];    atr_std    = ind["atr_std"]
        ts_col  = ind["ts"]
        N = len(closes)

        coin_equity = INITIAL_EQUITY
        position    = None  # (direction, entry, stop, tp)
        trades_today = 0; current_day = -1
        # Kelly state
        kel_wins = 0; kel_n = 0; kel_gross_w = 0.0; kel_gross_l = 0.0

        for i in range(min_bars, N - 1):
            price = closes[i]

            # Track day for max_trades_day
            if max_td > 0:
                day = int(ts_col[i] // 86_400_000)
                if day != current_day:
                    current_day = day; trades_today = 0

            # ── Gestión posición ──────────────────────────────────────────────
            if position is not None:
                direction, entry, stop, tp = position

                if trail_type == "atr_dynamic":
                    av = atr14[i] if atr14[i] == atr14[i] else entry * 0.01
                    if direction == "long":
                        ns = price - 2.0 * av
                        if ns > stop: stop = ns
                    else:
                        ns = price + 2.0 * av
                        if ns < stop: stop = ns
                    position = (direction, entry, stop, tp)
                elif trail_type == "chandelier":
                    w = 22
                    if i >= w:
                        if direction == "long":
                            peak = float(np.max(highs[i-w:i+1]))
                            av = atr14[i] if atr14[i] == atr14[i] else entry * 0.01
                            ns = peak - 3.0 * av
                            if ns > stop: stop = ns
                        else:
                            trough = float(np.min(lows[i-w:i+1]))
                            av = atr14[i] if atr14[i] == atr14[i] else entry * 0.01
                            ns = trough + 3.0 * av
                            if ns < stop: stop = ns
                    position = (direction, entry, stop, tp)
                elif sl_type == "trailing":
                    if direction == "long":
                        ns = price * (1.0 - trailing)
                        if ns > stop: stop = ns
                    else:
                        ns = price * (1.0 + trailing)
                        if ns < stop: stop = ns
                    position = (direction, entry, stop, tp)
                elif sl_type == "atr":
                    av = atr14[i]
                    if av == av:
                        if direction == "long":
                            ns = price - 2.0 * av
                            if ns > stop: stop = ns
                        else:
                            ns = price + 2.0 * av
                            if ns < stop: stop = ns
                        position = (direction, entry, stop, tp)

                exit_now = False
                if direction == "long":
                    exit_now = price <= stop or (tp is not None and price >= tp)
                else:
                    exit_now = price >= stop or (tp is not None and price <= tp)

                if not exit_now:
                    pf1 = maf[i-1]; ps1 = mas[i-1]; cf1 = maf[i]; cs1 = mas[i]
                    if pf1==pf1 and ps1==ps1 and cf1==cf1 and cs1==cs1:
                        if direction == "long"  and pf1 >= ps1 and cf1 < cs1: exit_now = True
                        elif direction == "short" and pf1 <= ps1 and cf1 > cs1: exit_now = True

                if exit_now:
                    # Position sizing on exit
                    if pos_siz == "kelly" and kel_n >= 5:
                        wr_k = kel_wins / kel_n
                        rr_k = (kel_gross_w / kel_wins) / (kel_gross_l / max(kel_n - kel_wins, 1)) if kel_wins > 0 and kel_n > kel_wins else 1.0
                        f_star = (wr_k * rr_k - (1.0 - wr_k)) / rr_k if rr_k > 0 else risk_lev
                        eff_risk = max(0.005, min(f_star, 0.25))
                    elif pos_siz == "atr_based":
                        av = atr14[i] if atr14[i] == atr14[i] and atr14[i] > 0 else entry * 0.01
                        eff_risk = (params.risk_pct * coin_equity) / (av * entry) if entry > 0 else risk_lev
                        eff_risk = min(eff_risk, 0.25)
                    elif pos_siz == "volatility_adjusted":
                        av = atr14[i] if atr14[i] == atr14[i] and atr14[i] > 0 else entry * 0.01
                        atp = av / price if price > 0 else atr_pct_ref
                        eff_risk = risk_lev * (atr_pct_ref / atp) if atp > 0 else risk_lev
                        eff_risk = min(eff_risk, 0.25)
                    else:
                        eff_risk = risk_lev

                    size = coin_equity * eff_risk / entry
                    pnl  = size * (price - entry) if direction == "long" else size * (entry - price)
                    if compound: coin_equity = max(coin_equity + pnl, 1.0)
                    # Kelly update
                    kel_n += 1
                    if pnl > 0: kel_wins += 1; kel_gross_w += pnl
                    else: kel_gross_l += -pnl
                    all_pnl    += pnl; all_trades += 1
                    if pnl > 0: all_wins += 1; gross_win  += pnl
                    else:       gross_loss += -pnl
                    if pnl > all_best:  all_best  = pnl
                    if pnl < all_worst: all_worst = pnl
                    equity += pnl
                    if equity > peak_equity: peak_equity = equity
                    elif peak_equity > 0:
                        dd = (peak_equity - equity) / peak_equity * 100.0
                        if dd > max_dd: max_dd = dd
                    eq_curve.append(equity)
                    trades_today += 1
                    position = None
                    continue

            # ── Búsqueda de entrada ───────────────────────────────────────────
            if max_td > 0 and trades_today >= max_td: continue

            pf1 = maf[i-1]; ps1 = mas[i-1]; cf1 = maf[i]; cs1 = mas[i]
            if not (pf1==pf1 and ps1==ps1 and cf1==cf1 and cs1==cs1): continue

            # Direction filter (--direction arg)
            if   DIRECTION == "long":
                if not (pf1 < ps1 and cf1 > cs1): continue
                signal = "long"
            elif DIRECTION == "short":
                if not (pf1 > ps1 and cf1 < cs1): continue
                signal = "short"
            else:
                if   pf1 < ps1 and cf1 > cs1: signal = "long"
                elif pf1 > ps1 and cf1 < cs1: signal = "short"
                else: continue

            # ── Session / hour filter ─────────────────────────────────────────
            if hour_ok is not None and not hour_ok[i]: continue

            # ── v5 hard filters ───────────────────────────────────────────────
            if e200_filt != "none":
                e200 = ema200[i]
                if e200 == e200 and e200 > 0:
                    if e200_filt == "soft":
                        if signal == "long"  and price < e200 * 0.98: continue
                        if signal == "short" and price > e200 * 1.02: continue
                    else:
                        if signal == "long"  and price <= e200: continue
                        if signal == "short" and price >= e200: continue

            if rsi_filt != "none":
                rv = rsi14[i]
                if rv == rv:
                    if rsi_filt == "rsi50":
                        if signal == "long"  and rv <= 50.0: continue
                        if signal == "short" and rv >= 50.0: continue
                    else:
                        if signal == "long"  and rv <= 55.0: continue
                        if signal == "short" and rv >= 45.0: continue

            if atr_filt != "none" and atr_pct_ref > 0:
                av = atr14[i]
                if av == av:
                    apt = av / price_mean
                    if atr_filt == "min" and apt < atr_pct_ref * 0.5: continue
                    if atr_filt == "max" and apt > atr_pct_ref * 1.8: continue

            if fib_mode != "disabled":
                hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
                if hi_sw == hi_sw and lo_sw == lo_sw:
                    if fib_mode == "required" and not _near_fibo_fast(price, hi_sw, lo_sw): continue

            if vol_prof != "disabled":
                vm = vol_ma[i]
                if vm == vm and vm > 0:
                    vn = vols[i]
                    if vol_prof == "strict"  and vn < vm * 1.5: continue
                    if vol_prof == "relaxed" and vn < vm * 0.7: continue

            # ── Static market regime filters (non-candle) ─────────────────────
            if fng_filt != "none" and fng_val is not None:
                if fng_filt == "extreme_fear":
                    if signal == "short": continue   # only long in extreme fear
                elif fng_filt == "extreme_greed":
                    if signal == "long":  continue   # only short in extreme greed
                elif fng_filt == "both":
                    if fng_val >= 25 and fng_val <= 75: continue  # no trades in neutral

            if fund_filt != "none" and fund_val is not None:
                abs_f = abs(fund_val)
                if fund_filt == "extreme" and abs_f <= 0.001: continue
                if fund_filt == "moderate" and abs_f <= 0.0003: continue
                # trade against the crowd: positive funding → longs paying → favor short
                if abs_f > 0:
                    if fund_val > 0 and signal == "long":  continue
                    if fund_val < 0 and signal == "short": continue

            # ── v6 confluence counting ────────────────────────────────────────
            confluences = 0

            # MACD
            if macd_filt != "none":
                ml = ind["macd_l"][i]; ms_ = ind["macd_s"][i]; mh = ind["macd_h"][i]
                if ml == ml and ms_ == ms_ and mh == mh:
                    if macd_filt == "histogram":
                        if (signal == "long"  and mh > 0) or (signal == "short" and mh < 0): confluences += 1
                    elif macd_filt == "signal_cross":
                        pml = ind["macd_l"][i-1]; pms = ind["macd_s"][i-1]
                        if pml == pml and pms == pms:
                            if signal == "long"  and pml <= pms and ml > ms_: confluences += 1
                            if signal == "short" and pml >= pms and ml < ms_: confluences += 1
                    elif macd_filt == "divergence":
                        if signal == "long"  and ind["macd_bull_div"][i]: confluences += 1
                        if signal == "short" and ind["macd_bear_div"][i]: confluences += 1

            # ADX
            if adx_filt != "none":
                av = ind["adx"][i]
                if av == av:
                    thr = float(adx_filt)
                    if av >= thr: confluences += 1

            # Supertrend
            if st_filt != "none":
                st_dir = ind["st"][i]
                if (signal == "long"  and st_dir > 0) or (signal == "short" and st_dir < 0): confluences += 1

            # Ichimoku
            if ichi_filt != "none":
                ct = ind["cloud_top"][i]; cb = ind["cloud_bot"][i]
                if ct == ct and cb == cb:
                    if ichi_filt == "above_cloud":
                        if signal == "long"  and price > ct: confluences += 1
                        if signal == "short" and price < cb: confluences += 1
                    elif ichi_filt == "tk_cross":
                        t = ind["tenkan"][i]; k_ = ind["kijun"][i]
                        pt = ind["tenkan"][i-1]; pk = ind["kijun"][i-1]
                        if t==t and k_==k_ and pt==pt and pk==pk:
                            if signal == "long"  and pt <= pk and t > k_: confluences += 1
                            if signal == "short" and pt >= pk and t < k_: confluences += 1

            # Stochastic RSI
            if stoch_filt != "none":
                sk_v = ind["sk"][i]; sd_v = ind["sd"][i]
                if sk_v == sk_v and sd_v == sd_v:
                    if stoch_filt == "oversold_ob":
                        if signal == "long"  and sk_v < 20: confluences += 1
                        if signal == "short" and sk_v > 80: confluences += 1
                    elif stoch_filt == "cross":
                        psk = ind["sk"][i-1]; psd = ind["sd"][i-1]
                        if psk == psk and psd == psd:
                            if signal == "long"  and psk <= psd and sk_v > sd_v: confluences += 1
                            if signal == "short" and psk >= psd and sk_v < sd_v: confluences += 1

            # CCI
            if cci_filt != "none":
                cv = ind["cci"][i]
                if cv == cv:
                    thr = float(cci_filt)
                    if signal == "long"  and cv > thr:   confluences += 1
                    if signal == "short" and cv < -thr:  confluences += 1

            # Williams %R
            if wr_filt != "none":
                wv = ind["wr"][i]
                if wv == wv:
                    if signal == "long"  and wv < -80: confluences += 1
                    if signal == "short" and wv > -20: confluences += 1

            # Momentum
            if mom_filt != "none":
                mv = ind["mom10"][i]
                if mv == mv:
                    if mom_filt == "positive":
                        if signal == "long"  and mv > 0: confluences += 1
                        if signal == "short" and mv < 0: confluences += 1
                    elif mom_filt == "accelerating":
                        pv = ind["mom10"][i-1]
                        if pv == pv:
                            if signal == "long"  and mv > 0 and mv > pv: confluences += 1
                            if signal == "short" and mv < 0 and mv < pv: confluences += 1

            # Bollinger Bands
            if bb_filt != "none":
                bu = ind["bb_u"][i]; bl_ = ind["bb_l"][i]
                bw = ind["bb_w"][i]; bw_s = ind["bb_w_sma"][i]
                if bu == bu and bl_ == bl_:
                    if bb_filt == "breakout":
                        if signal == "long"  and price > bu: confluences += 1
                        if signal == "short" and price < bl_: confluences += 1
                    elif bb_filt == "squeeze":
                        if bw == bw and bw_s == bw_s and bw < bw_s * 0.8: confluences += 1
                    elif bb_filt == "mean_reversion":
                        if signal == "long"  and price < bl_: confluences += 1
                        if signal == "short" and price > bu:  confluences += 1

            # ATR Volatility
            if atr_vol_f != "none" and atr_mean > 0:
                av = atr14[i]
                if av == av:
                    ratio = av / atr_mean
                    if atr_vol_f == "low"    and ratio < 0.7:  confluences += 1
                    if atr_vol_f == "medium" and 0.7 <= ratio < 1.5: confluences += 1
                    if atr_vol_f == "high"   and ratio >= 1.5: confluences += 1

            # Keltner
            if kelt_filt != "none":
                ku = ind["kelt_u"][i]; kl = ind["kelt_l"][i]
                if ku == ku and kl == kl:
                    bb_u_v = ind["bb_u"][i]; bb_l_v = ind["bb_l"][i]
                    if kelt_filt == "breakout":
                        if signal == "long"  and price > ku: confluences += 1
                        if signal == "short" and price < kl: confluences += 1
                    elif kelt_filt == "squeeze":
                        if bb_u_v == bb_u_v and bb_l_v == bb_l_v:
                            if bb_u_v < ku and bb_l_v > kl: confluences += 1

            # OBV
            if obv_filt != "none":
                ov = ind["obv"][i]; oe = ind["obv_ema"][i]
                if oe == oe:
                    if obv_filt == "trend":
                        if signal == "long"  and ov > oe: confluences += 1
                        if signal == "short" and ov < oe: confluences += 1
                    elif obv_filt == "divergence":
                        if signal == "long"  and ind["obv_bull_div"][i]: confluences += 1
                        if signal == "short" and ind["obv_bear_div"][i]: confluences += 1

            # VWAP
            if vwap_filt != "none":
                vv = ind["vwap"][i]
                if vv == vv:
                    if signal == "long"  and price > vv: confluences += 1
                    if signal == "short" and price < vv: confluences += 1

            # Volume Delta
            if vdelta_filt != "none":
                cv2 = ind["cvd"][i]; ce = ind["cvd_ema"][i]
                if ce == ce:
                    if vdelta_filt == "confirm":
                        if signal == "long"  and cv2 > ce: confluences += 1
                        if signal == "short" and cv2 < ce: confluences += 1
                    elif vdelta_filt == "divergence":
                        # Price up, CVD down = bear divergence
                        if i >= 1:
                            pc2 = closes[i-1]; pcvd = ind["cvd"][i-1]
                            if signal == "long"  and price < pc2 and cv2 > pcvd: confluences += 1
                            if signal == "short" and price > pc2 and cv2 < pcvd: confluences += 1

            # CVD filter
            if cvd_filt != "none":
                cv3 = ind["cvd"][i]; ce3 = ind["cvd_ema"][i]
                if ce3 == ce3:
                    if cvd_filt == "confirm":
                        if signal == "long"  and cv3 > 0: confluences += 1
                        if signal == "short" and cv3 < 0: confluences += 1
                    elif cvd_filt == "divergence":
                        # Negative CVD with bullish price → hidden bull
                        if signal == "long"  and cv3 < 0 and price > closes[i-1]: confluences += 1
                        if signal == "short" and cv3 > 0 and price < closes[i-1]: confluences += 1

            # Market structure
            if ms_filt != "none":
                ms_v = ind["ms"][i]
                if signal == "long"  and ms_v > 0: confluences += 1
                if signal == "short" and ms_v < 0: confluences += 1

            # Breakout range
            if bo_range != "none":
                n_bo = int(bo_range)
                rh_key = f"roll_hi{n_bo}"; rl_key = f"roll_lo{n_bo}"
                rh = ind.get(rh_key, roll_hi)[i]; rl = ind.get(rl_key, roll_lo)[i]
                if rh == rh and rl == rl:
                    if signal == "long"  and price > rh: confluences += 1
                    if signal == "short" and price < rl: confluences += 1

            # Candle pattern
            if cpat_filt != "none":
                op = opens[i]; hi = highs[i]; lo = lows[i]; cl = closes[i]
                body = abs(cl - op); total = hi - lo
                if cpat_filt == "hammer" and total > 0:
                    lower_wick = min(op, cl) - lo
                    if lower_wick > body * 2 and signal == "long": confluences += 1
                elif cpat_filt == "engulfing" and i >= 1:
                    pop = opens[i-1]; pcl = closes[i-1]
                    pbody = abs(pcl - pop)
                    if signal == "long"  and cl > op and cl > pop and op < pcl and body > pbody: confluences += 1
                    if signal == "short" and cl < op and cl < pop and op > pcl and body > pbody: confluences += 1
                elif cpat_filt == "doji" and total > 0:
                    if body / total < 0.1: confluences += 1

            # Order block
            if ob_filt != "none":
                if signal == "long"  and ind["ob_bull"][i]: confluences += 1
                if signal == "short" and ind["ob_bear"][i]: confluences += 1

            # Pivot filter
            if piv_filt != "none":
                if piv_filt == "daily":
                    s1 = ind["piv_s1"][i]; r1 = ind["piv_r1"][i]
                else:
                    s1 = ind["wpiv_s1"][i]; r1 = ind["wpiv_r1"][i]
                if s1 == s1 and r1 == r1:
                    if signal == "long"  and price > s1: confluences += 1
                    if signal == "short" and price < r1: confluences += 1

            # SR breakout
            if sr_bo_filt != "none":
                n_sr = int(sr_bo_filt)
                rh_k = f"roll_hi{n_sr}" if n_sr <= 100 else "roll_hi"
                rl_k = f"roll_lo{n_sr}" if n_sr <= 100 else "roll_lo"
                rh_v = ind.get(rh_k, roll_hi)
                rl_v = ind.get(rl_k, roll_lo)
                # Use i-1 (close above previous high)
                if i >= 1 and rh_v[i-1] == rh_v[i-1]:
                    if signal == "long"  and price > rh_v[i-1]: confluences += 1
                    if signal == "short" and price < rl_v[i-1]: confluences += 1

            # Fibonacci retracement
            if fib_ret_f != "none":
                hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
                if hi_sw == hi_sw and lo_sw == lo_sw and hi_sw > lo_sw:
                    rng = hi_sw - lo_sw
                    pct = float(fib_ret_f) / 100.0
                    level = lo_sw + rng * pct if signal == "long" else hi_sw - rng * pct
                    if abs(price - level) / price <= 0.015: confluences += 1

            # RSI divergence
            if rsi_div_f != "none":
                bull_d = ind["rsi_bull_div"][i]; bear_d = ind["rsi_bear_div"][i]
                if rsi_div_f == "required" or rsi_div_f == "bonus":
                    if signal == "long"  and bull_d: confluences += 1
                    if signal == "short" and bear_d: confluences += 1
                    if rsi_div_f == "required":
                        # Hard block if divergence not present
                        if signal == "long"  and not bull_d: continue
                        if signal == "short" and not bear_d: continue

            # BTC correlation
            if btc_corr != "none" and coin != "BTC":
                bt = ind.get("btc_trend")
                if bt is not None and i < len(bt):
                    if signal == "long"  and bt[i] > 0: confluences += 1
                    if signal == "short" and bt[i] < 0: confluences += 1
                    # Hard block if BTC going opposite
                    if signal == "long"  and bt[i] < 0: continue
                    if signal == "short" and bt[i] > 0: continue

            # ── Minimum confluences gate ──────────────────────────────────────
            if confluences < min_conf: continue

            # ── Entry ─────────────────────────────────────────────────────────
            entry  = price
            av_now = atr14[i] if atr14[i] == atr14[i] else entry * 0.01
            if sl_type == "fixed":
                sp   = trailing * 2.0
                stop = entry * (1.0 - sp) if signal == "long" else entry * (1.0 + sp)
            elif sl_type == "trailing":
                stop = entry * (1.0 - trailing) if signal == "long" else entry * (1.0 + trailing)
            else:
                stop = (entry - 2.0 * av_now) if signal == "long" else (entry + 2.0 * av_now)

            if trail_type == "chandelier":
                w = 22
                if i >= w:
                    av_c = av_now
                    if signal == "long":
                        stop = float(np.max(highs[i-w:i+1])) - 3.0 * av_c
                    else:
                        stop = float(np.min(lows[i-w:i+1]))  + 3.0 * av_c

            tp = None
            if tp_fib:
                hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
                if hi_sw == hi_sw and lo_sw == lo_sw:
                    rng_sw = hi_sw - lo_sw
                    if rng_sw > 0:
                        if signal == "long":
                            cands = [lo_sw + rng_sw * f for f in (0.618, 0.786, 1.0)
                                     if lo_sw + rng_sw * f > entry]
                            tp = min(cands) if cands else None
                        else:
                            cands = [lo_sw + rng_sw * f for f in (0.0, 0.236, 0.382)
                                     if lo_sw + rng_sw * f < entry]
                            tp = max(cands) if cands else None

            position = (signal, entry, stop, tp)

    # ── Métricas ──────────────────────────────────────────────────────────────
    if all_trades == 0:
        return {"total_pnl": 0.0, "total_pnl_pct": 0.0, "win_rate": 0.0,
                "total_trades": 0, "max_drawdown": 0.0, "sharpe": 0.0,
                "profit_factor": 0.0, "best_trade": 0.0, "worst_trade": 0.0}

    win_rate      = all_wins / all_trades * 100.0
    total_pnl_pct = all_pnl / INITIAL_EQUITY * 100.0

    if len(eq_curve) > 2:
        eq_arr  = np.array(eq_curve)
        returns = np.diff(eq_arr) / np.where(eq_arr[:-1] > 0, eq_arr[:-1], 1.0)
        std_r   = float(np.std(returns))
        sharpe  = float(np.mean(returns)) / std_r * np.sqrt(252) if std_r > 0 else 0.0
    else:
        sharpe = 0.0

    profit_factor = gross_win / gross_loss if gross_loss > 0 else (2.0 if gross_win > 0 else 1.0)

    return {
        "total_pnl":     round(all_pnl, 2),
        "total_pnl_pct": round(total_pnl_pct, 2),
        "win_rate":      round(win_rate, 2),
        "total_trades":  all_trades,
        "max_drawdown":  round(max_dd, 2),
        "sharpe":        round(sharpe, 3),
        "profit_factor": round(profit_factor, 3),
        "best_trade":    round(all_best, 2),
        "worst_trade":   round(all_worst, 2),
    }


# ─── WORKER FUNCTIONS ─────────────────────────────────────────────────────────
def _worker_run_segment(task: dict) -> List[dict]:
    n = task["n"]; seed = task["seed"]
    rng = random.Random(seed); results: List[dict] = []
    for _ in range(n):
        params = random_params(rng)
        coins_ind = _worker_ind.get(params.interval)
        if not coins_ind: continue
        metrics = _simulate_fast(params, coins_ind)
        if metrics["total_trades"] == 0: continue
        r = asdict(params); r.update(metrics); results.append(r)
    return results


def _worker_ping(_) -> bool: return True


def _worker_run_hc_segment(task: dict) -> List[dict]:
    results: List[dict] = []
    for pd in task.get("params", []):
        try:
            params = OptParams(**pd)
            coins_ind = _worker_ind.get(params.interval)
            if not coins_ind: continue
            metrics = _simulate_fast(params, coins_ind)
            if metrics["total_trades"] == 0: continue
            r = asdict(params); r.update(metrics); results.append(r)
        except Exception:
            continue
    return results


# ─── CHECKPOINT ───────────────────────────────────────────────────────────────
def _score(r: dict) -> float:
    pnl_pct  = r.get("total_pnl_pct", 0.0)
    win_rate = r.get("win_rate", 0.0)
    max_dd   = max(r.get("max_drawdown", 1.0), 1.0)
    return (pnl_pct * win_rate) / max_dd


def save_checkpoint(path: Path, processed: int,
                    results: List[dict], tf_results: Dict[str, List[dict]]):
    data = {
        "processed":  processed,
        "results":    sorted(results, key=_score, reverse=True)[:1000],
        "tf_results": {tf: sorted(r, key=_score, reverse=True)[:200]
                       for tf, r in tf_results.items()},
        "timestamp":  datetime.now().isoformat(),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f)


def load_checkpoint(path: Path) -> Optional[dict]:
    if not path.exists(): return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def deduplicate(results: List[dict], top_n: int = 30) -> List[dict]:
    seen = set(); unique = []
    for r in sorted(results, key=_score, reverse=True):
        p   = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
        key = hashlib.md5(json.dumps(p, sort_keys=True).encode()).hexdigest()
        if key not in seen:
            seen.add(key); unique.append(r)
            if len(unique) >= top_n: break
    return unique


def _add(r: dict, all_results: list, seen_keys: set, tf_results: dict):
    if not r or r.get("total_trades", 0) == 0: return
    p   = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
    key = hashlib.md5(json.dumps(p, sort_keys=True).encode()).hexdigest()
    if key in seen_keys: return
    seen_keys.add(key); all_results.append(r)
    tf = r.get("interval")
    if tf in tf_results: tf_results[tf].append(r)


# ─── JSON ─────────────────────────────────────────────────────────────────────
def save_json(out_dir: Path, global_top: List[dict], tf_tops: Dict[str, List[dict]],
              generated_at: str, total_tried: int):
    data = {
        "version": VERSION, "generated_at": generated_at, "total_tried": total_tried,
        "results": global_top,
        "by_timeframe": {tf: tf_tops.get(tf, []) for tf in TIMEFRAMES},
    }
    out_path = out_dir / f"top_params_{COIN_ID}_v6.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"[{ts()}] JSON: {out_path}")


# ─── EXCEL ────────────────────────────────────────────────────────────────────
def save_excel(out_dir: Path, global_top: List[dict], tf_tops: Dict[str, List[dict]],
               all_results: List[dict]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"[{ts()}] openpyxl no disponible — saltando Excel"); return

    PARAM_KEYS = list(_PARAM_FIELDS_SET) + [
        "total_pnl", "total_pnl_pct", "win_rate", "total_trades",
        "max_drawdown", "sharpe", "profit_factor", "best_trade", "worst_trade",
    ]
    HDR_COLS = ["Rank"] + PARAM_KEYS

    H_FILL = PatternFill("solid", fgColor="0D1F2D")
    H_FONT = Font(color="4FC3F7", bold=True, size=8)
    A_FILL = PatternFill("solid", fgColor="090F1A")

    def _fv(k, v):
        if k in ("trailing_pct", "risk_pct"): return f"{v*100:.1f}%"
        if isinstance(v, bool): return "Si" if v else "No"
        if isinstance(v, float): return round(v, 3)
        return v

    def write_sheet(ws, rows, title):
        ws.title = title[:31]
        ws.append(HDR_COLS)
        for cell in ws[1]:
            cell.font = H_FONT; cell.fill = H_FILL
            cell.alignment = Alignment(horizontal="center")
        for i, r in enumerate(rows):
            row_data = [i+1] + [_fv(k, r.get(k, "")) for k in PARAM_KEYS]
            ws.append(row_data)
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = A_FILL
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+2, 24)

    wb = Workbook()
    write_sheet(wb.active, global_top, "Top30_Global")
    for tf in TIMEFRAMES:
        write_sheet(wb.create_sheet(), tf_tops.get(tf, []), f"Top30_{tf}")

    sorted_all = sorted(all_results, key=_score, reverse=True)
    write_sheet(wb.create_sheet(), sorted_all[:10_000], "Todas_Combinaciones")

    out_path = out_dir / f"resultados_optimizacion_{COIN_ID}_v6.xlsx"
    try:
        wb.save(out_path)
        print(f"[{ts()}] Excel: {out_path}")
    except Exception as e:
        print(f"[{ts()}] Error Excel: {e}")


# ─── PDF ──────────────────────────────────────────────────────────────────────
def save_pdf(out_dir: Path, global_top: List[dict], tf_tops: Dict[str, List[dict]],
             total_tried: int, duration_sec: float):
    try:
        import matplotlib; matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except ImportError:
        print(f"[{ts()}] matplotlib no disponible — saltando PDF"); return

    BG = "#07090f"; FG = "#c9d4e0"; ACC = "#4fc3f7"
    plt.rcParams.update({
        "figure.facecolor": BG, "axes.facecolor": BG, "axes.edgecolor": "#162030",
        "text.color": FG, "axes.labelcolor": FG, "xtick.color": FG, "ytick.color": FG,
        "grid.color": "#0d1520", "font.family": "monospace", "font.size": 9,
    })

    out_path = out_dir / f"informe_optimizacion_{COIN_ID}_v6.pdf"
    try:
        with PdfPages(out_path) as pdf:
            fig, ax = plt.subplots(figsize=(11.69, 8.27))
            ax.axis("off")
            ax.text(0.5, 0.88, "AlphaChainBots", ha="center", fontsize=30, color=ACC, weight="bold")
            ax.text(0.5, 0.78, f"Informe de Optimizacion v6 — {COIN_ID.upper()} {DIRECTION.upper()}", ha="center", fontsize=16, color=FG)
            ax.text(0.5, 0.68, f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ha="center", fontsize=11, color="#78909c")
            ax.text(0.5, 0.60, f"Combinaciones: {total_tried:,}  |  Duracion: {duration_sec/60:.1f} min", ha="center", fontsize=11, color=FG)
            ax.text(0.5, 0.52, f"Monedas: {', '.join(COINS)}  |  TFs: {', '.join(TIMEFRAMES)}", ha="center", fontsize=9, color="#78909c")
            if global_top:
                best = global_top[0]
                ax.text(0.5, 0.40, "Mejor combinacion global:", ha="center", fontsize=11, color=ACC)
                summary = (f"TF:{best.get('interval','?')} "
                           f"{best.get('ma_type','').upper()} {best.get('ma_fast','?')}/{best.get('ma_slow','?')} "
                           f"Lev:{best.get('leverage','?')}x "
                           f"PnL:{best.get('total_pnl',0):+.0f}$ ({best.get('total_pnl_pct',0):+.1f}%) "
                           f"WR:{best.get('win_rate',0):.1f}% Sharpe:{best.get('sharpe',0):.2f}")
                ax.text(0.5, 0.30, summary, ha="center", fontsize=9, color=FG, family="monospace")
            pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()

            # Resumen por TF
            fig, axes = plt.subplots(2, 3, figsize=(11.69, 8.27))
            fig.suptitle("Resumen por Timeframe v6", fontsize=13, color=ACC)
            tf_names  = TIMEFRAMES
            best_pnls = [tf_tops[tf][0].get("total_pnl", 0) if tf_tops.get(tf) else 0 for tf in tf_names]
            axes[0,0].bar(tf_names, best_pnls,
                          color=["#00e676" if p >= 0 else "#ff4466" for p in best_pnls], alpha=0.85)
            axes[0,0].set_title("Mejor PnL por TF", color=ACC, fontsize=9)
            pnls = [r.get("total_pnl",0) for r in global_top]
            axes[0,1].hist(pnls, bins=min(10, len(pnls)), color=ACC, alpha=0.7)
            axes[0,1].set_title("Dist PnL Top30", color=ACC, fontsize=9)
            wrs = [r.get("win_rate",0) for r in global_top[:10]]
            axes[0,2].barh([f"#{i+1}" for i in range(len(wrs))], wrs, color="#4fc3f7", alpha=0.8)
            axes[0,2].set_title("WR% Top10", color=ACC, fontsize=9); axes[0,2].set_xlim(0,100)
            # Most common new filters in top 30
            for ax_idx, fld in enumerate(["macd_filter", "adx_filter", "supertrend_filter"]):
                cnts: Dict[str, int] = {}
                for r in global_top: cnts[r.get(fld,"?")] = cnts.get(r.get(fld,"?"),0)+1
                if cnts:
                    ax_ = axes[1, ax_idx]
                    ax_.bar(list(cnts.keys()), list(cnts.values()), color="#ce93d8", alpha=0.8)
                    ax_.set_title(fld, color=ACC, fontsize=8)
                    ax_.tick_params(labelsize=6)
                    plt.setp(ax_.get_xticklabels(), rotation=30, ha="right")
            plt.tight_layout()
            pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()

        print(f"[{ts()}] PDF: {out_path}")
    except Exception as e:
        print(f"[{ts()}] Error PDF: {e}"); traceback.print_exc()


# ─── PRINT SUMMARY ────────────────────────────────────────────────────────────
def print_summary(global_top: List[dict], tf_tops: Dict[str, List[dict]],
                  total_tried: int, duration_sec: float, out_dir: Path):
    sep = "=" * 76
    print(f"\n{sep}")
    print(f"[{ts()}]  OPTIMIZACION v6 COMPLETADA — {COIN_ID.upper()} {DIRECTION.upper()}")
    print(sep)
    print(f"  Combinaciones: {total_tried:,}  |  Duracion: {duration_sec/60:.1f} min")
    print(f"\n  TOP 5 GLOBAL:")
    print(f"  {'#':>2}  {'TF':>4}  {'MA':>12}  {'Lev':>4}  {'Trail':>6}  "
          f"{'PnL$':>10}  {'WR%':>5}  {'DD%':>5}  {'Sharpe':>7}  {'Conf':>4}")
    print(f"  {'-'*76}")
    for i, r in enumerate(global_top[:5]):
        ma = f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
        print(f"  {i+1:>2}  {r.get('interval','?'):>4}  {ma:>12}  "
              f"{r.get('leverage','?'):>3}x  "
              f"{r.get('trailing_pct',0)*100:>5.1f}%  "
              f"{r.get('total_pnl',0):>+10.2f}$  "
              f"{r.get('win_rate',0):>4.1f}%  "
              f"{r.get('max_drawdown',0):>4.1f}%  "
              f"{r.get('sharpe',0):>7.3f}  "
              f"{r.get('min_confluences',0):>4}")
    print(f"\n  MEJOR POR TIMEFRAME:")
    for tf in TIMEFRAMES:
        tops = tf_tops.get(tf, [])
        if tops:
            r = tops[0]
            ma = f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
            print(f"  {tf:>4}  {ma:>12}  {r.get('leverage','?'):>3}x  "
                  f"PnL:{r.get('total_pnl',0):>+10.2f}$  "
                  f"WR:{r.get('win_rate',0):>5.1f}%  "
                  f"Sharpe:{r.get('sharpe',0):>7.3f}")
    print(f"\n  Salida: {out_dir.resolve()}")
    print(f"{sep}\n")


# ─── EXTERNAL DATA ────────────────────────────────────────────────────────────
def _fetch_fear_greed() -> Optional[float]:
    try:
        r = requests.get(FNG_URL, timeout=10)
        data = r.json()
        return float(data["data"][0]["value"])
    except Exception as e:
        print(f"[{ts()}] Fear&Greed fetch failed: {e}")
        return None


def _fetch_funding_rate() -> Optional[float]:
    try:
        r = requests.get(BINANCE_FUND, params={"symbol": "BTCUSDT", "limit": 1}, timeout=10)
        data = r.json()
        if isinstance(data, list) and data:
            return float(data[-1].get("fundingRate", 0))
    except Exception as e:
        print(f"[{ts()}] Funding rate fetch failed: {e}")
    return None


# ─── DOWNLOAD TASK ────────────────────────────────────────────────────────────
def _dl_task(coin: str, tf: str, cache_dir: Path, days: int):
    arr = load_or_fetch(cache_dir, coin, tf, days)
    return coin, tf, arr


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    global COINS, COIN_ID, DIRECTION

    parser = argparse.ArgumentParser(description="AlphaChainBots Optimizer v6")
    parser.add_argument("--coin",      type=str, default="BTC",   help="Moneda (BTC, ETH, SOL…)")
    parser.add_argument("--direction", type=str, default="both",  help="long / short / both")
    parser.add_argument("--days",      type=int, default=365,     help="Dias de historial")
    parser.add_argument("--samples",   type=int, default=500_000, help="Combinaciones a probar")
    parser.add_argument("--out",       type=str, default=None,    help="Directorio de salida")
    parser.add_argument("--workers",   type=int, default=None,    help="Workers paralelos")
    parser.add_argument("--resume",    action="store_true",       help="Reanudar desde checkpoint")
    args = parser.parse_args()

    coin_id   = args.coin.upper().strip()
    direction = args.direction.lower().strip()
    if direction not in ("long", "short", "both"):
        print("--direction debe ser long, short o both"); sys.exit(1)

    COINS     = [coin_id]
    COIN_ID   = coin_id.lower()
    DIRECTION = direction
    cl        = f"{COIN_ID}_{direction}" if direction != "both" else COIN_ID

    out_default = f"./resultados_{cl}_v6"
    out_dir   = Path(args.out if args.out else out_default)
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = out_dir / f"cache_{cl}"
    cache_dir.mkdir(parents=True, exist_ok=True)
    ckpt_path = out_dir / f"checkpoint_{cl}_v6.json"
    n_workers = args.workers or os.cpu_count() or 4

    sep = "=" * 76
    print(f"\n{sep}")
    print(f"[{ts()}]  AlphaChainBots Optimizer v6 — {VERSION}")
    print(sep)
    print(f"  Moneda       : {coin_id}  |  Direccion: {direction.upper()}")
    print(f"  Historial    : {args.days}d  |  Muestras: {args.samples:,}  |  Workers: {n_workers}")
    print(f"  Salida       : {out_dir.resolve()}")
    print(f"{sep}\n")

    # ── Fetch external data ───────────────────────────────────────────────────
    print(f"[{ts()}] Obteniendo Fear & Greed Index…", end=" ", flush=True)
    fng = _fetch_fear_greed()
    print(f"{fng:.0f}" if fng is not None else "N/A")

    print(f"[{ts()}] Obteniendo Funding Rate BTC…", end=" ", flush=True)
    funding = _fetch_funding_rate()
    print(f"{funding:.5f}" if funding is not None else "N/A")

    # ── Resume checkpoint ─────────────────────────────────────────────────────
    processed_start = 0
    all_results: List[dict] = []
    tf_results:  Dict[str, List[dict]] = {tf: [] for tf in TIMEFRAMES}
    seen_keys:   set = set()

    if args.resume:
        ckpt = load_checkpoint(ckpt_path)
        if ckpt:
            processed_start = ckpt.get("processed", 0)
            all_results     = ckpt.get("results", [])
            for tf in TIMEFRAMES:
                tf_results[tf] = ckpt.get("tf_results", {}).get(tf, [])
            for r in all_results:
                p = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
                seen_keys.add(hashlib.md5(json.dumps(p, sort_keys=True).encode()).hexdigest())
            print(f"[{ts()}] Reanudando: {processed_start:,} combinaciones ya procesadas\n")
        else:
            print(f"[{ts()}] No se encontro checkpoint — iniciando desde cero\n")

    # ── Download candles ──────────────────────────────────────────────────────
    all_coins_needed = list(set([coin_id, "BTC"]))  # always include BTC for correlation
    total_tasks = len(all_coins_needed) * len(TIMEFRAMES)
    print(f"[{ts()}] Descargando velas: {len(all_coins_needed)} monedas x {len(TIMEFRAMES)} TF…")
    all_candles: Dict[str, Dict[str, np.ndarray]] = {tf: {} for tf in TIMEFRAMES}
    btc_candles: Dict[str, np.ndarray] = {}
    dl_ok = 0; dl_fail = 0

    with ThreadPoolExecutor(max_workers=min(total_tasks, 20)) as ex:
        futures = {ex.submit(_dl_task, coin, tf, cache_dir, args.days): (coin, tf)
                   for tf in TIMEFRAMES for coin in all_coins_needed}
        for fut in as_completed(futures):
            coin_f, tf_key = futures[fut]
            try:
                _, _, arr = fut.result()
            except Exception as e:
                print(f"  [ERR] {coin_f}/{tf_key}: {e}"); dl_fail += 1; continue
            if arr is not None and len(arr) > 50:
                if coin_f == coin_id:
                    all_candles[tf_key][coin_f] = arr
                if coin_f == "BTC":
                    btc_candles[tf_key] = arr
                print(f"  [OK] {coin_f:6s}/{tf_key:4s}  {len(arr):,} velas")
                dl_ok += 1
            else:
                print(f"  [--] {coin_f:6s}/{tf_key:4s}  sin datos"); dl_fail += 1

    print(f"\n[{ts()}] Descarga: {dl_ok} OK / {dl_fail} fallidos\n")

    # ── Pack worker cache ─────────────────────────────────────────────────────
    worker_pkl = out_dir / f"_wcache_{cl}.pkl"
    with open(worker_pkl, "wb") as f:
        pickle.dump({"candles": all_candles, "btc_candles": btc_candles,
                     "_fng": fng, "_funding": funding,
                     "_direction": direction, "_coin_id": COIN_ID,
                     "_coins": COINS}, f)

    # ── Test pool ─────────────────────────────────────────────────────────────
    print(f"[{ts()}] Iniciando ProcessPoolExecutor ({n_workers} workers)…", end=" ", flush=True)
    with ProcessPoolExecutor(max_workers=n_workers, initializer=_worker_init,
                             initargs=(str(worker_pkl),)) as _tp:
        list(_tp.map(_worker_ping, range(n_workers)))
    print(f"listos\n")

    # ── Random search ─────────────────────────────────────────────────────────
    rng        = random.Random(42)
    start_time = time.time()
    processed  = processed_start
    SEG_SIZE   = 2_000   # slightly smaller due to richer params
    last_top_print = processed_start; last_ckpt = processed_start
    initargs = (str(worker_pkl),)

    print(f"[{ts()}] Random search: {args.samples - processed_start:,} combinaciones…")

    with ProcessPoolExecutor(max_workers=n_workers, initializer=_worker_init,
                             initargs=initargs) as pool:
        while processed < args.samples:
            remaining  = args.samples - processed
            n_tasks    = max(1, min(n_workers * 2, remaining // max(SEG_SIZE // 2, 1)))
            actual_seg = max(1, min(SEG_SIZE, remaining // max(n_tasks, 1)))
            tasks = [{"n": actual_seg, "seed": rng.randint(0, 2**32)}
                     for _ in range(n_tasks)]
            try:
                batch = list(pool.map(_worker_run_segment, tasks))
            except Exception as e:
                print(f"\n[{ts()}] Error pool.map: {e} — continuando…")
                processed += n_tasks * actual_seg; continue
            for seg in batch:
                for r in seg:
                    _add(r, all_results, seen_keys, tf_results)
            processed += sum(t["n"] for t in tasks)

            elapsed  = max(time.time() - start_time, 0.001)
            done_now = processed - processed_start
            speed    = done_now / elapsed
            eta_sec  = (args.samples - processed) / speed if speed > 0 else 0
            eta_str  = (f"{eta_sec/3600:.1f}h" if eta_sec >= 3600
                        else f"{eta_sec/60:.1f}min" if eta_sec >= 60
                        else f"{eta_sec:.0f}s")
            pct = processed / args.samples * 100.0
            print(f"\r[{ts()}] {processed:>9,}/{args.samples:,} ({pct:5.1f}%) "
                  f"| {speed:>7,.0f}/s | ETA:{eta_str:>8} | unicos:{len(all_results):,}",
                  end="", flush=True)

            if processed - last_top_print >= PRINT_TOP_EVERY:
                last_top_print = processed
                top3 = sorted(all_results, key=_score, reverse=True)[:3]
                print(f"\n[{ts()}] Top 3:")
                for i, r in enumerate(top3):
                    ma = f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
                    print(f"  #{i+1}  {r.get('interval','?'):>4}  {ma:>14}  "
                          f"Lev:{r.get('leverage','?'):>2}x  "
                          f"PnL:{r.get('total_pnl',0):>+10.2f}$  "
                          f"WR:{r.get('win_rate',0):>5.1f}%  "
                          f"Sharpe:{r.get('sharpe',0):>7.3f}  "
                          f"Conf:{r.get('min_confluences',0)}")
                print()

            if processed - last_ckpt >= CHECKPOINT_EVERY:
                last_ckpt = processed
                save_checkpoint(ckpt_path, processed, all_results, tf_results)
                print(f"\n[{ts()}] Checkpoint ({processed:,})\n")

    print(f"\n\n[{ts()}] Random search completado.")

    # ── Hill climbing ─────────────────────────────────────────────────────────
    if all_results:
        sorted_all = sorted(all_results, key=_score, reverse=True)
        n_seeds    = max(1, len(sorted_all) // 5)
        n_hc       = min(n_seeds * 8, 40_000)
        print(f"[{ts()}] Hill climbing: {n_seeds} semillas x 8 = {n_hc:,} intentos…")
        hc_batch: List[dict] = []
        for seed in sorted_all[:n_hc // 8]:
            sp = OptParams(**{k: seed[k] for k in _PARAM_FIELDS_SET})
            for _ in range(8): hc_batch.append(asdict(perturb(sp, rng)))
        chunk_sz = max(1, len(hc_batch) // n_workers)
        hc_tasks = [{"params": hc_batch[i:i+chunk_sz]}
                    for i in range(0, len(hc_batch), chunk_sz)]
        with ProcessPoolExecutor(max_workers=n_workers, initializer=_worker_init,
                                 initargs=initargs) as pool:
            try:
                for seg in pool.map(_worker_run_hc_segment, hc_tasks):
                    for r in seg: _add(r, all_results, seen_keys, tf_results)
                processed += len(hc_batch)
            except Exception as e:
                print(f"\n[{ts()}] Error HC: {e}")
        print(f"[{ts()}] Hill climbing completado.")

    # ── Rankings ──────────────────────────────────────────────────────────────
    print(f"\n[{ts()}] Construyendo rankings…")
    global_top = deduplicate(all_results, 30)
    tf_tops    = {tf: deduplicate(tf_results.get(tf, []), 30) for tf in TIMEFRAMES}
    duration   = time.time() - start_time
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    save_json(out_dir, global_top, tf_tops, generated_at, processed)
    save_excel(out_dir, global_top, tf_tops, all_results)
    save_pdf(out_dir, global_top, tf_tops, processed, duration)
    save_checkpoint(ckpt_path, processed, all_results, tf_results)

    try: worker_pkl.unlink()
    except Exception: pass

    print_summary(global_top, tf_tops, processed, duration, out_dir)


if __name__ == "__main__":
    mp.freeze_support()
    main()
