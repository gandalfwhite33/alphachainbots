#!/usr/bin/env python3
"""
optimizer_master.py — Optimizador maestro multi-moneda y multi-dirección.
Prueba 10 monedas × 3 direcciones en un solo run de 6M muestras.
Todos los parámetros de entrada y salida de v7 conservados.

Genera:
  master_results.xlsx  — tabs por moneda + TOP_PNL + TOP_CONSISTENCIA + RECOMENDADOS_PITCH + RECOMENDADOS_LIVE
  master_bots.json     — top 30 listos para importar a sim_engine.py

Uso:
    python optimizer_master.py [--samples 6000000] [--out ../resultados_master]
                               [--workers N] [--resume] [--focus-coins ETH,BTC]
"""

import os, sys, json, time, random, math, argparse, hashlib, pickle
import traceback, multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Optional, List, Dict, Tuple
from pathlib import Path

import numpy as np
import requests

# ─── VERSIÓN Y CONSTANTES ─────────────────────────────────────────────────────
VERSION        = "master_4.2.0"
INITIAL_EQUITY = 10_000.0
HL_URL         = "https://api.hyperliquid.xyz/info"
BINANCE_FUND   = "https://fapi.binance.com/fapi/v1/fundingRate"
FNG_URL        = "https://api.alternative.me/fng/?limit=1"

MASTER_COINS      = ["BTC","ETH","SOL","DOGE","WIF","SUI","ARB","AVAX","OP","BNB"]
ACTIVE_COINS      = list(MASTER_COINS)  # puede reducirse con --focus-coins
MASTER_DIRECTIONS = ["long","short","both"]
TIMEFRAMES        = ["15m","30m","1h","2h","4h"]

CHECKPOINT_EVERY = 25_000
PRINT_TOP_EVERY  = 30_000
MIN_TRADES_DISPLAY   = 5   # para Excel sheets y TOP_PNL (permite SMA / 4h)
MIN_TRADES_DASHBOARD = 15  # para master_bots.json y RECOMENDADOS_LIVE

# ─── OPCIONES DE PARÁMETROS (v5/v6/v7 — sin cambios) ─────────────────────────
MA_PAIRS      = [("ema",8,21),("ema",13,34),("ema",21,55),("ema",20,50),
                 ("sma",50,100),("sma",100,200)]
LEVERAGES     = [2, 3, 5, 10, 15]
TRAILING_PCTS = [0.003,0.005,0.008,0.010,0.015,0.020,0.030]
SL_TYPES      = ["fixed","trailing","atr"]
FIB_MODES     = ["required","optional","disabled"]
RSI_FILTERS   = ["none","rsi50","rsi55"]
EMA200_FILT   = ["none","soft","strict"]
ATR_FILTERS   = ["none","min","max"]
COMPOUNDS     = [True, False]
TIME_FILTERS  = ["none","london_ny","asia"]
VOL_PROFILES  = ["disabled","strict","relaxed"]
LIQ_CONFIRMS  = [True, False]
RISK_PCTS     = [0.01,0.02,0.03,0.05]

MACD_FILTERS     = ["none","signal_cross","histogram","divergence"]
ADX_FILTERS      = ["none","20","25","30"]
ST_FILTERS       = ["none","required"]
ICHI_FILTERS     = ["none","above_cloud","tk_cross"]
STOCH_RSI        = ["none","oversold_ob","cross"]
CCI_FILTERS      = ["none","100","200"]
WILLIAMS_R       = ["none","required"]
MOM_FILTERS      = ["none","positive","accelerating"]
BB_FILTERS       = ["none","breakout","squeeze","mean_reversion"]
ATR_VOL          = ["none","low","medium","high"]
KELT_FILTERS     = ["none","breakout","squeeze"]
OBV_FILTERS      = ["none","trend","divergence"]
VWAP_FILTERS     = ["none","required"]
VOL_DELTA        = ["none","confirm","divergence"]
CVD_FILTERS      = ["none","confirm","divergence"]
MKT_STRUCT       = ["none","required"]
BREAKOUT_RANGE   = ["none","10","20","50"]
CANDLE_PAT       = ["none","hammer","engulfing","doji"]
ORDER_BLOCK      = ["none","required"]
PIVOT_FILTERS    = ["none","daily","weekly"]
SR_BREAKOUT      = ["none","20","50","100"]
FIB_RET          = ["none","38","50","61"]
RSI_DIV          = ["none","required","bonus"]
BTC_CORR         = ["none","required"]
FUNDING_FILT     = ["none","extreme","moderate"]
FNG_FILTERS      = ["none","extreme_fear","extreme_greed","both"]
SESSION_FILT     = ["none","asia","london","ny","london_ny_overlap","all_sessions"]
POS_SIZING       = ["fixed","kelly","atr_based","volatility_adjusted"]
MAX_TRADES_DAY   = [0, 1, 2, 5, 10]
MIN_CONF         = [0, 1, 2, 3]
TRAIL_TYPES      = ["fixed","atr_dynamic","chandelier","parabolic_sar","ma_trailing"]
TP_TYPES         = ["none","fixed_pct","atr_multiple","bb_opposite","vwap_target",
                    "ma_cross_opposite","ichimoku_cloud","fib_127","fib_161","fib_261",
                    "pivot_r1","pivot_r2","pivot_s1","pivot_s2","sr_level"]
TP_PCT_OPTS      = [2, 5, 10, 15, 20, 30]
TP_ATR_OPTS      = [1.5, 2.0, 3.0, 5.0]
TRAIL_ACT        = ["none","1pct","2pct","5pct","10pct"]
TRAIL_PROG       = [True, False]
ATR_TP_ADJ       = ["none","reduce_high","extend_low","both"]
PARTIAL_CLOSE    = ["none","2level","3level"]
PARTIAL_TRIG     = ["1atr","2atr","5pct","10pct","fib_127","fib_161"]
BREAKEVEN_OPTS   = ["none","1atr","2atr","5pct","10pct"]
TIME_EXIT_OPTS   = ["none","5","10","20","50"]
SESSION_EXIT     = [True, False]
WEEKEND_EXIT     = [True, False]
RR_MIN_OPTS      = ["none","1.5","2.0","3.0"]

_FEAR_GREED_IDX: Optional[float] = None
_FUNDING_RATE:   Optional[float] = None

_PARAM_FIELDS_SET = {
    "coin","direction",
    "interval","ma_type","ma_fast","ma_slow","leverage","trailing_pct",
    "sl_type","fib_mode","rsi_filter","ema200_filter","atr_filter",
    "compound","time_filter","vol_profile","liq_confirm","risk_pct",
    "macd_filter","adx_filter","supertrend_filter","ichimoku_filter",
    "stoch_rsi","cci_filter","williams_r","momentum_filter",
    "bb_filter","atr_volatility","keltner_filter",
    "obv_filter","vwap_filter","volume_delta","cvd_filter",
    "market_structure","breakout_range","candle_pattern","order_block",
    "pivot_filter","sr_breakout","fib_retracement","rsi_divergence",
    "btc_correlation","funding_filter","fear_greed_filter","session_filter",
    "position_sizing","max_trades_day","trailing_type","min_confluences",
    "tp_type","tp_pct","tp_atr","trailing_activation","trailing_progressive",
    "atr_tp_adjust","partial_close","partial_trigger","breakeven",
    "time_exit","session_exit","weekend_exit","rr_min",
}

# Filter fields used to build "Filtros activos" summary
_FILTER_FIELDS = [
    "macd_filter","adx_filter","supertrend_filter","ichimoku_filter",
    "stoch_rsi","cci_filter","williams_r","momentum_filter","bb_filter",
    "atr_volatility","keltner_filter","obv_filter","vwap_filter",
    "volume_delta","cvd_filter","market_structure","breakout_range",
    "candle_pattern","order_block","pivot_filter","sr_breakout",
    "fib_retracement","rsi_divergence","btc_correlation","funding_filter",
    "fear_greed_filter","session_filter",
]


def ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


# ─── DATACLASS ────────────────────────────────────────────────────────────────
@dataclass
class OptParams:
    interval:      str
    ma_type:       str
    ma_fast:       int
    ma_slow:       int
    leverage:      float
    trailing_pct:  float
    sl_type:       str
    fib_mode:      str
    rsi_filter:    str
    ema200_filter: str
    atr_filter:    str
    compound:      bool
    time_filter:   str
    vol_profile:   str
    liq_confirm:   bool
    risk_pct:      float
    macd_filter:        str
    adx_filter:         str
    supertrend_filter:  str
    ichimoku_filter:    str
    stoch_rsi:          str
    cci_filter:         str
    williams_r:         str
    momentum_filter:    str
    bb_filter:          str
    atr_volatility:     str
    keltner_filter:     str
    obv_filter:         str
    vwap_filter:        str
    volume_delta:       str
    cvd_filter:         str
    market_structure:   str
    breakout_range:     str
    candle_pattern:     str
    order_block:        str
    pivot_filter:       str
    sr_breakout:        str
    fib_retracement:    str
    rsi_divergence:     str
    btc_correlation:    str
    funding_filter:     str
    fear_greed_filter:  str
    session_filter:     str
    position_sizing:    str
    max_trades_day:     int
    trailing_type:      str
    min_confluences:    int
    tp_type:              str
    tp_pct:               int
    tp_atr:               float
    trailing_activation:  str
    trailing_progressive: bool
    atr_tp_adjust:        str
    partial_close:        str
    partial_trigger:      str
    breakeven:            str
    time_exit:            str
    session_exit:         bool
    weekend_exit:         bool
    rr_min:               str
    # master — coin y direction forman parte del espacio de búsqueda
    coin:      str = "BTC"
    direction: str = "both"

    def param_key(self) -> str:
        d = {k: v for k, v in asdict(self).items()}
        return hashlib.md5(json.dumps(d, sort_keys=True).encode()).hexdigest()


def random_params(rng: random.Random) -> OptParams:
    ma = rng.choices(MA_PAIRS, weights=[2, 2, 2, 2, 1, 1])[0]

    # Limitar filtros activos a máximo 5
    _filter_keys = [
        "macd_filter","adx_filter","supertrend_filter","ichimoku_filter",
        "stoch_rsi","cci_filter","williams_r","momentum_filter","bb_filter",
        "atr_volatility","keltner_filter","obv_filter","vwap_filter",
        "volume_delta","cvd_filter","market_structure","breakout_range",
        "candle_pattern","order_block","pivot_filter","sr_breakout",
        "fib_retracement","rsi_divergence","btc_correlation","funding_filter",
        "fear_greed_filter","session_filter"
    ]
    _filter_vals = {
        "macd_filter": rng.choice(MACD_FILTERS), "adx_filter": rng.choice(ADX_FILTERS),
        "supertrend_filter": rng.choice(ST_FILTERS), "ichimoku_filter": rng.choice(ICHI_FILTERS),
        "stoch_rsi": rng.choice(STOCH_RSI), "cci_filter": rng.choice(CCI_FILTERS),
        "williams_r": rng.choice(WILLIAMS_R), "momentum_filter": rng.choice(MOM_FILTERS),
        "bb_filter": rng.choice(BB_FILTERS), "atr_volatility": rng.choice(ATR_VOL),
        "keltner_filter": rng.choice(KELT_FILTERS), "obv_filter": rng.choice(OBV_FILTERS),
        "vwap_filter": rng.choice(VWAP_FILTERS), "volume_delta": rng.choice(VOL_DELTA),
        "cvd_filter": rng.choice(CVD_FILTERS), "market_structure": rng.choice(MKT_STRUCT),
        "breakout_range": rng.choice(BREAKOUT_RANGE), "candle_pattern": rng.choice(CANDLE_PAT),
        "order_block": rng.choice(ORDER_BLOCK), "pivot_filter": rng.choice(PIVOT_FILTERS),
        "sr_breakout": rng.choice(SR_BREAKOUT), "fib_retracement": rng.choice(FIB_RET),
        "rsi_divergence": rng.choice(RSI_DIV), "btc_correlation": rng.choice(BTC_CORR),
        "funding_filter": rng.choice(FUNDING_FILT), "fear_greed_filter": rng.choice(FNG_FILTERS),
        "session_filter": rng.choice(SESSION_FILT)
    }
    return OptParams(
        interval=rng.choices(TIMEFRAMES, weights=[1, 2, 3, 3, 1])[0],
        ma_type=ma[0], ma_fast=ma[1], ma_slow=ma[2],
        leverage=rng.choices(LEVERAGES, weights=[1, 1, 2, 3, 5])[0],
        trailing_pct=rng.choice(TRAILING_PCTS),
        sl_type=rng.choice(SL_TYPES), fib_mode=rng.choice(FIB_MODES),
        rsi_filter=rng.choice(RSI_FILTERS), ema200_filter=rng.choice(EMA200_FILT),
        atr_filter=rng.choice(ATR_FILTERS),
        compound=rng.choices([True, False], weights=[80, 20])[0],
        time_filter=rng.choice(TIME_FILTERS),
        vol_profile=rng.choice(VOL_PROFILES), liq_confirm=rng.choice(LIQ_CONFIRMS),
        risk_pct=rng.choice(RISK_PCTS),
        macd_filter=_filter_vals["macd_filter"], adx_filter=_filter_vals["adx_filter"],
        supertrend_filter=_filter_vals["supertrend_filter"], ichimoku_filter=_filter_vals["ichimoku_filter"],
        stoch_rsi=_filter_vals["stoch_rsi"], cci_filter=_filter_vals["cci_filter"],
        williams_r=_filter_vals["williams_r"], momentum_filter=_filter_vals["momentum_filter"],
        bb_filter=_filter_vals["bb_filter"], atr_volatility=_filter_vals["atr_volatility"],
        keltner_filter=_filter_vals["keltner_filter"], obv_filter=_filter_vals["obv_filter"],
        vwap_filter=_filter_vals["vwap_filter"], volume_delta=_filter_vals["volume_delta"],
        cvd_filter=_filter_vals["cvd_filter"], market_structure=_filter_vals["market_structure"],
        breakout_range=_filter_vals["breakout_range"], candle_pattern=_filter_vals["candle_pattern"],
        order_block=_filter_vals["order_block"], pivot_filter=_filter_vals["pivot_filter"],
        sr_breakout=_filter_vals["sr_breakout"], fib_retracement=_filter_vals["fib_retracement"],
        rsi_divergence=_filter_vals["rsi_divergence"], btc_correlation=_filter_vals["btc_correlation"],
        funding_filter=_filter_vals["funding_filter"], fear_greed_filter=_filter_vals["fear_greed_filter"],
        session_filter=_filter_vals["session_filter"],
        position_sizing=rng.choice(POS_SIZING),
        max_trades_day=rng.choice(MAX_TRADES_DAY), trailing_type=rng.choice(TRAIL_TYPES),
        min_confluences=rng.choice(MIN_CONF),
        tp_type=rng.choice(TP_TYPES), tp_pct=rng.choice(TP_PCT_OPTS),
        tp_atr=rng.choice(TP_ATR_OPTS), trailing_activation=rng.choice(TRAIL_ACT),
        trailing_progressive=rng.choice(TRAIL_PROG), atr_tp_adjust=rng.choice(ATR_TP_ADJ),
        partial_close=rng.choice(PARTIAL_CLOSE), partial_trigger=rng.choice(PARTIAL_TRIG),
        breakeven=rng.choice(BREAKEVEN_OPTS), time_exit=rng.choice(TIME_EXIT_OPTS),
        session_exit=rng.choice(SESSION_EXIT), weekend_exit=rng.choice(WEEKEND_EXIT),
        rr_min=rng.choice(RR_MIN_OPTS),
        coin=rng.choice(ACTIVE_COINS),
        direction=rng.choice(MASTER_DIRECTIONS),
    )


def perturb(p: OptParams, rng: random.Random) -> OptParams:
    d = asdict(p)
    all_keys = list(_PARAM_FIELDS_SET - {"ma_type","ma_fast","ma_slow"}) + ["ma_pair"]
    key = rng.choice(all_keys)
    opts = {
        "coin": ACTIVE_COINS, "direction": MASTER_DIRECTIONS,
        "interval": TIMEFRAMES, "leverage": LEVERAGES, "trailing_pct": TRAILING_PCTS,
        "sl_type": SL_TYPES, "fib_mode": FIB_MODES, "rsi_filter": RSI_FILTERS,
        "ema200_filter": EMA200_FILT, "atr_filter": ATR_FILTERS,
        "compound": COMPOUNDS, "time_filter": TIME_FILTERS,
        "vol_profile": VOL_PROFILES, "liq_confirm": LIQ_CONFIRMS, "risk_pct": RISK_PCTS,
        "macd_filter": MACD_FILTERS, "adx_filter": ADX_FILTERS,
        "supertrend_filter": ST_FILTERS, "ichimoku_filter": ICHI_FILTERS,
        "stoch_rsi": STOCH_RSI, "cci_filter": CCI_FILTERS,
        "williams_r": WILLIAMS_R, "momentum_filter": MOM_FILTERS,
        "bb_filter": BB_FILTERS, "atr_volatility": ATR_VOL,
        "keltner_filter": KELT_FILTERS, "obv_filter": OBV_FILTERS,
        "vwap_filter": VWAP_FILTERS, "volume_delta": VOL_DELTA,
        "cvd_filter": CVD_FILTERS, "market_structure": MKT_STRUCT,
        "breakout_range": BREAKOUT_RANGE, "candle_pattern": CANDLE_PAT,
        "order_block": ORDER_BLOCK, "pivot_filter": PIVOT_FILTERS,
        "sr_breakout": SR_BREAKOUT, "fib_retracement": FIB_RET,
        "rsi_divergence": RSI_DIV, "btc_correlation": BTC_CORR,
        "funding_filter": FUNDING_FILT, "fear_greed_filter": FNG_FILTERS,
        "session_filter": SESSION_FILT, "position_sizing": POS_SIZING,
        "max_trades_day": MAX_TRADES_DAY, "trailing_type": TRAIL_TYPES,
        "min_confluences": MIN_CONF,
        "tp_type": TP_TYPES, "tp_pct": TP_PCT_OPTS, "tp_atr": TP_ATR_OPTS,
        "trailing_activation": TRAIL_ACT, "trailing_progressive": TRAIL_PROG,
        "atr_tp_adjust": ATR_TP_ADJ, "partial_close": PARTIAL_CLOSE,
        "partial_trigger": PARTIAL_TRIG, "breakeven": BREAKEVEN_OPTS,
        "time_exit": TIME_EXIT_OPTS, "session_exit": SESSION_EXIT,
        "weekend_exit": WEEKEND_EXIT, "rr_min": RR_MIN_OPTS,
    }
    if key == "ma_pair":
        ma = rng.choices(MA_PAIRS, weights=[2, 2, 2, 2, 1, 1])[0]
        d["ma_type"] = ma[0]; d["ma_fast"] = ma[1]; d["ma_slow"] = ma[2]
    elif key in opts:
        d[key] = rng.choice(opts[key])
    return OptParams(**d)


# ─── CANDLE FETCHING ──────────────────────────────────────────────────────────
_INTERVAL_MS = {
    "15m": 900_000, "30m": 1_800_000, "1h": 3_600_000,
    "2h": 7_200_000, "4h": 14_400_000,
}
_HL_MAX_CANDLES = 5000  # Hyperliquid max per request


def _fetch_hl(coin: str, interval: str, days: int) -> Optional[np.ndarray]:
    """Fetch candles with backward pagination when API returns max (5000) candles."""
    end_ms   = int(time.time() * 1000)
    start_ms = end_ms - days * 86_400_000
    iv_ms    = _INTERVAL_MS.get(interval, 3_600_000)

    # First request: get latest data
    arr = _fetch_hl_single(coin, interval, start_ms, end_ms)
    if arr is None or len(arr) < 50:
        return arr

    # If we got less than max, API has no more data
    if len(arr) < _HL_MAX_CANDLES:
        return arr

    # Got exactly max — paginate backward to get older data
    all_rows = [arr]
    oldest_ts = int(arr[0, 0])

    while oldest_ts > start_ms:
        chunk_end = oldest_ts - 1  # just before current oldest
        chunk = _fetch_hl_single(coin, interval, start_ms, chunk_end)
        if chunk is None or len(chunk) == 0:
            break
        all_rows.append(chunk)
        new_oldest = int(chunk[0, 0])
        if new_oldest >= oldest_ts:
            break  # no progress
        oldest_ts = new_oldest
        time.sleep(0.1)

    if len(all_rows) == 1:
        return arr

    combined = np.concatenate(all_rows, axis=0)
    _, idx = np.unique(combined[:, 0], return_index=True)
    combined = combined[np.sort(idx)]
    return combined if len(combined) > 50 else None


def _fetch_hl_single(coin: str, interval: str,
                     start_ms: int, end_ms: int) -> Optional[np.ndarray]:
    """Single API call to Hyperliquid candleSnapshot."""
    try:
        r = requests.post(HL_URL, json={
            "type": "candleSnapshot",
            "req":  {"coin": coin, "interval": interval,
                     "startTime": start_ms, "endTime": end_ms},
        }, timeout=30)
        if not r.ok: return None
        candles = r.json()
        if not candles: return None
        return np.array([
            [c["T"], float(c["o"]), float(c["h"]), float(c["l"]),
             float(c["c"]), float(c.get("v", 0))]
            for c in candles
        ], dtype=np.float64)
    except Exception as e:
        print(f"[{ts()}]   ! Error {coin}/{interval}: {e}")
        return None


def load_or_fetch(cache_dir: Path, coin: str, interval: str, days: int) -> Optional[np.ndarray]:
    cache_file = cache_dir / f"{coin}_{interval}_{days}d.pkl"
    if cache_file.exists():
        try:
            with open(cache_file, "rb") as f:
                data = pickle.load(f)
            if data is not None and len(data) > 50:
                return data
        except Exception:
            pass
    time.sleep(random.uniform(0.05, 0.25))
    arr = _fetch_hl(coin, interval, days)
    if arr is not None and len(arr) > 50:
        try:
            with open(cache_file, "wb") as f:
                pickle.dump(arr, f)
        except Exception:
            pass
    return arr


# ─── INDICADORES ─────────────────────────────────────────────────────────────
def _safe_nanmin(arr: np.ndarray) -> float:
    if len(arr) == 0: return np.nan
    valid = arr[~np.isnan(arr)]
    return float(np.min(valid)) if len(valid) > 0 else np.nan

def _safe_nanmax(arr: np.ndarray) -> float:
    if len(arr) == 0: return np.nan
    valid = arr[~np.isnan(arr)]
    return float(np.max(valid)) if len(valid) > 0 else np.nan

def _safe_nanmean(arr: np.ndarray) -> float:
    if len(arr) == 0: return np.nan
    valid = arr[~np.isnan(arr)]
    return float(np.mean(valid)) if len(valid) > 0 else np.nan


def _ema(arr: np.ndarray, n: int) -> np.ndarray:
    if len(arr) < n: return np.full_like(arr, np.nan)
    out = np.empty_like(arr); out[:n-1] = np.nan
    out[n-1] = _safe_nanmean(arr[:n])
    k = 2.0 / (n + 1); k1 = 1.0 - k; prev = out[n-1]
    for i in range(n, len(arr)):
        if arr[i] == arr[i]: prev = arr[i] * k + prev * k1
        out[i] = prev
    return out

def _sma(arr: np.ndarray, n: int) -> np.ndarray:
    if len(arr) < n: return np.full_like(arr, np.nan)
    out = np.full_like(arr, np.nan)
    cs = np.cumsum(np.where(np.isnan(arr), 0, arr))
    out[n-1] = cs[n-1] / n; out[n:] = (cs[n:] - cs[:len(arr)-n]) / n
    return out

def _rsi(closes: np.ndarray, n: int = 14) -> np.ndarray:
    out = np.full_like(closes, np.nan)
    if len(closes) < n + 1: return out
    d = np.diff(closes)
    gains = np.maximum(d, 0.0); losses = np.maximum(-d, 0.0)
    ag = _safe_nanmean(gains[:n]); al = _safe_nanmean(losses[:n])
    if np.isnan(ag): ag = 0.0
    if np.isnan(al): al = 0.0
    for i in range(n, len(d)):
        ag = (ag*(n-1) + gains[i]) / n; al = (al*(n-1) + losses[i]) / n
        rs = ag / al if al > 0 else 100.0
        out[i+1] = 100.0 - 100.0 / (1.0 + rs)
    return out

def _atr(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 14) -> np.ndarray:
    out = np.full_like(close, np.nan)
    if len(close) < n + 1: return out
    tr = np.maximum(high[1:]-low[1:],
         np.maximum(np.abs(high[1:]-close[:-1]), np.abs(low[1:]-close[:-1])))
    out[n] = _safe_nanmean(tr[:n])
    for i in range(n, len(tr)): out[i+1] = (out[i]*(n-1) + tr[i]) / n
    return out

def _roll_max(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w: return out
    shape = arr.shape[:-1] + (arr.shape[-1]-w+1, w)
    strides = arr.strides + (arr.strides[-1],)
    wins = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w-1:] = wins.max(axis=1); return out

def _roll_min(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w: return out
    shape = arr.shape[:-1] + (arr.shape[-1]-w+1, w)
    strides = arr.strides + (arr.strides[-1],)
    wins = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w-1:] = wins.min(axis=1); return out

def _roll_std(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w: return out
    shape = arr.shape[:-1] + (arr.shape[-1]-w+1, w)
    strides = arr.strides + (arr.strides[-1],)
    wins = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w-1:] = wins.std(axis=1); return out

def _macd(closes: np.ndarray, fast: int = 12, slow: int = 26,
          sig: int = 9) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    L = len(closes); ef = _ema(closes, fast); es = _ema(closes, slow)
    line = ef - es; signal = np.full(L, np.nan)
    valid_idx = np.where(~np.isnan(line))[0]
    if len(valid_idx) >= sig:
        start = valid_idx[0]; sig_sub = _ema(line[start:], sig)
        signal[start:start + len(sig_sub)] = sig_sub
    return line, signal, line - signal

def _adx(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 14) -> np.ndarray:
    L = len(close); adx = np.full(L, np.nan)
    if L < n * 2 + 2: return adx
    tr  = np.maximum(high[1:]-low[1:], np.maximum(np.abs(high[1:]-close[:-1]), np.abs(low[1:]-close[:-1])))
    dmp = np.where((high[1:]-high[:-1]) > (low[:-1]-low[1:]), np.maximum(high[1:]-high[:-1], 0.0), 0.0)
    dmm = np.where((low[:-1]-low[1:]) > (high[1:]-high[:-1]), np.maximum(low[:-1]-low[1:], 0.0), 0.0)
    N = len(tr); atr_w = np.zeros(N); dmp_w = np.zeros(N); dmm_w = np.zeros(N)
    if n - 1 < N:
        atr_w[n-1] = float(np.sum(tr[:n])); dmp_w[n-1] = float(np.sum(dmp[:n])); dmm_w[n-1] = float(np.sum(dmm[:n]))
    for i in range(n, N):
        atr_w[i] = atr_w[i-1] - atr_w[i-1]/n + tr[i]
        dmp_w[i] = dmp_w[i-1] - dmp_w[i-1]/n + dmp[i]
        dmm_w[i] = dmm_w[i-1] - dmm_w[i-1]/n + dmm[i]
    with np.errstate(divide="ignore", invalid="ignore"):
        dip = np.where(atr_w > 0, dmp_w/atr_w*100, 0.0)
        dim = np.where(atr_w > 0, dmm_w/atr_w*100, 0.0)
        dx  = np.where(dip+dim > 0, np.abs(dip-dim)/(dip+dim)*100, 0.0)
    adx_arr = np.full(N, np.nan); first = n * 2 - 2
    if first < N:
        adx_arr[first] = _safe_nanmean(dx[n-1:n*2-1])
        for i in range(first + 1, N): adx_arr[i] = (adx_arr[i-1]*(n-1) + dx[i]) / n
    adx[1:] = adx_arr; return adx

def _supertrend(high: np.ndarray, low: np.ndarray, close: np.ndarray,
                n: int = 10, m: float = 3.0) -> np.ndarray:
    direction = np.zeros(len(close)); atr = _atr(high, low, close, n)
    hl2 = (high + low) / 2.0; ub_basic = hl2 + m * atr; lb_basic = hl2 - m * atr
    ub = np.copy(ub_basic); lb = np.copy(lb_basic)
    for i in range(1, len(close)):
        if np.isnan(atr[i]): continue
        ub[i] = ub_basic[i] if ub_basic[i] < ub[i-1] or close[i-1] > ub[i-1] else ub[i-1]
        lb[i] = lb_basic[i] if lb_basic[i] > lb[i-1] or close[i-1] < lb[i-1] else lb[i-1]
        if close[i] > ub[i]:   direction[i] = 1
        elif close[i] < lb[i]: direction[i] = -1
        else:                  direction[i] = direction[i-1]
    return direction

def _ichimoku(high: np.ndarray, low: np.ndarray, close: np.ndarray):
    tenkan = (_roll_max(high, 9) + _roll_min(low, 9)) / 2.0
    kijun  = (_roll_max(high, 26) + _roll_min(low, 26)) / 2.0
    span_a = (tenkan + kijun) / 2.0
    span_b = (_roll_max(high, 52) + _roll_min(low, 52)) / 2.0
    return tenkan, kijun, span_a, span_b

def _stoch_rsi(closes: np.ndarray, rsi_n: int = 14, stoch_n: int = 14,
               k_s: int = 3, d_s: int = 3) -> Tuple[np.ndarray, np.ndarray]:
    rsi = _rsi(closes, rsi_n); rsi_max = _roll_max(rsi, stoch_n); rsi_min = _roll_min(rsi, stoch_n)
    with np.errstate(invalid="ignore", divide="ignore"):
        raw_k = np.where(rsi_max - rsi_min > 0, (rsi - rsi_min) / (rsi_max - rsi_min) * 100.0, 50.0)
    raw_k[np.isnan(rsi)] = np.nan
    return _sma(raw_k, k_s), _sma(_sma(raw_k, k_s), d_s)

def _cci(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 20) -> np.ndarray:
    tp = (high + low + close) / 3.0; tp_sma = _sma(tp, n); out = np.full_like(close, np.nan)
    for i in range(n-1, len(close)):
        md = _safe_nanmean(np.abs(tp[i-n+1:i+1] - tp_sma[i]))
        if md > 0: out[i] = (tp[i] - tp_sma[i]) / (0.015 * md)
    return out

def _williams_r_arr(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 14) -> np.ndarray:
    hh = _roll_max(high, n); ll = _roll_min(low, n)
    with np.errstate(invalid="ignore", divide="ignore"):
        return -100.0 * (hh - close) / np.where(hh - ll > 0, hh - ll, np.nan)

def _bbands(closes: np.ndarray, n: int = 20, k: float = 2.0) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    sma = _sma(closes, n); std = _roll_std(closes, n)
    upper = sma + k * std; lower = sma - k * std
    with np.errstate(invalid="ignore", divide="ignore"):
        width = np.where(sma > 0, (upper - lower) / sma, np.nan)
    return upper, lower, width

def _keltner(high: np.ndarray, low: np.ndarray, close: np.ndarray,
             n: int = 20, m: float = 2.0) -> Tuple[np.ndarray, np.ndarray]:
    ema = _ema(close, n); atr = _atr(high, low, close, n)
    return ema + m * atr, ema - m * atr

def _obv(close: np.ndarray, volume: np.ndarray) -> np.ndarray:
    obv = np.zeros(len(close))
    for i in range(1, len(close)):
        if close[i] > close[i-1]:   obv[i] = obv[i-1] + volume[i]
        elif close[i] < close[i-1]: obv[i] = obv[i-1] - volume[i]
        else:                        obv[i] = obv[i-1]
    return obv

def _vwap_daily(timestamps: np.ndarray, high: np.ndarray,
                low: np.ndarray, close: np.ndarray, volume: np.ndarray) -> np.ndarray:
    tp = (high + low + close) / 3.0; vwap = np.full_like(close, np.nan)
    cum_tv = 0.0; cum_v = 0.0; prev_day = -1; days = (timestamps // 86_400_000).astype(np.int64)
    for i in range(len(close)):
        if days[i] != prev_day: cum_tv = 0.0; cum_v = 0.0; prev_day = int(days[i])
        cum_tv += tp[i] * volume[i]; cum_v += volume[i]
        vwap[i] = cum_tv / cum_v if cum_v > 0 else close[i]
    return vwap

def _cvd(opens: np.ndarray, closes: np.ndarray, volumes: np.ndarray) -> np.ndarray:
    return np.cumsum(np.where(closes >= opens, volumes, -volumes))

def _market_structure(high: np.ndarray, low: np.ndarray, n: int = 10) -> np.ndarray:
    struct = np.zeros(len(high))
    for i in range(n, len(high)):
        ph = high[i-n:i]; pl = low[i-n:i]
        ph_max = _safe_nanmax(ph); pl_min = _safe_nanmin(pl)
        if np.isnan(ph_max) or np.isnan(pl_min): continue
        if high[i] > ph_max and low[i] > pl_min:   struct[i] = 1
        elif high[i] < ph_max and low[i] < pl_min: struct[i] = -1
    return struct

def _order_blocks(closes: np.ndarray, atr: np.ndarray, multiplier: float = 2.0):
    ob_bull = np.zeros(len(closes), dtype=bool); ob_bear = np.zeros(len(closes), dtype=bool)
    for i in range(1, len(closes)):
        av = atr[i] if atr[i] == atr[i] else 0.0
        if av <= 0: continue
        body = abs(closes[i] - closes[i-1])
        if body > multiplier * av:
            if closes[i] > closes[i-1]: ob_bear[i] = True
            else:                        ob_bull[i] = True
    return ob_bull, ob_bear

def _pivot_levels(timestamps: np.ndarray, high: np.ndarray, low: np.ndarray,
                  close: np.ndarray, weekly: bool = False):
    period_ms = 604_800_000 if weekly else 86_400_000
    periods = (timestamps // period_ms).astype(np.int64)
    P  = np.full_like(close, np.nan); R1 = np.full_like(close, np.nan)
    R2 = np.full_like(close, np.nan); S1 = np.full_like(close, np.nan)
    S2 = np.full_like(close, np.nan)
    prev_p = -1; ph = pl = pc = 0.0
    for i in range(len(close)):
        p = int(periods[i])
        if p != prev_p and prev_p >= 0:
            pivot = (ph + pl + pc) / 3.0
            R1[i] = 2*pivot - pl; S1[i] = 2*pivot - ph
            R2[i] = pivot + (ph - pl); S2[i] = pivot - (ph - pl); P[i] = pivot; prev_p = p
        elif p == prev_p and i > 0:
            P[i] = P[i-1]; R1[i] = R1[i-1]; R2[i] = R2[i-1]; S1[i] = S1[i-1]; S2[i] = S2[i-1]
        if p != prev_p: prev_p = p; ph = high[i]; pl = low[i]; pc = close[i]
        else: ph = max(ph, high[i]); pl = min(pl, low[i]); pc = close[i]
    return P, R1, R2, S1, S2

def _rsi_divergence_arr(close: np.ndarray, rsi: np.ndarray, lb: int = 14):
    bull = np.zeros(len(close), dtype=bool); bear = np.zeros(len(close), dtype=bool)
    for i in range(lb*2, len(close)):
        if close[i] < close[i-lb] and rsi[i] > rsi[i-lb] and rsi[i] == rsi[i]: bull[i] = True
        if close[i] > close[i-lb] and rsi[i] < rsi[i-lb] and rsi[i] == rsi[i]: bear[i] = True
    return bull, bear

def _parabolic_sar(high: np.ndarray, low: np.ndarray,
                   af_start: float = 0.02, af_step: float = 0.02, af_max: float = 0.2) -> np.ndarray:
    n = len(high); sar = np.full(n, np.nan)
    if n < 3: return sar
    bull = True; sar[0] = low[0]; ep = high[0]; af = af_start
    for i in range(1, n):
        if bull:
            sar[i] = sar[i-1] + af * (ep - sar[i-1])
            sar[i] = min(sar[i], low[i-1], low[i-2]) if i >= 2 else min(sar[i], low[i-1])
            if low[i] < sar[i]:
                bull = False; sar[i] = ep; ep = low[i]; af = af_start
            else:
                if high[i] > ep: ep = high[i]; af = min(af + af_step, af_max)
        else:
            sar[i] = sar[i-1] + af * (ep - sar[i-1])
            sar[i] = max(sar[i], high[i-1], high[i-2]) if i >= 2 else max(sar[i], high[i-1])
            if high[i] > sar[i]:
                bull = True; sar[i] = ep; ep = high[i]; af = af_start
            else:
                if low[i] < ep: ep = low[i]; af = min(af + af_step, af_max)
    return sar

def _hour_mask(ts_arr: np.ndarray, allowed: set) -> np.ndarray:
    hours = (ts_arr // 3_600_000 % 24).astype(np.int32)
    mask = np.zeros(len(ts_arr), dtype=bool)
    for h in allowed: mask |= (hours == h)
    return mask

_LONDON_NY = set(range(7, 22)); _ASIA = set(range(0, 9))
_NY = set(range(13, 22)); _LDN_NY_OV = set(range(13, 18))
_FIBO_L = np.array([0.236, 0.382, 0.500, 0.618, 0.786])

def _near_fibo_fast(price: float, hi: float, lo: float) -> bool:
    rng = hi - lo
    if rng <= 0 or price <= 0: return False
    return bool(np.any(np.abs(lo + rng * _FIBO_L - price) / price <= 0.015))


# ─── WORKER INIT ──────────────────────────────────────────────────────────────
_worker_ind: Dict[str, Dict[str, dict]] = {}


def _worker_init(cache_pkl: str):
    global _worker_ind, _FEAR_GREED_IDX, _FUNDING_RATE
    try:
        with open(cache_pkl, "rb") as f:
            raw: dict = pickle.load(f)
        _FEAR_GREED_IDX = raw.get("_fng")
        _FUNDING_RATE   = raw.get("_funding")
        candles         = raw.get("candles", {})
    except Exception:
        _worker_ind = {}; return

    # BTC trend per TF (for btc_correlation filter on non-BTC coins)
    btc_trend: Dict[str, np.ndarray] = {}
    for tf in TIMEFRAMES:
        btc_arr = candles.get(tf, {}).get("BTC")
        if btc_arr is not None and len(btc_arr) >= 60:
            c = btc_arr[:, 4]; e20 = _ema(c, 20); e50 = _ema(c, 50)
            btc_trend[tf] = np.where(e20 > e50, 1.0, -1.0)

    _worker_ind = {}
    for tf, coins_dict in candles.items():
        _worker_ind[tf] = {}
        for coin, arr in coins_dict.items():
            if arr is None or len(arr) < 250: continue
            ts_col = arr[:, 0]; opens = arr[:, 1]; highs = arr[:, 2]
            lows = arr[:, 3]; closes = arr[:, 4]; volumes = arr[:, 5]

            atr14 = _atr(highs, lows, closes, 14); rsi14 = _rsi(closes, 14)
            ema200 = _ema(closes, 200); vol_ma = _sma(volumes, 20)
            roll_hi = _roll_max(highs, 60); roll_lo = _roll_min(lows, 60)

            macd_l, macd_s, macd_h = _macd(closes)
            lb = 14
            macd_bull_div = np.zeros(len(closes), dtype=bool)
            macd_bear_div = np.zeros(len(closes), dtype=bool)
            for i in range(lb*2, len(closes)):
                if macd_h[i] != macd_h[i]: continue
                sl_c = closes[i-lb:i]; sl_mh = macd_h[i-lb:i]
                lo_c = _safe_nanmin(sl_c); hi_c = _safe_nanmax(sl_c)
                lo_mh = _safe_nanmin(sl_mh); hi_mh = _safe_nanmax(sl_mh)
                if np.isnan(lo_c) or np.isnan(hi_c) or np.isnan(lo_mh) or np.isnan(hi_mh): continue
                if closes[i] <= lo_c and macd_h[i] > lo_mh: macd_bull_div[i] = True
                if closes[i] >= hi_c and macd_h[i] < hi_mh: macd_bear_div[i] = True

            adx = _adx(highs, lows, closes); st = _supertrend(highs, lows, closes)
            tenkan, kijun, span_a, span_b = _ichimoku(highs, lows, closes)
            cloud_top = np.maximum(span_a, span_b); cloud_bot = np.minimum(span_a, span_b)
            sk, sd = _stoch_rsi(closes); cci = _cci(highs, lows, closes)
            wr = _williams_r_arr(highs, lows, closes)
            mom10 = closes - np.roll(closes, 10); mom10[:10] = np.nan
            bb_u, bb_l, bb_w = _bbands(closes); kelt_u, kelt_l = _keltner(highs, lows, closes)
            bb_w_sma = _sma(bb_w, 20)
            obv_arr = _obv(closes, volumes); obv_ema = _ema(obv_arr, 20)
            obv_bull_div = np.zeros(len(closes), dtype=bool)
            obv_bear_div = np.zeros(len(closes), dtype=bool)
            for i in range(lb*2, len(closes)):
                sl_c = closes[i-lb:i]; sl_o = obv_arr[i-lb:i]
                lo_c = _safe_nanmin(sl_c); hi_c = _safe_nanmax(sl_c)
                lo_o = _safe_nanmin(sl_o); hi_o = _safe_nanmax(sl_o)
                if np.isnan(lo_c) or np.isnan(hi_c) or np.isnan(lo_o) or np.isnan(hi_o): continue
                if closes[i] <= lo_c and obv_arr[i] > lo_o: obv_bull_div[i] = True
                if closes[i] >= hi_c and obv_arr[i] < hi_o: obv_bear_div[i] = True

            vwap_arr = _vwap_daily(ts_col, highs, lows, closes, volumes)
            cvd_arr  = _cvd(opens, closes, volumes); cvd_ema = _ema(cvd_arr.astype(float), 14)
            ms_arr   = _market_structure(highs, lows)
            ob_bull, ob_bear = _order_blocks(closes, atr14)
            piv_p, piv_r1, piv_r2, piv_s1, piv_s2 = _pivot_levels(ts_col, highs, lows, closes, weekly=False)
            wpiv_p, wpiv_r1, wpiv_r2, wpiv_s1, wpiv_s2 = _pivot_levels(ts_col, highs, lows, closes, weekly=True)
            rsi_bull_div, rsi_bear_div = _rsi_divergence_arr(closes, rsi14)
            psar_arr = _parabolic_sar(highs, lows)

            price_mean  = float(np.nanmean(closes))
            atr_valid   = atr14[~np.isnan(atr14)]
            atr_pct_ref = (float(np.nanmean(atr_valid)) / price_mean
                           if len(atr_valid) > 0 and price_mean > 0 else 0.0)
            atr_std     = float(np.nanstd(atr_valid)) if len(atr_valid) > 1 else 0.0
            atr_mean    = float(np.nanmean(atr_valid)) if len(atr_valid) > 0 else 0.0

            roll_hi10  = _roll_max(highs, 10); roll_lo10  = _roll_min(lows, 10)
            roll_hi20  = _roll_max(highs, 20); roll_lo20  = _roll_min(lows, 20)
            roll_hi50  = _roll_max(highs, 50); roll_lo50  = _roll_min(lows, 50)
            roll_hi100 = _roll_max(highs, 100); roll_lo100 = _roll_min(lows, 100)

            ind: dict = {
                "ts": ts_col, "opens": opens, "highs": highs, "lows": lows,
                "closes": closes, "volumes": volumes,
                "ema200": ema200, "rsi14": rsi14, "atr14": atr14,
                "vol_ma": vol_ma, "roll_hi": roll_hi, "roll_lo": roll_lo,
                "h_london": _hour_mask(ts_col, _LONDON_NY),
                "h_asia":   _hour_mask(ts_col, _ASIA),
                "h_ny":     _hour_mask(ts_col, _NY),
                "h_lno":    _hour_mask(ts_col, _LDN_NY_OV),
                "atr_pct_ref": atr_pct_ref, "price_mean": price_mean,
                "atr_mean": atr_mean, "atr_std": atr_std,
                "macd_l": macd_l, "macd_s": macd_s, "macd_h": macd_h,
                "macd_bull_div": macd_bull_div, "macd_bear_div": macd_bear_div,
                "adx": adx, "st": st,
                "cloud_top": cloud_top, "cloud_bot": cloud_bot, "tenkan": tenkan, "kijun": kijun,
                "sk": sk, "sd": sd, "cci": cci, "wr": wr, "mom10": mom10,
                "bb_u": bb_u, "bb_l": bb_l, "bb_w": bb_w, "bb_w_sma": bb_w_sma,
                "kelt_u": kelt_u, "kelt_l": kelt_l,
                "obv": obv_arr, "obv_ema": obv_ema,
                "obv_bull_div": obv_bull_div, "obv_bear_div": obv_bear_div,
                "vwap": vwap_arr, "cvd": cvd_arr.astype(float), "cvd_ema": cvd_ema,
                "ms": ms_arr, "ob_bull": ob_bull, "ob_bear": ob_bear,
                "piv_p": piv_p, "piv_r1": piv_r1, "piv_r2": piv_r2,
                "piv_s1": piv_s1, "piv_s2": piv_s2,
                "wpiv_p": wpiv_p, "wpiv_r1": wpiv_r1, "wpiv_r2": wpiv_r2,
                "wpiv_s1": wpiv_s1, "wpiv_s2": wpiv_s2,
                "rsi_bull_div": rsi_bull_div, "rsi_bear_div": rsi_bear_div,
                "psar": psar_arr,
                "roll_hi10": roll_hi10, "roll_lo10": roll_lo10,
                "roll_hi20": roll_hi20, "roll_lo20": roll_lo20,
                "roll_hi50": roll_hi50, "roll_lo50": roll_lo50,
                "roll_hi100": roll_hi100, "roll_lo100": roll_lo100,
                "btc_trend": btc_trend.get(tf),
            }
            for (ma_type, fast, slow) in MA_PAIRS:
                k = f"{ma_type}_{fast}_{slow}"; fn = _ema if ma_type == "ema" else _sma
                ind[f"maf_{k}"] = fn(closes, fast); ind[f"mas_{k}"] = fn(closes, slow)

            _worker_ind[tf][coin] = ind


# ─── HELPERS TP / PARTIAL / BREAKEVEN (idénticos a v7) ───────────────────────
def _calc_tp(tp_type: str, tp_pct: int, tp_atr_mult: float,
             signal: str, entry: float, stop: float,
             av: float, ind: dict, i: int) -> Optional[float]:
    if tp_type == "none": return None
    risk_dist = abs(entry - stop); is_long = (signal == "long")
    if tp_type == "fixed_pct":
        pct = tp_pct / 100.0
        return entry * (1.0 + pct) if is_long else entry * (1.0 - pct)
    elif tp_type == "atr_multiple":
        return (entry + tp_atr_mult * av) if is_long else (entry - tp_atr_mult * av)
    elif tp_type == "bb_opposite":
        bu = ind["bb_u"][i]; bl = ind["bb_l"][i]
        if bu == bu and bl == bl: return bu if is_long else bl
    elif tp_type == "vwap_target":
        vv = ind["vwap"][i]
        if vv == vv:
            if is_long and vv > entry: return vv
            if not is_long and vv < entry: return vv
    elif tp_type == "ichimoku_cloud":
        ct = ind["cloud_top"][i]; cb = ind["cloud_bot"][i]
        if ct == ct and cb == cb: return ct if is_long else cb
    elif tp_type == "fib_127":
        return (entry + 1.272 * risk_dist) if is_long else (entry - 1.272 * risk_dist)
    elif tp_type == "fib_161":
        return (entry + 1.618 * risk_dist) if is_long else (entry - 1.618 * risk_dist)
    elif tp_type == "fib_261":
        return (entry + 2.618 * risk_dist) if is_long else (entry - 2.618 * risk_dist)
    elif tp_type == "pivot_r1":
        r1 = ind["piv_r1"][i]; s1 = ind["piv_s1"][i]
        if is_long and r1 == r1 and r1 > entry: return r1
        if not is_long and s1 == s1 and s1 < entry: return s1
    elif tp_type == "pivot_r2":
        r2 = ind["piv_r2"][i]; s2 = ind["piv_s2"][i]
        if is_long and r2 == r2 and r2 > entry: return r2
        if not is_long and s2 == s2 and s2 < entry: return s2
    elif tp_type in ("pivot_s1", "pivot_s2"):
        s = ind["piv_s1"][i] if tp_type == "pivot_s1" else ind["piv_s2"][i]
        r = ind["piv_r1"][i] if tp_type == "pivot_s1" else ind["piv_r2"][i]
        if not is_long and s == s and s < entry: return s
        if is_long and r == r and r > entry: return r
    elif tp_type == "sr_level":
        rh = ind["roll_hi"][i]; rl = ind["roll_lo"][i]
        if is_long and rh == rh and rh > entry: return rh
        if not is_long and rl == rl and rl < entry: return rl
    elif tp_type == "ma_cross_opposite":
        for k in ind:
            if k.startswith("mas_"):
                v = ind[k][i]
                if v == v and v > 0:
                    if is_long and v > entry: return v
                    if not is_long and v < entry: return v
                break
    return None

def _calc_partial_trigger(partial_trigger: str, signal: str, entry: float, stop: float, av: float) -> float:
    risk_dist = abs(entry - stop); is_long = (signal == "long")
    dist = {"1atr": av, "2atr": 2.0*av, "5pct": entry*0.05, "10pct": entry*0.10,
            "fib_127": 1.272*risk_dist, "fib_161": 1.618*risk_dist}.get(partial_trigger, av)
    return (entry + dist) if is_long else (entry - dist)

def _calc_breakeven_trigger(breakeven: str, signal: str, entry: float, av: float) -> Optional[float]:
    if breakeven == "none": return None
    is_long = (signal == "long")
    dist = {"1atr": av, "2atr": 2.0*av, "5pct": entry*0.05, "10pct": entry*0.10}.get(breakeven, av)
    return (entry + dist) if is_long else (entry - dist)


# ─── SIMULACIÓN ───────────────────────────────────────────────────────────────
def _simulate_fast(params: OptParams, ind: dict,
                   return_events: bool = False,
                   fng_override: Optional[float] = None,
                   fund_override: Optional[float] = None) -> dict:
    """Simula un bot en una sola moneda. ind = indicadores precalculados para esa moneda.
    Si return_events=True, devuelve (metrics, events, trades_list) para backtest_engine."""
    all_pnl = 0.0; all_trades = 0; all_wins = 0
    gross_win = 0.0; gross_loss = 0.0
    all_best = 0.0; all_worst = 0.0; total_partial_exits = 0
    equity = INITIAL_EQUITY; peak_equity = INITIAL_EQUITY; max_dd = 0.0
    eq_curve = [INITIAL_EQUITY]
    _evt = [] if return_events else None
    _trl = [] if return_events else None

    direction = params.direction
    min_bars  = max(params.ma_slow + 50, 260)
    ma_key    = f"{params.ma_type}_{params.ma_fast}_{params.ma_slow}"
    sl_type   = params.sl_type; fib_mode = params.fib_mode
    rsi_filt  = params.rsi_filter; e200_filt = params.ema200_filter
    atr_filt  = params.atr_filter; vol_prof = params.vol_profile
    time_filt = params.time_filter; trailing = params.trailing_pct
    compound  = params.compound; risk_lev = params.risk_pct * params.leverage
    macd_filt = params.macd_filter; adx_filt = params.adx_filter
    st_filt   = params.supertrend_filter; ichi_filt = params.ichimoku_filter
    stoch_filt = params.stoch_rsi; cci_filt = params.cci_filter
    wr_filt   = params.williams_r; mom_filt = params.momentum_filter
    bb_filt   = params.bb_filter; atr_vol_f = params.atr_volatility
    kelt_filt = params.keltner_filter; obv_filt = params.obv_filter
    vwap_filt = params.vwap_filter; vdelta_filt = params.volume_delta
    cvd_filt  = params.cvd_filter; ms_filt = params.market_structure
    bo_range  = params.breakout_range; cpat_filt = params.candle_pattern
    ob_filt   = params.order_block; piv_filt = params.pivot_filter
    sr_bo_filt = params.sr_breakout; fib_ret_f = params.fib_retracement
    rsi_div_f = params.rsi_divergence; btc_corr = params.btc_correlation
    fund_filt = params.funding_filter; fng_filt = params.fear_greed_filter
    sess_filt = params.session_filter; pos_siz = params.position_sizing
    max_td    = params.max_trades_day; trail_type = params.trailing_type
    min_conf  = params.min_confluences
    tp_type_p = params.tp_type; tp_pct_p = params.tp_pct; tp_atr_p = params.tp_atr
    trail_act = params.trailing_activation; trail_prog = params.trailing_progressive
    atr_tp_adj = params.atr_tp_adjust; part_close = params.partial_close
    part_trig = params.partial_trigger; be_mode = params.breakeven
    time_exit_p = params.time_exit; sess_exit = params.session_exit
    wknd_exit = params.weekend_exit; rr_min_p = params.rr_min
    rr_min_val = float(rr_min_p) if rr_min_p != "none" else None
    time_exit_n = int(time_exit_p) if time_exit_p != "none" else None
    fng_val = fng_override if fng_override is not None else _FEAR_GREED_IDX
    fund_val = fund_override if fund_override is not None else _FUNDING_RATE
    is_btc = (params.coin == "BTC")

    closes = ind.get("closes")
    _empty = {"total_pnl": 0.0, "total_pnl_pct": 0.0, "win_rate": 0.0, "total_trades": 0,
              "max_drawdown": 0.0, "sharpe": 0.0, "profit_factor": 0.0,
              "best_trade": 0.0, "worst_trade": 0.0, "partial_exits": 0}
    if closes is None or len(closes) < min_bars + 10:
        return (_empty, [], []) if return_events else _empty

    maf = ind.get(f"maf_{ma_key}"); mas = ind.get(f"mas_{ma_key}")
    if maf is None or mas is None:
        return (_empty, [], []) if return_events else _empty

    if time_filt == "london_ny":            hour_ok = ind["h_london"]
    elif time_filt == "asia":               hour_ok = ind["h_asia"]
    elif sess_filt == "asia":               hour_ok = ind["h_asia"]
    elif sess_filt == "london":             hour_ok = ind["h_london"]
    elif sess_filt == "ny":                 hour_ok = ind["h_ny"]
    elif sess_filt == "london_ny_overlap":  hour_ok = ind["h_lno"]
    else:                                   hour_ok = None

    highs = ind["highs"]; lows = ind["lows"]; opens = ind["opens"]; vols = ind["volumes"]
    atr14 = ind["atr14"]; ema200 = ind["ema200"]; rsi14 = ind["rsi14"]; vol_ma = ind["vol_ma"]
    roll_hi = ind["roll_hi"]; roll_lo = ind["roll_lo"]
    atr_pct_ref = ind["atr_pct_ref"]; price_mean = ind["price_mean"]
    atr_mean = ind["atr_mean"]; ts_col = ind["ts"]; psar = ind["psar"]
    N = len(closes)

    coin_equity = INITIAL_EQUITY; position = None
    trades_today = 0; current_day = -1
    kel_wins = 0; kel_n = 0; kel_gross_w = 0.0; kel_gross_l = 0.0
    prev_hour_in_session = False

    for i in range(min_bars, N - 1):
        price = closes[i]
        hour  = int(ts_col[i] // 3_600_000 % 24)
        dow   = int((ts_col[i] // 86_400_000 + 3) % 7)
        if max_td > 0:
            day = int(ts_col[i] // 86_400_000)
            if day != current_day: current_day = day; trades_today = 0
        cur_hour_in_session = (hour_ok is None or bool(hour_ok[i]))

        if position is not None:
            dir_  = position["dir"]; entry = position["entry"]
            stop  = position["stop"]; tp    = position["tp"]
            be_done = position["be_done"]; part_done = position["pdone"]
            frac  = position["frac"]; trig1 = position["trig1"]; trig2 = position["trig2"]
            entry_bar = position["bar"]; trail_act_flag = position["trail_act"]
            av_pos = atr14[i] if atr14[i] == atr14[i] else entry * 0.01

            exit_now = False
            if wknd_exit and dow >= 4 and hour >= 20: exit_now = True
            if not exit_now and sess_exit and prev_hour_in_session and not cur_hour_in_session: exit_now = True
            if not exit_now and time_exit_n is not None and (i - entry_bar) >= time_exit_n: exit_now = True

            if not exit_now:
                if not trail_act_flag:
                    threshold = None
                    if trail_act == "1pct":   threshold = entry * 1.01 if dir_ == "long" else entry * 0.99
                    elif trail_act == "2pct": threshold = entry * 1.02 if dir_ == "long" else entry * 0.98
                    elif trail_act == "5pct": threshold = entry * 1.05 if dir_ == "long" else entry * 0.95
                    elif trail_act == "10pct":threshold = entry * 1.10 if dir_ == "long" else entry * 0.90
                    if threshold is None: trail_act_flag = True
                    elif dir_ == "long"  and price >= threshold: trail_act_flag = True
                    elif dir_ == "short" and price <= threshold: trail_act_flag = True
                    position["trail_act"] = trail_act_flag

                if trail_act_flag:
                    eff_trail = trailing
                    if trail_prog:
                        profit_pct = (price - entry) / entry if dir_ == "long" else (entry - price) / entry
                        if profit_pct > 0.10:   eff_trail = trailing * 0.5
                        elif profit_pct > 0.05: eff_trail = trailing * 0.7
                    if trail_type == "atr_dynamic":
                        ns = (price - 2.0 * av_pos) if dir_ == "long" else (price + 2.0 * av_pos)
                    elif trail_type == "chandelier":
                        w = 22
                        if i >= w:
                            if dir_ == "long":
                                peak = _safe_nanmax(highs[i-w:i+1])
                                ns = (peak - 3.0 * av_pos) if not np.isnan(peak) else stop
                            else:
                                trough = _safe_nanmin(lows[i-w:i+1])
                                ns = (trough + 3.0 * av_pos) if not np.isnan(trough) else stop
                        else: ns = stop
                    elif trail_type == "parabolic_sar":
                        ps_v = psar[i]; ns = ps_v if ps_v == ps_v else stop
                    elif trail_type == "ma_trailing":
                        ns = maf[i] if maf[i] == maf[i] else stop
                    elif sl_type == "trailing":
                        ns = (price * (1.0 - eff_trail) if dir_ == "long" else price * (1.0 + eff_trail))
                    elif sl_type == "atr":
                        ns = (price - 2.0 * av_pos) if dir_ == "long" else (price + 2.0 * av_pos)
                    else: ns = stop
                    if dir_ == "long"  and ns > stop: stop = ns; position["stop"] = stop
                    if dir_ == "short" and ns < stop: stop = ns; position["stop"] = stop

                if not be_done and be_mode != "none":
                    be_trig = position.get("be_trig")
                    if be_trig is not None:
                        if dir_ == "long"  and price >= be_trig:
                            if entry > stop: stop = entry; position["stop"] = stop
                            position["be_done"] = True; be_done = True
                        elif dir_ == "short" and price <= be_trig:
                            if entry < stop: stop = entry; position["stop"] = stop
                            position["be_done"] = True; be_done = True

                if part_close != "none" and part_done < (1 if part_close == "2level" else 2):
                    trig = trig1 if part_done == 0 else trig2
                    if trig is not None:
                        triggered = ((dir_ == "long" and price >= trig) or (dir_ == "short" and price <= trig))
                        if triggered:
                            part_frac = 0.50 if part_close == "2level" else 0.333
                            eff_risk_p = risk_lev
                            size_p = coin_equity * eff_risk_p / entry * part_frac * frac
                            pnl_p  = size_p * (trig - entry) if dir_ == "long" else size_p * (entry - trig)
                            all_pnl += pnl_p; total_partial_exits += 1
                            if _evt is not None: _evt.append((float(ts_col[i]), pnl_p)); _trl.append({"pnl": round(pnl_p, 2)})
                            if compound: coin_equity = max(coin_equity + pnl_p, 1.0)
                            equity += pnl_p
                            if equity > peak_equity: peak_equity = equity
                            elif peak_equity > 0:
                                dd = (peak_equity - equity) / peak_equity * 100.0
                                if dd > max_dd: max_dd = dd
                            eq_curve.append(equity)
                            new_frac = frac * (1.0 - part_frac)
                            position["frac"] = new_frac; frac = new_frac
                            position["pdone"] = part_done + 1; part_done += 1
                            if stop < entry and dir_ == "long": stop = entry; position["stop"] = stop
                            if stop > entry and dir_ == "short": stop = entry; position["stop"] = stop

                if dir_ == "long":  exit_now = price <= stop or (tp is not None and price >= tp)
                else:               exit_now = price >= stop or (tp is not None and price <= tp)
                if not exit_now:
                    pf1 = maf[i-1]; ps1 = mas[i-1]; cf1 = maf[i]; cs1 = mas[i]
                    if pf1==pf1 and ps1==ps1 and cf1==cf1 and cs1==cs1:
                        if dir_ == "long"  and pf1 >= ps1 and cf1 < cs1: exit_now = True
                        elif dir_ == "short" and pf1 <= ps1 and cf1 > cs1: exit_now = True

            if exit_now:
                eff_risk = risk_lev
                if pos_siz == "kelly" and kel_n >= 5:
                    wr_k = kel_wins / kel_n
                    rr_k = ((kel_gross_w / kel_wins) / (kel_gross_l / max(kel_n - kel_wins, 1))
                            if kel_wins > 0 and kel_n > kel_wins else 1.0)
                    f_s  = (wr_k * rr_k - (1.0 - wr_k)) / rr_k if rr_k > 0 else risk_lev
                    eff_risk = max(0.005, min(f_s, 0.25))
                elif pos_siz == "atr_based":
                    av2 = atr14[i] if atr14[i] == atr14[i] and atr14[i] > 0 else entry * 0.01
                    eff_risk = min((params.risk_pct * coin_equity) / (av2 * entry) if entry > 0 else risk_lev, 0.25)
                elif pos_siz == "volatility_adjusted":
                    av2 = atr14[i] if atr14[i] == atr14[i] and atr14[i] > 0 else entry * 0.01
                    atp = av2 / price if price > 0 else atr_pct_ref
                    eff_risk = min(risk_lev * (atr_pct_ref / atp) if atp > 0 else risk_lev, 0.25)
                size = coin_equity * eff_risk / entry * frac
                pnl  = size * (price - entry) if dir_ == "long" else size * (entry - price)
                if compound: coin_equity = max(coin_equity + pnl, 1.0)
                kel_n += 1
                if pnl > 0: kel_wins += 1; kel_gross_w += pnl
                else: kel_gross_l += -pnl
                all_pnl += pnl; all_trades += 1
                if _evt is not None: _evt.append((float(ts_col[i]), pnl)); _trl.append({"pnl": round(pnl, 2)})
                if pnl > 0: all_wins += 1; gross_win += pnl
                else: gross_loss += -pnl
                if pnl > all_best:  all_best  = pnl
                if pnl < all_worst: all_worst = pnl
                equity += pnl
                if equity > peak_equity: peak_equity = equity
                elif peak_equity > 0:
                    dd = (peak_equity - equity) / peak_equity * 100.0
                    if dd > max_dd: max_dd = dd
                eq_curve.append(equity)
                trades_today += 1; position = None
                prev_hour_in_session = cur_hour_in_session; continue

        prev_hour_in_session = cur_hour_in_session
        if max_td > 0 and trades_today >= max_td: continue
        pf1 = maf[i-1]; ps1 = mas[i-1]; cf1 = maf[i]; cs1 = mas[i]
        if not (pf1==pf1 and ps1==ps1 and cf1==cf1 and cs1==cs1): continue

        if direction == "long":
            if not (pf1 < ps1 and cf1 > cs1): continue
            signal = "long"
        elif direction == "short":
            if not (pf1 > ps1 and cf1 < cs1): continue
            signal = "short"
        else:
            if   pf1 < ps1 and cf1 > cs1: signal = "long"
            elif pf1 > ps1 and cf1 < cs1: signal = "short"
            else: continue

        if hour_ok is not None and not hour_ok[i]: continue

        if e200_filt != "none":
            e200 = ema200[i]
            if e200 == e200 and e200 > 0:
                if e200_filt == "soft":
                    if signal == "long"  and price < e200 * 0.98: continue
                    if signal == "short" and price > e200 * 1.02: continue
                else:
                    if signal == "long"  and price <= e200: continue
                    if signal == "short" and price >= e200: continue

        if rsi_filt != "none":
            rv = rsi14[i]
            if rv == rv:
                if rsi_filt == "rsi50":
                    if signal == "long"  and rv <= 50.0: continue
                    if signal == "short" and rv >= 50.0: continue
                else:
                    if signal == "long"  and rv <= 55.0: continue
                    if signal == "short" and rv >= 45.0: continue

        if atr_filt != "none" and atr_pct_ref > 0:
            av = atr14[i]
            if av == av:
                apt = av / price_mean
                if atr_filt == "min" and apt < atr_pct_ref * 0.5: continue
                if atr_filt == "max" and apt > atr_pct_ref * 1.8: continue

        if fib_mode != "disabled":
            hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
            if hi_sw == hi_sw and lo_sw == lo_sw:
                if fib_mode == "required" and not _near_fibo_fast(price, hi_sw, lo_sw): continue

        if vol_prof != "disabled":
            vm = vol_ma[i]
            if vm == vm and vm > 0:
                vn = vols[i]
                if vol_prof == "strict"  and vn < vm * 1.5: continue
                if vol_prof == "relaxed" and vn < vm * 0.7: continue

        if fng_filt != "none" and fng_val is not None:
            if fng_filt == "extreme_fear" and signal == "short": continue
            if fng_filt == "extreme_greed" and signal == "long": continue
            if fng_filt == "both" and fng_val >= 25 and fng_val <= 75: continue

        if fund_filt != "none" and fund_val is not None:
            abs_f = abs(fund_val)
            if fund_filt == "extreme"  and abs_f <= 0.001:  continue
            if fund_filt == "moderate" and abs_f <= 0.0003: continue
            if abs_f > 0:
                if fund_val > 0 and signal == "long":  continue
                if fund_val < 0 and signal == "short": continue

        confluences = 0
        # MACD
        if macd_filt != "none":
            ml = ind["macd_l"][i]; ms_ = ind["macd_s"][i]; mh = ind["macd_h"][i]
            if ml == ml and ms_ == ms_ and mh == mh:
                if macd_filt == "histogram":
                    if (signal == "long" and mh > 0) or (signal == "short" and mh < 0): confluences += 1
                elif macd_filt == "signal_cross":
                    pml = ind["macd_l"][i-1]; pms = ind["macd_s"][i-1]
                    if pml == pml and pms == pms:
                        if signal == "long"  and pml <= pms and ml > ms_: confluences += 1
                        if signal == "short" and pml >= pms and ml < ms_: confluences += 1
                elif macd_filt == "divergence":
                    if signal == "long"  and ind["macd_bull_div"][i]: confluences += 1
                    if signal == "short" and ind["macd_bear_div"][i]: confluences += 1
        # ADX
        if adx_filt != "none":
            av = ind["adx"][i]
            if av == av and av >= float(adx_filt): confluences += 1
        # Supertrend
        if st_filt != "none":
            st_dir = ind["st"][i]
            if (signal == "long" and st_dir > 0) or (signal == "short" and st_dir < 0): confluences += 1
        # Ichimoku
        if ichi_filt != "none":
            ct = ind["cloud_top"][i]; cb = ind["cloud_bot"][i]
            if ct == ct and cb == cb:
                if ichi_filt == "above_cloud":
                    if signal == "long"  and price > ct: confluences += 1
                    if signal == "short" and price < cb: confluences += 1
                elif ichi_filt == "tk_cross":
                    t = ind["tenkan"][i]; k_ = ind["kijun"][i]
                    pt = ind["tenkan"][i-1]; pk = ind["kijun"][i-1]
                    if t==t and k_==k_ and pt==pt and pk==pk:
                        if signal == "long"  and pt <= pk and t > k_: confluences += 1
                        if signal == "short" and pt >= pk and t < k_: confluences += 1
        # Stoch RSI
        if stoch_filt != "none":
            sk_v = ind["sk"][i]; sd_v = ind["sd"][i]
            if sk_v == sk_v and sd_v == sd_v:
                if stoch_filt == "oversold_ob":
                    if signal == "long"  and sk_v < 20: confluences += 1
                    if signal == "short" and sk_v > 80: confluences += 1
                elif stoch_filt == "cross":
                    psk = ind["sk"][i-1]; psd = ind["sd"][i-1]
                    if psk == psk and psd == psd:
                        if signal == "long"  and psk <= psd and sk_v > sd_v: confluences += 1
                        if signal == "short" and psk >= psd and sk_v < sd_v: confluences += 1
        # CCI
        if cci_filt != "none":
            cv = ind["cci"][i]
            if cv == cv:
                thr = float(cci_filt)
                if signal == "long"  and cv > thr:  confluences += 1
                if signal == "short" and cv < -thr: confluences += 1
        # Williams %R
        if wr_filt != "none":
            wv = ind["wr"][i]
            if wv == wv:
                if signal == "long"  and wv < -80: confluences += 1
                if signal == "short" and wv > -20: confluences += 1
        # Momentum
        if mom_filt != "none":
            mv = ind["mom10"][i]
            if mv == mv:
                if mom_filt == "positive":
                    if signal == "long"  and mv > 0: confluences += 1
                    if signal == "short" and mv < 0: confluences += 1
                elif mom_filt == "accelerating":
                    pv = ind["mom10"][i-1]
                    if pv == pv:
                        if signal == "long"  and mv > 0 and mv > pv: confluences += 1
                        if signal == "short" and mv < 0 and mv < pv: confluences += 1
        # BB
        if bb_filt != "none":
            bu = ind["bb_u"][i]; bl_ = ind["bb_l"][i]
            bw = ind["bb_w"][i]; bw_s = ind["bb_w_sma"][i]
            if bu == bu and bl_ == bl_:
                if bb_filt == "breakout":
                    if signal == "long"  and price > bu: confluences += 1
                    if signal == "short" and price < bl_: confluences += 1
                elif bb_filt == "squeeze":
                    if bw == bw and bw_s == bw_s and bw < bw_s * 0.8: confluences += 1
                elif bb_filt == "mean_reversion":
                    if signal == "long"  and price < bl_: confluences += 1
                    if signal == "short" and price > bu:  confluences += 1
        # ATR Volatility
        if atr_vol_f != "none" and atr_mean > 0:
            av = atr14[i]
            if av == av:
                ratio = av / atr_mean
                if atr_vol_f == "low"    and ratio < 0.7:           confluences += 1
                if atr_vol_f == "medium" and 0.7 <= ratio < 1.5:    confluences += 1
                if atr_vol_f == "high"   and ratio >= 1.5:          confluences += 1
        # Keltner
        if kelt_filt != "none":
            ku = ind["kelt_u"][i]; kl = ind["kelt_l"][i]
            if ku == ku and kl == kl:
                if kelt_filt == "breakout":
                    if signal == "long"  and price > ku: confluences += 1
                    if signal == "short" and price < kl: confluences += 1
                elif kelt_filt == "squeeze":
                    bu2 = ind["bb_u"][i]; bl2 = ind["bb_l"][i]
                    if bu2 == bu2 and bl2 == bl2 and bu2 < ku and bl2 > kl: confluences += 1
        # OBV
        if obv_filt != "none":
            ov = ind["obv"][i]; oe = ind["obv_ema"][i]
            if oe == oe:
                if obv_filt == "trend":
                    if signal == "long"  and ov > oe: confluences += 1
                    if signal == "short" and ov < oe: confluences += 1
                elif obv_filt == "divergence":
                    if signal == "long"  and ind["obv_bull_div"][i]: confluences += 1
                    if signal == "short" and ind["obv_bear_div"][i]: confluences += 1
        # VWAP
        if vwap_filt != "none":
            vv = ind["vwap"][i]
            if vv == vv:
                if signal == "long"  and price > vv: confluences += 1
                if signal == "short" and price < vv: confluences += 1
        # Volume Delta
        if vdelta_filt != "none":
            cv2 = ind["cvd"][i]; ce = ind["cvd_ema"][i]
            if ce == ce:
                if vdelta_filt == "confirm":
                    if signal == "long"  and cv2 > ce: confluences += 1
                    if signal == "short" and cv2 < ce: confluences += 1
                elif vdelta_filt == "divergence" and i >= 1:
                    pc2 = closes[i-1]; pcvd = ind["cvd"][i-1]
                    if signal == "long"  and price < pc2 and cv2 > pcvd: confluences += 1
                    if signal == "short" and price > pc2 and cv2 < pcvd: confluences += 1
        # CVD
        if cvd_filt != "none":
            cv3 = ind["cvd"][i]; ce3 = ind["cvd_ema"][i]
            if ce3 == ce3:
                if cvd_filt == "confirm":
                    if signal == "long"  and cv3 > 0: confluences += 1
                    if signal == "short" and cv3 < 0: confluences += 1
                elif cvd_filt == "divergence":
                    if signal == "long"  and cv3 < 0 and price > closes[i-1]: confluences += 1
                    if signal == "short" and cv3 > 0 and price < closes[i-1]: confluences += 1
        # Market Structure
        if ms_filt != "none":
            ms_v = ind["ms"][i]
            if signal == "long"  and ms_v > 0: confluences += 1
            if signal == "short" and ms_v < 0: confluences += 1
        # Breakout Range
        if bo_range != "none":
            n_bo = int(bo_range)
            rh = ind.get(f"roll_hi{n_bo}", roll_hi)[i]; rl = ind.get(f"roll_lo{n_bo}", roll_lo)[i]
            if rh == rh and rl == rl:
                if signal == "long"  and price > rh: confluences += 1
                if signal == "short" and price < rl: confluences += 1
        # Candle Pattern
        if cpat_filt != "none":
            op = opens[i]; hi = highs[i]; lo = lows[i]; cl = closes[i]
            body = abs(cl - op); total = hi - lo
            if cpat_filt == "hammer" and total > 0:
                if min(op, cl) - lo > body * 2 and signal == "long": confluences += 1
            elif cpat_filt == "engulfing" and i >= 1:
                pop = opens[i-1]; pcl = closes[i-1]; pbody = abs(pcl - pop)
                if signal == "long"  and cl > op and cl > pop and op < pcl and body > pbody: confluences += 1
                if signal == "short" and cl < op and cl < pop and op > pcl and body > pbody: confluences += 1
            elif cpat_filt == "doji" and total > 0:
                if body / total < 0.1: confluences += 1
        # Order Block
        if ob_filt != "none":
            if signal == "long"  and ind["ob_bull"][i]: confluences += 1
            if signal == "short" and ind["ob_bear"][i]: confluences += 1
        # Pivot
        if piv_filt != "none":
            s1 = ind["piv_s1" if piv_filt == "daily" else "wpiv_s1"][i]
            r1 = ind["piv_r1" if piv_filt == "daily" else "wpiv_r1"][i]
            if s1 == s1 and r1 == r1:
                if signal == "long"  and price > s1: confluences += 1
                if signal == "short" and price < r1: confluences += 1
        # SR Breakout
        if sr_bo_filt != "none":
            n_sr = int(sr_bo_filt)
            rh_v = ind.get(f"roll_hi{n_sr}" if n_sr <= 100 else "roll_hi", ind.get("roll_hi", roll_hi))
            rl_v = ind.get(f"roll_lo{n_sr}" if n_sr <= 100 else "roll_lo", ind.get("roll_lo", roll_lo))
            if i >= 1 and rh_v[i-1] == rh_v[i-1]:
                if signal == "long"  and price > rh_v[i-1]: confluences += 1
                if signal == "short" and price < rl_v[i-1]: confluences += 1
        # Fib Retracement
        if fib_ret_f != "none":
            hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
            if hi_sw == hi_sw and lo_sw == lo_sw and hi_sw > lo_sw:
                rng = hi_sw - lo_sw; pct = float(fib_ret_f) / 100.0
                level = (lo_sw + rng * pct) if signal == "long" else (hi_sw - rng * pct)
                if abs(price - level) / price <= 0.015: confluences += 1
        # RSI Divergence
        if rsi_div_f != "none":
            bull_d = ind["rsi_bull_div"][i]; bear_d = ind["rsi_bear_div"][i]
            if rsi_div_f in ("required","bonus"):
                if signal == "long"  and bull_d: confluences += 1
                if signal == "short" and bear_d: confluences += 1
                if rsi_div_f == "required":
                    if signal == "long"  and not bull_d: continue
                    if signal == "short" and not bear_d: continue
        # BTC Correlation
        if btc_corr != "none" and not is_btc:
            bt = ind.get("btc_trend")
            if bt is not None and i < len(bt):
                if signal == "long"  and bt[i] > 0: confluences += 1
                if signal == "short" and bt[i] < 0: confluences += 1
                if signal == "long"  and bt[i] < 0: continue
                if signal == "short" and bt[i] > 0: continue

        if confluences < min_conf: continue

        # ── Entrada ──────────────────────────────────────────────────────────
        entry  = price
        av_now = atr14[i] if atr14[i] == atr14[i] else entry * 0.01
        if sl_type == "fixed":
            stop = entry * (1.0 - trailing * 2) if signal == "long" else entry * (1.0 + trailing * 2)
        elif sl_type == "trailing":
            stop = entry * (1.0 - trailing) if signal == "long" else entry * (1.0 + trailing)
        else:
            stop = (entry - 2.0 * av_now) if signal == "long" else (entry + 2.0 * av_now)

        if trail_type == "chandelier":
            w = 22
            if i >= w:
                if signal == "long":
                    peak = _safe_nanmax(highs[i-w:i+1])
                    stop = (peak - 3.0 * av_now) if not np.isnan(peak) else stop
                else:
                    trough = _safe_nanmin(lows[i-w:i+1])
                    stop = (trough + 3.0 * av_now) if not np.isnan(trough) else stop
        elif trail_type == "parabolic_sar":
            ps_v = psar[i]
            if ps_v == ps_v: stop = ps_v
        elif trail_type == "ma_trailing":
            maf_v = maf[i]
            if maf_v == maf_v: stop = maf_v

        tp = _calc_tp(tp_type_p, tp_pct_p, tp_atr_p, signal, entry, stop, av_now, ind, i)
        if tp is not None and atr_tp_adj != "none" and atr_mean > 0:
            ratio = av_now / atr_mean; adj = 1.0
            if atr_tp_adj in ("reduce_high","both") and ratio > 1.5: adj = 0.80
            elif atr_tp_adj in ("extend_low","both") and ratio < 0.7: adj = 1.20
            if adj != 1.0:
                tp = (entry + abs(tp - entry) * adj) if signal == "long" else (entry - abs(tp - entry) * adj)

        if rr_min_val is not None and tp is not None:
            risk_d = abs(entry - stop); reward_d = abs(tp - entry)
            if risk_d > 0 and reward_d / risk_d < rr_min_val: continue

        trig1 = None; trig2 = None
        if part_close != "none":
            trig1 = _calc_partial_trigger(part_trig, signal, entry, stop, av_now)
            if part_close == "3level":
                dist1 = abs(trig1 - entry)
                trig2 = (entry + 2.0 * dist1) if signal == "long" else (entry - 2.0 * dist1)

        be_trig = _calc_breakeven_trigger(be_mode, signal, entry, av_now)
        trail_act_at_entry = (trail_act == "none")
        position = {
            "dir": signal, "entry": entry, "stop": stop, "tp": tp,
            "trig1": trig1, "trig2": trig2, "bar": i,
            "be_done": False, "be_trig": be_trig, "pdone": 0,
            "frac": 1.0, "trail_act": trail_act_at_entry,
        }

    if all_trades == 0:
        _empty0 = {"total_pnl": 0.0, "total_pnl_pct": 0.0, "win_rate": 0.0, "total_trades": 0,
                    "max_drawdown": 0.0, "sharpe": 0.0, "profit_factor": 0.0,
                    "best_trade": 0.0, "worst_trade": 0.0, "partial_exits": 0}
        return (_empty0, [], []) if return_events else _empty0

    win_rate      = all_wins / all_trades * 100.0
    total_pnl_pct = all_pnl / INITIAL_EQUITY * 100.0
    if len(eq_curve) > 2:
        eq_arr  = np.array(eq_curve); returns = np.diff(eq_arr) / np.where(eq_arr[:-1] > 0, eq_arr[:-1], 1.0)
        std_r   = float(np.std(returns))
        sharpe  = float(np.mean(returns)) / std_r * np.sqrt(252) if std_r > 0 else 0.0
    else:
        sharpe = 0.0
    profit_factor = gross_win / gross_loss if gross_loss > 0 else (2.0 if gross_win > 0 else 1.0)
    _result = {
        "total_pnl":     round(all_pnl, 2), "total_pnl_pct": round(total_pnl_pct, 2),
        "win_rate":      round(win_rate, 2), "total_trades":  all_trades,
        "max_drawdown":  round(max_dd, 2),   "sharpe":        round(sharpe, 3),
        "profit_factor": round(profit_factor, 3), "best_trade": round(all_best, 2),
        "worst_trade":   round(all_worst, 2), "partial_exits": total_partial_exits,
    }
    return (_result, _evt, _trl) if return_events else _result


# ─── WORKER FUNCTIONS ─────────────────────────────────────────────────────────
def _worker_run_segment(task: dict) -> List[dict]:
    n = task["n"]; seed = task["seed"]
    rng = random.Random(seed); results: List[dict] = []
    for _ in range(n):
        params = random_params(rng)
        tf_coins = _worker_ind.get(params.interval)
        if not tf_coins: continue
        ind = tf_coins.get(params.coin)
        if not ind: continue
        metrics = _simulate_fast(params, ind)
        if metrics["total_trades"] == 0: continue
        r = asdict(params); r.update(metrics); results.append(r)
    return results

def _worker_ping(_) -> bool: return True

def _worker_run_hc_segment(task: dict) -> List[dict]:
    results: List[dict] = []
    for pd in task.get("params", []):
        try:
            params = OptParams(**pd)
            tf_coins = _worker_ind.get(params.interval)
            if not tf_coins: continue
            ind = tf_coins.get(params.coin)
            if not ind: continue
            metrics = _simulate_fast(params, ind)
            if metrics["total_trades"] == 0: continue
            r = asdict(params); r.update(metrics); results.append(r)
        except Exception:
            continue
    return results


# ─── SCORING Y DEDUPLICACIÓN ──────────────────────────────────────────────────
def _score(r: dict) -> float:
    pnl_pct  = r.get("total_pnl_pct", 0.0)
    win_rate = r.get("win_rate", 0.0)
    max_dd   = max(r.get("max_drawdown", 1.0), 1.0)
    trades   = r.get("total_trades", 0)
    trade_factor = math.log10(max(trades, 1)) / math.log10(10)  # 10t=1.0, 30t=1.48, 79t=1.90
    return (pnl_pct * win_rate * trade_factor) / max_dd

def _key(r: dict) -> str:
    p = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
    return hashlib.md5(json.dumps(p, sort_keys=True).encode()).hexdigest()

def deduplicate(results: List[dict], top_n: int = 30,
                min_trades: int = MIN_TRADES_DASHBOARD) -> List[dict]:
    seen = set(); unique = []
    for r in sorted(results, key=lambda x: _score(x), reverse=True):
        if r.get("total_trades", 0) < min_trades: continue
        k = _key(r)
        if k not in seen:
            seen.add(k); unique.append(r)
            if len(unique) >= top_n: break
    return unique

def _struct_key(r: dict) -> tuple:
    """Clave estructural: (coin, direction, interval, ma_type, leverage).
    Dos configs con misma estructura pero filtros distintos son variantes;
    solo guardamos la mejor."""
    return (r.get("coin",""), r.get("direction",""), r.get("interval",""),
            r.get("ma_type",""), r.get("leverage",0))

def _add(r: dict, struct_best: Dict[tuple, dict], seen_keys: set):
    """Añade resultado si es nuevo o mejor que la variante existente.
    struct_best: dict struct_key → mejor resultado (1 entrada por estructura).
    Exactamente 1 entrada por (coin, direction, interval, ma_type, leverage)."""
    if not r or r.get("total_trades", 0) == 0: return
    k = _key(r)
    if k in seen_keys: return
    seen_keys.add(k)
    sk = _struct_key(r)
    pnl = r.get("total_pnl_pct", 0.0)
    existing = struct_best.get(sk)
    if existing is not None and existing.get("total_pnl_pct", 0.0) >= pnl:
        return  # ya tenemos una variante mejor
    struct_best[sk] = r  # reemplaza la anterior (o inserta nueva)

def _all_results(struct_best: Dict[tuple, dict]) -> List[dict]:
    """Vista: todos los resultados únicos del struct_best."""
    return list(struct_best.values())

def _coin_results(struct_best: Dict[tuple, dict]) -> Dict[str, List[dict]]:
    """Vista: resultados agrupados por moneda."""
    cr: Dict[str, List[dict]] = {}
    for r in struct_best.values():
        coin = r.get("coin", "UNK")
        if coin not in cr: cr[coin] = []
        cr[coin].append(r)
    return cr


# ─── CHECKPOINT ───────────────────────────────────────────────────────────────
def save_checkpoint(path: Path, processed: int,
                    struct_best: Dict[tuple, dict]):
    all_res = _all_results(struct_best)
    coin_res = _coin_results(struct_best)
    data = {
        "processed":     processed,
        "results":       sorted(all_res, key=lambda x: _score(x), reverse=True)[:5000],
        "coin_results":  {c: sorted(r, key=lambda x: _score(x), reverse=True)[:300]
                          for c, r in coin_res.items()},
        "timestamp":     datetime.now().isoformat(),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f)

def load_checkpoint(path: Path) -> Optional[dict]:
    if not path.exists(): return None
    try:
        with open(path, "r", encoding="utf-8") as f: return json.load(f)
    except Exception: return None


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────
def _active_filters(r: dict) -> str:
    parts = []
    for k in _FILTER_FIELDS:
        v = r.get(k, "none")
        if v not in ("none", "disabled", False, 0, "False", "0", ""):
            short = k.replace("_filter","").replace("_"," ")
            parts.append(f"{short}={v}")
    return " | ".join(parts[:10])  # cap at 10 for readability


def save_excel(out_dir: Path, all_results: List[dict],
               coin_results: Dict[str, list], total_tried: int):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"[{ts()}] openpyxl no disponible — saltando Excel"); return

    BG_HDR  = PatternFill("solid", fgColor="0D1F2D")
    FT_HDR  = Font(color="4FC3F7", bold=True, size=8)
    BG_ALT  = PatternFill("solid", fgColor="090F1A")
    BG_TOP  = PatternFill("solid", fgColor="0A2010")
    FT_GLD  = Font(color="FFD700", bold=True, size=8)
    ALIGN_C = Alignment(horizontal="center")

    # Columnas para pestañas de moneda
    COIN_COLS = [
        ("Rank", lambda r, i: i+1),
        ("Direction", lambda r, i: r.get("direction", "?")),
        ("TF", lambda r, i: r.get("interval", "?")),
        ("MA", lambda r, i: f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"),
        ("Lev", lambda r, i: f"{r.get('leverage','?')}x"),
        ("SL", lambda r, i: r.get("sl_type", "?")),
        ("TP", lambda r, i: r.get("tp_type", "none")),
        ("Trailing", lambda r, i: r.get("trailing_type", "none")),
        ("Breakeven", lambda r, i: r.get("breakeven", "none")),
        ("Partial", lambda r, i: r.get("partial_close", "none")),
        ("Filtros activos", lambda r, i: _active_filters(r)),
        ("PnL%", lambda r, i: round(r.get("total_pnl_pct", 0), 2)),
        ("WR%", lambda r, i: round(r.get("win_rate", 0), 1)),
        ("Trades", lambda r, i: r.get("total_trades", 0)),
        ("Sharpe", lambda r, i: round(r.get("sharpe", 0), 3)),
        ("MaxDD%", lambda r, i: round(r.get("max_drawdown", 0), 2)),
    ]

    def write_coin_sheet(ws, rows, title):
        ws.title = title[:31]
        headers = [c[0] for c in COIN_COLS]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = FT_HDR; cell.fill = BG_HDR; cell.alignment = ALIGN_C
        for idx, r in enumerate(rows):
            row_data = [fn(r, idx) for _, fn in COIN_COLS]
            ws.append(row_data)
            fill = BG_TOP if idx < 3 else (BG_ALT if idx % 2 == 0 else None)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
            if idx < 3:
                for cell in ws[ws.max_row]:
                    cell.font = FT_GLD
        col_widths = [6, 10, 6, 14, 6, 10, 16, 14, 10, 10, 60, 8, 7, 7, 8, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Columnas para RANKING_GLOBAL
    GLOBAL_COLS = [("Rank", lambda r, i: i+1), ("Coin", lambda r, i: r.get("coin","?"))] + COIN_COLS[1:]

    def write_global_sheet(ws, rows, title):
        ws.title = title[:31]
        headers = [c[0] for c in GLOBAL_COLS]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = FT_HDR; cell.fill = BG_HDR; cell.alignment = ALIGN_C
        for idx, r in enumerate(rows):
            row_data = [fn(r, idx) for _, fn in GLOBAL_COLS]
            ws.append(row_data)
            fill = BG_TOP if idx < 3 else (BG_ALT if idx % 2 == 0 else None)
            if fill:
                for cell in ws[ws.max_row]: cell.fill = fill
            if idx < 3:
                for cell in ws[ws.max_row]: cell.font = FT_GLD

    wb = Workbook()
    first = True
    for coin in ACTIVE_COINS:
        coin_data = coin_results.get(coin, [])
        qualified = [r for r in coin_data if r.get("total_trades", 0) >= MIN_TRADES_DISPLAY]
        # Deduplicar por (direction, interval, ma_type) para evitar variantes repetidas
        seen_coin: set = set()
        deduped: List[dict] = []
        for r in sorted(qualified, key=lambda x: x.get("total_pnl_pct", 0), reverse=True):
            sig = (r.get("direction",""), r.get("interval",""), r.get("ma_type",""))
            if sig not in seen_coin:
                seen_coin.add(sig); deduped.append(r)
                if len(deduped) >= 30: break
        ws = wb.active if first else wb.create_sheet()
        first = False
        write_coin_sheet(ws, deduped, coin)

    # TOP_PNL: top 100 ordenados por PnL% puro (MIN_TRADES_DISPLAY para incluir SMA)
    top_pnl = sorted([r for r in all_results if r.get("total_trades", 0) >= MIN_TRADES_DISPLAY],
                      key=lambda x: x.get("total_pnl_pct", 0), reverse=True)[:100]
    ws_pnl = wb.create_sheet()
    write_global_sheet(ws_pnl, top_pnl, "TOP_PNL")

    # TOP_CONSISTENCIA: top 100 ordenados por _score (PnL*WR/DD)
    top_consist = sorted([r for r in all_results if r.get("total_trades", 0) >= MIN_TRADES_DISPLAY],
                         key=lambda x: _score(x), reverse=True)[:100]
    ws_consist = wb.create_sheet()
    write_global_sheet(ws_consist, top_consist, "TOP_CONSISTENCIA")

    # RECOMENDADOS_PITCH: top 30 por PnL%, sin repetir (coin, direction, interval)
    # Para demos a clientes — máximo impacto visual
    seen_pitch: set = set()
    pitch: List[dict] = []
    for r in sorted(all_results, key=lambda x: x.get("total_pnl_pct", 0), reverse=True):
        if r.get("total_trades", 0) < MIN_TRADES_DASHBOARD: continue
        cdt = (r.get("coin",""), r.get("direction",""), r.get("interval",""))
        if cdt not in seen_pitch:
            seen_pitch.add(cdt); pitch.append(r)
            if len(pitch) >= 30: break
    ws_pitch = wb.create_sheet()
    write_global_sheet(ws_pitch, pitch, "RECOMENDADOS_PITCH")

    # RECOMENDADOS_LIVE: top 30 por _score, sin repetir (coin, direction, interval)
    # Para uso en producción — balance entre PnL, WR y drawdown
    seen_live: set = set()
    recommended: List[dict] = []
    for r in sorted(all_results, key=lambda x: _score(x), reverse=True):
        if r.get("total_trades", 0) < MIN_TRADES_DASHBOARD: continue
        cdt = (r.get("coin",""), r.get("direction",""), r.get("interval",""))
        if cdt not in seen_live:
            seen_live.add(cdt); recommended.append(r)
            if len(recommended) >= 30: break
    ws_live = wb.create_sheet()
    write_global_sheet(ws_live, recommended, "RECOMENDADOS_LIVE")

    out_path = out_dir / "master_results.xlsx"
    try:
        wb.save(out_path)
        print(f"[{ts()}] Excel: {out_path}")
    except Exception as e:
        print(f"[{ts()}] Error Excel: {e}")

    return recommended


# ─── JSON BOTS ────────────────────────────────────────────────────────────────
def save_master_bots(out_dir: Path, recommended: List[dict]):
    """Genera master_bots.json con los 30 bots listos para sim_engine.py."""
    bots = []
    for i, r in enumerate(recommended[:30]):
        coin = r.get("coin", "BTC")
        direction = r.get("direction", "both")
        interval  = r.get("interval", "1h")
        ma_type   = r.get("ma_type", "ema")
        ma_fast   = r.get("ma_fast", 20)
        ma_slow   = r.get("ma_slow", 50)
        dir_label = {"long": "L", "short": "S", "both": "LS"}.get(direction, "?")
        candle_limit = {"15m": 300, "30m": 250, "1h": 200, "2h": 250, "4h": 300}.get(interval, 200)
        bot = {
            "idx":           i,
            "name":          f"master_{i}_{coin.lower()}_{direction}",
            "label":         f"BOT·{interval.upper()}·{coin}·{dir_label}",
            "interval":      interval,
            "ma_type":       ma_type,
            "ma_fast":       ma_fast,
            "ma_slow":       ma_slow,
            "trailing_pct":  r.get("trailing_pct", 0.01),
            "min_vol_ratio": 1.0,
            "sr_near_pct":   0.01,
            "fibo_zone_pct": 0.012,
            "candle_limit":  candle_limit,
            "leverage":      r.get("leverage", 10),
            "risk_per_trade": r.get("risk_pct", 0.05),
            "sl_type":       r.get("sl_type", "trailing"),
            "direction":     direction,
            "coins":         [coin],
            "require_fib":   False,
            "require_sr":    False,
            "time_filter":   r.get("time_filter", "none"),
            "rsi_filter":    r.get("rsi_filter", "none"),
            "ema200_filter": r.get("ema200_filter", "none"),
            "atr_filter":    r.get("atr_filter", "none"),
            "tp_type":       r.get("tp_type", "none"),
            "trailing_type": r.get("trailing_type", "none"),
            "macd_filter":   r.get("macd_filter", "none"),
            "adx_filter":    r.get("adx_filter", "none"),
            "supertrend_filter": r.get("supertrend_filter", "none"),
            "ichimoku_filter":   r.get("ichimoku_filter", "none"),
            "stoch_rsi":     r.get("stoch_rsi", "none"),
            "cci_filter":    r.get("cci_filter", "none"),
            "williams_r":    r.get("williams_r", "none"),
            "momentum_filter": r.get("momentum_filter", "none"),
            "bb_filter":     r.get("bb_filter", "none"),
            "atr_volatility": r.get("atr_volatility", "none"),
            "keltner_filter": r.get("keltner_filter", "none"),
            "obv_filter":    r.get("obv_filter", "none"),
            "vwap_filter":   r.get("vwap_filter", "none"),
            "volume_delta":  r.get("volume_delta", "none"),
            "cvd_filter":    r.get("cvd_filter", "none"),
            "market_structure": r.get("market_structure", "none"),
            "breakout_range": r.get("breakout_range", "none"),
            "candle_pattern": r.get("candle_pattern", "none"),
            "order_block":   r.get("order_block", "none"),
            "pivot_filter":  r.get("pivot_filter", "none"),
            "sr_breakout":   r.get("sr_breakout", "none"),
            "fib_retracement": r.get("fib_retracement", "none"),
            "rsi_divergence": r.get("rsi_divergence", "none"),
            "btc_correlation": r.get("btc_correlation", "none"),
            "funding_filter": r.get("funding_filter", "none"),
            "fear_greed_filter": r.get("fear_greed_filter", "none"),
            "session_filter": r.get("session_filter", "none"),
            "partial_close": r.get("partial_close", "none"),
            "partial_trigger": r.get("partial_trigger", "1atr"),
            "breakeven":     r.get("breakeven", "none"),
            "min_confluences": r.get("min_confluences", 0),
            "position_sizing": r.get("position_sizing", "fixed"),
            "fib_mode":      r.get("fib_mode", "disabled"),
            "compound":      r.get("compound", True),
            "vol_profile":   r.get("vol_profile", "disabled"),
            "max_trades_day": r.get("max_trades_day", 0),
            "tp_pct":        r.get("tp_pct", 10),
            "tp_atr":        r.get("tp_atr", 2.0),
            "trailing_activation": r.get("trailing_activation", "none"),
            "trailing_progressive": r.get("trailing_progressive", False),
            "atr_tp_adjust": r.get("atr_tp_adjust", "none"),
            "time_exit":     r.get("time_exit", "none"),
            "session_exit":  r.get("session_exit", False),
            "weekend_exit":  r.get("weekend_exit", False),
            "rr_min":        r.get("rr_min", "none"),
            # métricas de referencia
            "_sharpe":       r.get("sharpe", 0),
            "_pnl_pct":      r.get("total_pnl_pct", 0),
            "_win_rate":     r.get("win_rate", 0),
            "_trades":       r.get("total_trades", 0),
            "_max_dd":       r.get("max_drawdown", 0),
        }
        bots.append(bot)
    out_path = out_dir / "master_bots.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(bots, f, indent=2, ensure_ascii=False)
    print(f"[{ts()}] Bots JSON: {out_path}")


# ─── PRINT SUMMARY ────────────────────────────────────────────────────────────
def print_summary(all_results: List[dict], coin_results: Dict[str, list],
                  recommended: List[dict], total_tried: int,
                  duration_sec: float, out_dir: Path):
    sep = "=" * 84
    print(f"\n{sep}")
    print(f"[{ts()}]  OPTIMIZACION MASTER COMPLETADA — {VERSION}")
    print(sep)
    print(f"  Monedas: {', '.join(ACTIVE_COINS)}")
    print(f"  Combinaciones: {total_tried:,}  |  Duracion: {duration_sec/60:.1f} min")
    print(f"\n  RECOMENDADOS_LIVE (top 30 sin duplicar moneda/dir/TF, min {MIN_TRADES_DASHBOARD} trades):")
    print(f"  {'#':>2}  {'Coin':>5}  {'Dir':>6}  {'TF':>4}  {'MA':>14}  "
          f"{'Lev':>4}  {'PnL%':>7}  {'WR%':>5}  {'T':>4}  {'Sharpe':>7}")
    print(f"  {'-'*84}")
    for i, r in enumerate(recommended):
        ma = f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
        print(f"  {i+1:>2}  {r.get('coin','?'):>5}  {r.get('direction','?'):>6}  "
              f"{r.get('interval','?'):>4}  {ma:>14}  "
              f"{r.get('leverage','?'):>3}x  "
              f"{r.get('total_pnl_pct',0):>+6.1f}%  "
              f"{r.get('win_rate',0):>4.1f}%  "
              f"{r.get('total_trades',0):>4}  "
              f"{r.get('sharpe',0):>7.3f}")
    print(f"\n  Salida: {out_dir.resolve()}")
    print(f"{sep}\n")


# ─── EXTERNAL DATA ────────────────────────────────────────────────────────────
def _fetch_fear_greed() -> Optional[float]:
    try:
        r = requests.get(FNG_URL, timeout=10)
        return float(r.json()["data"][0]["value"])
    except Exception as e:
        print(f"[{ts()}] Fear&Greed fetch failed: {e}"); return None

def _fetch_funding_rate() -> Optional[float]:
    try:
        r = requests.get(BINANCE_FUND, params={"symbol": "BTCUSDT", "limit": 1}, timeout=10)
        data = r.json()
        if isinstance(data, list) and data: return float(data[-1].get("fundingRate", 0))
    except Exception as e:
        print(f"[{ts()}] Funding rate fetch failed: {e}")
    return None

def _dl_task(coin: str, tf: str, cache_dir: Path, days: int):
    return coin, tf, load_or_fetch(cache_dir, coin, tf, days)


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    global _FEAR_GREED_IDX, _FUNDING_RATE

    parser = argparse.ArgumentParser(description="AlphaChainBots Optimizer Master")
    parser.add_argument("--days",    type=int, default=365,       help="Dias de historial")
    parser.add_argument("--samples", type=int, default=6_000_000, help="Combinaciones totales")
    parser.add_argument("--out",     type=str, default=None,      help="Directorio de salida")
    parser.add_argument("--workers", type=int, default=None,      help="Workers paralelos")
    parser.add_argument("--resume",  action="store_true",         help="Reanudar desde checkpoint")
    parser.add_argument("--focus-coins", type=str, default=None,
                        help="Monedas especificas separadas por coma (ej: ETH,BTC)")
    args = parser.parse_args()

    # --focus-coins: sobrescribir ACTIVE_COINS global
    global ACTIVE_COINS
    if args.focus_coins:
        fc = [c.strip().upper() for c in args.focus_coins.split(",") if c.strip()]
        invalid = [c for c in fc if c not in MASTER_COINS]
        if invalid:
            print(f"[ERROR] Monedas no válidas: {invalid}. Válidas: {MASTER_COINS}")
            sys.exit(1)
        ACTIVE_COINS = fc
        print(f"[FOCUS] Monedas activas: {', '.join(ACTIVE_COINS)}")

    out_dir = Path(args.out if args.out else "../resultados_master")
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = out_dir / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    ckpt_path = out_dir / "checkpoint_master.json"
    n_workers = args.workers or os.cpu_count() or 4

    sep = "=" * 84
    print(f"\n{sep}")
    print(f"[{ts()}]  AlphaChainBots Optimizer Master — {VERSION}")
    print(sep)
    print(f"  Monedas      : {', '.join(ACTIVE_COINS)}" +
          (f"  (focus mode)" if args.focus_coins else ""))
    print(f"  Direcciones  : {', '.join(MASTER_DIRECTIONS)}")
    print(f"  Historial    : {args.days}d  |  Muestras: {args.samples:,}  |  Workers: {n_workers}")
    print(f"  Salida       : {out_dir.resolve()}")
    print(f"{sep}\n")

    print(f"[{ts()}] Obteniendo Fear & Greed Index…", end=" ", flush=True)
    fng = _fetch_fear_greed(); print(f"{fng:.0f}" if fng is not None else "N/A")

    print(f"[{ts()}] Obteniendo Funding Rate BTC…", end=" ", flush=True)
    funding = _fetch_funding_rate(); print(f"{funding:.5f}" if funding is not None else "N/A")

    processed_start = 0
    struct_best: Dict[tuple, dict] = {}  # struct_key → mejor resultado
    seen_keys: set = set()

    if args.resume:
        ckpt = load_checkpoint(ckpt_path)
        if ckpt:
            processed_start = ckpt.get("processed", 0)
            # Reconstruir struct_best desde checkpoint
            for r in ckpt.get("results", []):
                sk = _struct_key(r)
                existing = struct_best.get(sk)
                if existing is None or r.get("total_pnl_pct", 0) > existing.get("total_pnl_pct", 0):
                    struct_best[sk] = r
                seen_keys.add(_key(r))
            for c in ACTIVE_COINS:
                for r in ckpt.get("coin_results", {}).get(c, []):
                    sk = _struct_key(r)
                    existing = struct_best.get(sk)
                    if existing is None or r.get("total_pnl_pct", 0) > existing.get("total_pnl_pct", 0):
                        struct_best[sk] = r
                    seen_keys.add(_key(r))
            print(f"[{ts()}] Reanudando: {processed_start:,} combinaciones, {len(struct_best)} structs únicos\n")
        else:
            print(f"[{ts()}] No se encontro checkpoint — iniciando desde cero\n")

    # ── Descarga de velas (todas las monedas × todos los TFs) ────────────────
    total_dl = len(ACTIVE_COINS) * len(TIMEFRAMES)
    print(f"[{ts()}] Descargando velas: {len(ACTIVE_COINS)} monedas × {len(TIMEFRAMES)} TF…")
    all_candles: Dict[str, Dict[str, np.ndarray]] = {tf: {} for tf in TIMEFRAMES}
    dl_ok = 0; dl_fail = 0

    with ThreadPoolExecutor(max_workers=min(total_dl, 20)) as ex:
        futures = {ex.submit(_dl_task, coin, tf, cache_dir, args.days): (coin, tf)
                   for tf in TIMEFRAMES for coin in ACTIVE_COINS}
        for fut in as_completed(futures):
            coin_f, tf_key = futures[fut]
            try:
                _, _, arr = fut.result()
            except Exception as e:
                print(f"  [ERR] {coin_f}/{tf_key}: {e}"); dl_fail += 1; continue
            if arr is not None and len(arr) > 50:
                all_candles[tf_key][coin_f] = arr
                print(f"  [OK] {coin_f:6s}/{tf_key:4s}  {len(arr):,} velas"); dl_ok += 1
            else:
                print(f"  [--] {coin_f:6s}/{tf_key:4s}  sin datos"); dl_fail += 1

    print(f"\n[{ts()}] Descarga: {dl_ok} OK / {dl_fail} fallidos\n")

    # ── Worker cache ──────────────────────────────────────────────────────────
    worker_pkl = out_dir / "_wcache_master.pkl"
    with open(worker_pkl, "wb") as f:
        pickle.dump({"candles": all_candles, "_fng": fng, "_funding": funding}, f)

    initargs = (str(worker_pkl),)
    print(f"[{ts()}] Iniciando ProcessPoolExecutor ({n_workers} workers)…", end=" ", flush=True)
    with ProcessPoolExecutor(max_workers=n_workers, initializer=_worker_init,
                             initargs=initargs) as _tp:
        list(_tp.map(_worker_ping, range(n_workers)))
    print(f"listos\n")

    # ── Random search ─────────────────────────────────────────────────────────
    rng        = random.Random(42)
    start_time = time.time()
    processed  = processed_start
    SEG_SIZE   = 1_200
    last_top_print = processed_start; last_ckpt = processed_start

    print(f"[{ts()}] Random search: {args.samples - processed_start:,} combinaciones…")
    with ProcessPoolExecutor(max_workers=n_workers, initializer=_worker_init,
                             initargs=initargs) as pool:
        while processed < args.samples:
            remaining  = args.samples - processed
            n_tasks    = max(1, min(n_workers * 2, remaining // max(SEG_SIZE // 2, 1)))
            actual_seg = max(1, min(SEG_SIZE, remaining // max(n_tasks, 1)))
            tasks = [{"n": actual_seg, "seed": rng.randint(0, 2**32)} for _ in range(n_tasks)]
            try:
                batch = list(pool.map(_worker_run_segment, tasks))
            except Exception as e:
                print(f"\n[{ts()}] Error pool.map: {e} — continuando…")
                processed += n_tasks * actual_seg; continue
            for seg in batch:
                for r in seg: _add(r, struct_best, seen_keys)
            processed += sum(t["n"] for t in tasks)

            elapsed  = max(time.time() - start_time, 0.001)
            done_now = processed - processed_start; speed = done_now / elapsed
            eta_sec  = (args.samples - processed) / speed if speed > 0 else 0
            eta_str  = (f"{eta_sec/3600:.1f}h" if eta_sec >= 3600
                        else f"{eta_sec/60:.1f}min" if eta_sec >= 60 else f"{eta_sec:.0f}s")
            pct = processed / args.samples * 100.0
            print(f"\r[{ts()}] {processed:>10,}/{args.samples:,} ({pct:5.1f}%) "
                  f"| {speed:>6,.0f}/s | ETA:{eta_str:>8} | unicos:{len(struct_best):,}",
                  end="", flush=True)

            if processed - last_top_print >= PRINT_TOP_EVERY:
                last_top_print = processed
                top5 = [r for r in sorted(struct_best.values(), key=lambda x: _score(x), reverse=True)
                        if r.get("total_trades", 0) >= MIN_TRADES_DASHBOARD][:5]
                print(f"\n[{ts()}] Top 5 (Sharpe, min {MIN_TRADES_DASHBOARD} trades):")
                for i, r in enumerate(top5):
                    print(f"  #{i+1}  {r.get('coin','?'):>5}  {r.get('direction','?'):>6}  "
                          f"{r.get('interval','?'):>4}  "
                          f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?'):>3}  "
                          f"PnL:{r.get('total_pnl_pct',0):>+7.1f}%  "
                          f"WR:{r.get('win_rate',0):>5.1f}%  "
                          f"T:{r.get('total_trades',0):>4}  "
                          f"Sharpe:{r.get('sharpe',0):>7.3f}")
                print()

            if processed - last_ckpt >= CHECKPOINT_EVERY:
                last_ckpt = processed
                save_checkpoint(ckpt_path, processed, struct_best)
                print(f"\n[{ts()}] Checkpoint ({processed:,})\n")

    print(f"\n\n[{ts()}] Random search completado.")

    # ── Hill climbing ─────────────────────────────────────────────────────────
    qualified = [r for r in struct_best.values() if r.get("total_trades", 0) >= MIN_TRADES_DASHBOARD]
    if qualified:
        sorted_q = sorted(qualified, key=lambda x: _score(x), reverse=True)
        n_seeds  = max(1, min(len(sorted_q) // 5, 2000))
        n_hc     = min(n_seeds * 8, 50_000)
        print(f"[{ts()}] Hill climbing: {n_seeds} semillas x 8 = {n_hc:,} intentos…")
        hc_batch: List[dict] = []
        for seed in sorted_q[:n_hc // 8]:
            sp = OptParams(**{k: seed[k] for k in _PARAM_FIELDS_SET if k in seed})
            for _ in range(8): hc_batch.append(asdict(perturb(sp, rng)))
        chunk_sz = max(1, len(hc_batch) // n_workers)
        hc_tasks = [{"params": hc_batch[i:i+chunk_sz]} for i in range(0, len(hc_batch), chunk_sz)]
        with ProcessPoolExecutor(max_workers=n_workers, initializer=_worker_init,
                                 initargs=initargs) as pool:
            try:
                for seg in pool.map(_worker_run_hc_segment, hc_tasks):
                    for r in seg: _add(r, struct_best, seen_keys)
                processed += len(hc_batch)
            except Exception as e:
                print(f"\n[{ts()}] Error HC: {e}")
        print(f"[{ts()}] Hill climbing completado.")

    # ── Outputs ───────────────────────────────────────────────────────────────
    print(f"\n[{ts()}] Generando salidas…")
    save_checkpoint(ckpt_path, processed, struct_best)
    all_results = _all_results(struct_best)
    coin_results = _coin_results(struct_best)
    recommended = save_excel(out_dir, all_results, coin_results, processed) or []
    save_master_bots(out_dir, recommended)

    duration = time.time() - start_time
    print_summary(all_results, coin_results, recommended, processed, duration, out_dir)

    try: worker_pkl.unlink()
    except Exception: pass


if __name__ == "__main__":
    mp.freeze_support()
    main()
