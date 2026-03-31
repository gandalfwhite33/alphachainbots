#!/usr/bin/env python3
"""
optimizer_v3.py — Optimizador de parámetros v3 para AlphaChainBots
Máximo número de combinaciones únicas, resultados por timeframe, variedad de indicadores.

Uso:
    python optimizer_v3.py [--days 365] [--samples 500000] [--out ./resultados]
                           [--workers N] [--resume]
"""

import os
import sys
import json
import time
import random
import argparse
import hashlib
import pickle
import platform
import traceback
import multiprocessing as mp
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict, fields
from datetime import datetime
from typing import Optional, List, Dict, Tuple
from pathlib import Path

import numpy as np
import requests

# ─── VERSIÓN Y CONSTANTES ─────────────────────────────────────────────────────
VERSION        = "3.0.0"
INITIAL_EQUITY = 10_000.0
HL_URL         = "https://api.hyperliquid.xyz/info"

COINS      = ["BTC", "ETH", "SOL", "ARB", "OP", "AVAX", "DOGE", "WIF", "PEPE", "SUI"]
TIMEFRAMES = ["15m", "30m", "1h", "2h", "4h"]

INTERVAL_MS = {
    "15m": 900_000, "30m": 1_800_000, "1h": 3_600_000,
    "2h": 7_200_000, "4h": 14_400_000,
}

CHECKPOINT_EVERY = 20_000
PRINT_TOP_EVERY  = 25_000

# ─── OPCIONES DE PARÁMETROS ───────────────────────────────────────────────────
MA_PAIRS = [
    ("ema", 8,  21),
    ("ema", 13, 34),
    ("ema", 21, 55),
    ("ema", 20, 50),
    ("sma", 50, 100),
    ("sma", 100, 200),
]
LEVERAGES       = [2, 3, 5, 10, 15]
TRAILING_PCTS   = [0.003, 0.005, 0.008, 0.010, 0.015, 0.020, 0.030]
SL_TYPES        = ["fixed", "trailing", "atr"]
FIB_MODES       = ["required", "optional", "disabled"]
RSI_FILTERS     = ["none", "rsi50", "rsi55"]
EMA200_FILTERS  = ["none", "soft", "strict"]
ATR_FILTERS     = ["none", "min", "max"]
COMPOUNDS       = [True, False]
TP_FIBS         = [True, False]
TIME_FILTERS    = ["none", "london_ny", "asia"]
VOL_PROFILES    = ["disabled", "strict", "relaxed"]
LIQ_CONFIRMS    = [True, False]
RISK_PCTS       = [0.01, 0.02, 0.03, 0.05]

TOTAL_COMBOS = (
    len(TIMEFRAMES) * len(MA_PAIRS) * len(LEVERAGES) * len(TRAILING_PCTS) *
    len(SL_TYPES) * len(FIB_MODES) * len(RSI_FILTERS) * len(EMA200_FILTERS) *
    len(ATR_FILTERS) * len(COMPOUNDS) * len(TP_FIBS) * len(TIME_FILTERS) *
    len(VOL_PROFILES) * len(LIQ_CONFIRMS) * len(RISK_PCTS)
)

_PARAM_FIELDS_SET = {
    "interval", "ma_type", "ma_fast", "ma_slow", "leverage", "trailing_pct",
    "sl_type", "fib_mode", "rsi_filter", "ema200_filter", "atr_filter",
    "compound", "tp_fib", "time_filter", "vol_profile", "liq_confirm", "risk_pct",
}
_METRIC_FIELDS = {
    "total_pnl", "total_pnl_pct", "win_rate", "total_trades",
    "max_drawdown", "sharpe", "profit_factor", "best_trade", "worst_trade",
}


def ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


# ─── DATACLASS DE PARÁMETROS ──────────────────────────────────────────────────
@dataclass
class OptParams:
    interval:      str
    ma_type:       str
    ma_fast:       int
    ma_slow:       int
    leverage:      float
    trailing_pct:  float
    sl_type:       str
    fib_mode:      str
    rsi_filter:    str
    ema200_filter: str
    atr_filter:    str
    compound:      bool
    tp_fib:        bool
    time_filter:   str
    vol_profile:   str
    liq_confirm:   bool
    risk_pct:      float

    def param_key(self) -> str:
        d = {k: v for k, v in asdict(self).items()}
        return hashlib.md5(json.dumps(d, sort_keys=True).encode()).hexdigest()


def random_params(rng: random.Random) -> OptParams:
    ma = rng.choice(MA_PAIRS)
    return OptParams(
        interval      = rng.choice(TIMEFRAMES),
        ma_type       = ma[0],
        ma_fast       = ma[1],
        ma_slow       = ma[2],
        leverage      = rng.choice(LEVERAGES),
        trailing_pct  = rng.choice(TRAILING_PCTS),
        sl_type       = rng.choice(SL_TYPES),
        fib_mode      = rng.choice(FIB_MODES),
        rsi_filter    = rng.choice(RSI_FILTERS),
        ema200_filter = rng.choice(EMA200_FILTERS),
        atr_filter    = rng.choice(ATR_FILTERS),
        compound      = rng.choice(COMPOUNDS),
        tp_fib        = rng.choice(TP_FIBS),
        time_filter   = rng.choice(TIME_FILTERS),
        vol_profile   = rng.choice(VOL_PROFILES),
        liq_confirm   = rng.choice(LIQ_CONFIRMS),
        risk_pct      = rng.choice(RISK_PCTS),
    )


def perturb(p: OptParams, rng: random.Random) -> OptParams:
    """Hill climbing: perturba un parámetro aleatorio."""
    d = asdict(p)
    choosable = [
        "interval", "leverage", "trailing_pct", "sl_type", "fib_mode",
        "rsi_filter", "ema200_filter", "atr_filter", "compound", "tp_fib",
        "time_filter", "vol_profile", "liq_confirm", "risk_pct", "ma_pair",
    ]
    key = rng.choice(choosable)
    if key == "ma_pair":
        ma = rng.choice(MA_PAIRS)
        d["ma_type"] = ma[0]; d["ma_fast"] = ma[1]; d["ma_slow"] = ma[2]
    else:
        opts = {
            "interval":      TIMEFRAMES,
            "leverage":      LEVERAGES,
            "trailing_pct":  TRAILING_PCTS,
            "sl_type":       SL_TYPES,
            "fib_mode":      FIB_MODES,
            "rsi_filter":    RSI_FILTERS,
            "ema200_filter": EMA200_FILTERS,
            "atr_filter":    ATR_FILTERS,
            "compound":      COMPOUNDS,
            "tp_fib":        TP_FIBS,
            "time_filter":   TIME_FILTERS,
            "vol_profile":   VOL_PROFILES,
            "liq_confirm":   LIQ_CONFIRMS,
            "risk_pct":      RISK_PCTS,
        }
        if key in opts:
            d[key] = rng.choice(opts[key])
    return OptParams(**d)


# ─── CANDLE FETCHING ──────────────────────────────────────────────────────────
def _fetch_hl(coin: str, interval: str, days: int) -> Optional[np.ndarray]:
    end_ms   = int(time.time() * 1000)
    start_ms = end_ms - days * 86_400_000
    try:
        r = requests.post(HL_URL, json={
            "type": "candleSnapshot",
            "req":  {"coin": coin, "interval": interval,
                     "startTime": start_ms, "endTime": end_ms},
        }, timeout=30)
        if not r.ok:
            return None
        candles = r.json()
        if not candles:
            return None
        return np.array([
            [c["T"], float(c["o"]), float(c["h"]), float(c["l"]),
             float(c["c"]), float(c.get("v", 0))]
            for c in candles
        ], dtype=np.float64)
    except Exception as e:
        print(f"[{ts()}]   ⚠ Error {coin}/{interval}: {e}")
        return None


def load_or_fetch(cache_dir: Path, coin: str, interval: str, days: int) -> Optional[np.ndarray]:
    cache_file = cache_dir / f"{coin}_{interval}_{days}d.pkl"
    if cache_file.exists():
        try:
            with open(cache_file, "rb") as f:
                data = pickle.load(f)
            if data is not None and len(data) > 50:
                return data
        except Exception:
            pass
    # Pequeña pausa aleatoria para evitar rate-limit 429 en descargas paralelas
    time.sleep(random.uniform(0.05, 0.25))
    arr = _fetch_hl(coin, interval, days)
    if arr is not None and len(arr) > 50:
        try:
            with open(cache_file, "wb") as f:
                pickle.dump(arr, f)
        except Exception:
            pass
    return arr


# ─── INDICADORES VECTORIZADOS ─────────────────────────────────────────────────
def _ema(arr: np.ndarray, n: int) -> np.ndarray:
    """EMA vectorizado — asume que arr no tiene NaN (precios de cierre limpios)."""
    if len(arr) < n:
        return np.full_like(arr, np.nan)
    out = np.empty_like(arr)
    out[:n - 1] = np.nan
    out[n - 1]  = float(np.mean(arr[:n]))
    k = 2.0 / (n + 1)
    k1 = 1.0 - k
    # Bucle Cython-friendly: sin branches dentro del loop
    prev = out[n - 1]
    for i in range(n, len(arr)):
        prev = arr[i] * k + prev * k1
        out[i] = prev
    return out


def _sma(arr: np.ndarray, n: int) -> np.ndarray:
    """SMA vectorizado con cumsum — O(n), sin bucle Python."""
    if len(arr) < n:
        return np.full_like(arr, np.nan)
    out = np.full_like(arr, np.nan)
    cs = np.cumsum(arr)
    # out[n-1] = mean(arr[0:n]), out[i] = (cs[i] - cs[i-n]) / n
    out[n - 1] = cs[n - 1] / n
    out[n:]    = (cs[n:] - cs[: len(arr) - n]) / n
    return out


def _rsi(closes: np.ndarray, n: int = 14) -> np.ndarray:
    out = np.full_like(closes, np.nan)
    if len(closes) < n + 1:
        return out
    d     = np.diff(closes)
    gains = np.maximum(d, 0.0)
    losses = np.maximum(-d, 0.0)
    avg_g = float(np.mean(gains[:n]))
    avg_l = float(np.mean(losses[:n]))
    for i in range(n, len(d)):
        avg_g = (avg_g * (n - 1) + gains[i]) / n
        avg_l = (avg_l * (n - 1) + losses[i]) / n
        rs    = avg_g / avg_l if avg_l > 0 else 100.0
        out[i + 1] = 100.0 - 100.0 / (1.0 + rs)
    return out


def _atr(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int = 14) -> np.ndarray:
    out = np.full_like(close, np.nan)
    if len(close) < n + 1:
        return out
    tr = np.maximum(
        high[1:] - low[1:],
        np.maximum(np.abs(high[1:] - close[:-1]), np.abs(low[1:] - close[:-1])),
    )
    out[n] = float(np.mean(tr[:n]))
    for i in range(n, len(tr)):
        out[i + 1] = (out[i] * (n - 1) + tr[i]) / n
    return out


# ─── UTILIDADES DE PRE-CÓMPUTO ────────────────────────────────────────────────
_LONDON_NY = set(range(7, 22))
_ASIA      = set(range(0, 9))


def _hour_mask(ts_arr: np.ndarray, allowed: set) -> np.ndarray:
    """Máscara booleana: True si la hora UTC de esa vela está en `allowed`."""
    hours = (ts_arr // 3_600_000 % 24).astype(np.int32)
    mask  = np.zeros(len(ts_arr), dtype=bool)
    for h in allowed:
        mask |= (hours == h)
    return mask


def _roll_max(arr: np.ndarray, w: int) -> np.ndarray:
    """Rolling max de ventana w — vectorizado con stride tricks."""
    out = np.full_like(arr, np.nan)
    if len(arr) < w:
        return out
    shape   = arr.shape[:-1] + (arr.shape[-1] - w + 1, w)
    strides = arr.strides + (arr.strides[-1],)
    windows = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w - 1:] = windows.max(axis=1)
    return out


def _roll_min(arr: np.ndarray, w: int) -> np.ndarray:
    out = np.full_like(arr, np.nan)
    if len(arr) < w:
        return out
    shape   = arr.shape[:-1] + (arr.shape[-1] - w + 1, w)
    strides = arr.strides + (arr.strides[-1],)
    windows = np.lib.stride_tricks.as_strided(arr, shape=shape, strides=strides)
    out[w - 1:] = windows.min(axis=1)
    return out


# ─── FIBONACCI ────────────────────────────────────────────────────────────────
_FIBO_L = np.array([0.236, 0.382, 0.500, 0.618, 0.786])


def _near_fibo_fast(price: float, hi: float, lo: float) -> bool:
    rng = hi - lo
    if rng <= 0 or price <= 0:
        return False
    levels = lo + rng * _FIBO_L
    return bool(np.any(np.abs(levels - price) / price <= 0.015))


# ─── SIMULACIÓN FAST (usa indicadores pre-computados) ────────────────────────
def _simulate_fast(params: OptParams, coins_ind: Dict[str, dict]) -> dict:
    """Simula usando indicadores pre-computados — sin recalcular nada por eval."""
    all_pnl     = 0.0
    all_trades  = 0
    all_wins    = 0
    gross_win   = 0.0
    gross_loss  = 0.0
    all_best    = 0.0
    all_worst   = 0.0
    equity      = INITIAL_EQUITY
    peak_equity = INITIAL_EQUITY
    max_dd      = 0.0
    eq_curve    = [INITIAL_EQUITY]

    min_bars  = max(params.ma_slow + 20, 220)
    ma_key    = f"{params.ma_type}_{params.ma_fast}_{params.ma_slow}"
    sl_type   = params.sl_type
    fib_mode  = params.fib_mode
    rsi_filt  = params.rsi_filter
    e200_filt = params.ema200_filter
    atr_filt  = params.atr_filter
    vol_prof  = params.vol_profile
    time_filt = params.time_filter
    trailing  = params.trailing_pct
    compound  = params.compound
    tp_fib    = params.tp_fib
    risk_lev  = params.risk_pct * params.leverage

    for coin, ind in coins_ind.items():
        closes  = ind.get("closes")
        if closes is None or len(closes) < min_bars + 10:
            continue

        maf = ind.get(f"maf_{ma_key}")
        mas = ind.get(f"mas_{ma_key}")
        if maf is None or mas is None:
            continue

        ema200      = ind["ema200"]
        rsi14       = ind["rsi14"]
        atr14       = ind["atr14"]
        vol_ma      = ind["vol_ma"]
        highs       = ind["highs"]
        lows        = ind["lows"]
        volumes     = ind["volumes"]
        roll_hi     = ind["roll_hi"]
        roll_lo     = ind["roll_lo"]
        atr_pct_ref = ind["atr_pct_ref"]
        price_mean  = ind["price_mean"]
        N           = len(closes)

        if time_filt == "london_ny":
            hour_ok = ind["h_london"]
        elif time_filt == "asia":
            hour_ok = ind["h_asia"]
        else:
            hour_ok = None

        coin_equity = INITIAL_EQUITY
        position    = None  # (direction, entry, stop, tp)

        for i in range(min_bars, N - 1):
            price = closes[i]

            # ── Gestión posición existente ────────────────────────────────────
            if position is not None:
                direction, entry, stop, tp = position

                if sl_type == "trailing":
                    if direction == "long":
                        ns = price * (1.0 - trailing)
                        if ns > stop: stop = ns
                    else:
                        ns = price * (1.0 + trailing)
                        if ns < stop: stop = ns
                    position = (direction, entry, stop, tp)
                elif sl_type == "atr":
                    av = atr14[i]
                    if av == av:   # not NaN
                        if direction == "long":
                            ns = price - 2.0 * av
                            if ns > stop: stop = ns
                        else:
                            ns = price + 2.0 * av
                            if ns < stop: stop = ns
                        position = (direction, entry, stop, tp)

                exit_now = False
                if direction == "long":
                    exit_now = price <= stop or (tp is not None and price >= tp)
                else:
                    exit_now = price >= stop or (tp is not None and price <= tp)

                if not exit_now:
                    pf1 = maf[i-1]; ps1 = mas[i-1]; cf1 = maf[i]; cs1 = mas[i]
                    if pf1==pf1 and ps1==ps1 and cf1==cf1 and cs1==cs1:
                        if direction == "long"  and pf1 >= ps1 and cf1 < cs1: exit_now = True
                        elif direction == "short" and pf1 <= ps1 and cf1 > cs1: exit_now = True

                if exit_now:
                    size = coin_equity * risk_lev / entry
                    pnl  = size * (price - entry) if direction == "long" else size * (entry - price)
                    if compound: coin_equity = max(coin_equity + pnl, 1.0)
                    all_pnl    += pnl
                    all_trades += 1
                    if pnl > 0:  all_wins += 1; gross_win  += pnl
                    else:        gross_loss += -pnl
                    if pnl > all_best:  all_best  = pnl
                    if pnl < all_worst: all_worst = pnl
                    equity += pnl
                    if equity > peak_equity:
                        peak_equity = equity
                    elif peak_equity > 0:
                        dd = (peak_equity - equity) / peak_equity * 100.0
                        if dd > max_dd: max_dd = dd
                    eq_curve.append(equity)
                    position = None
                    continue

            # ── Búsqueda de entrada ───────────────────────────────────────────
            pf1 = maf[i-1]; ps1 = mas[i-1]; cf1 = maf[i]; cs1 = mas[i]
            if not (pf1==pf1 and ps1==ps1 and cf1==cf1 and cs1==cs1):
                continue
            if   pf1 < ps1 and cf1 > cs1: signal = "long"
            elif pf1 > ps1 and cf1 < cs1: signal = "short"
            else: continue

            if hour_ok is not None and not hour_ok[i]: continue

            if e200_filt != "none":
                e200 = ema200[i]
                if e200 == e200 and e200 > 0:
                    if e200_filt == "soft":
                        if signal == "long"  and price < e200 * 0.98: continue
                        if signal == "short" and price > e200 * 1.02: continue
                    else:
                        if signal == "long"  and price <= e200: continue
                        if signal == "short" and price >= e200: continue

            if rsi_filt != "none":
                rv = rsi14[i]
                if rv == rv:
                    if rsi_filt == "rsi50":
                        if signal == "long"  and rv <= 50.0: continue
                        if signal == "short" and rv >= 50.0: continue
                    else:
                        if signal == "long"  and rv <= 55.0: continue
                        if signal == "short" and rv >= 45.0: continue

            if atr_filt != "none" and atr_pct_ref > 0:
                av = atr14[i]
                if av == av:
                    apt = av / price_mean
                    if atr_filt == "min" and apt < atr_pct_ref * 0.5: continue
                    if atr_filt == "max" and apt > atr_pct_ref * 1.8: continue

            if fib_mode != "disabled":
                hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
                if hi_sw == hi_sw and lo_sw == lo_sw:
                    if fib_mode == "required" and not _near_fibo_fast(price, hi_sw, lo_sw):
                        continue

            if vol_prof != "disabled":
                vm = vol_ma[i]
                if vm == vm and vm > 0:
                    vn = volumes[i]
                    if vol_prof == "strict"  and vn < vm * 1.5: continue
                    if vol_prof == "relaxed" and vn < vm * 0.7: continue

            entry  = price
            av_now = atr14[i] if atr14[i] == atr14[i] else entry * 0.01
            if sl_type == "fixed":
                sp   = trailing * 2.0
                stop = entry * (1.0 - sp) if signal == "long" else entry * (1.0 + sp)
            elif sl_type == "trailing":
                stop = entry * (1.0 - trailing) if signal == "long" else entry * (1.0 + trailing)
            else:
                stop = (entry - 2.0 * av_now) if signal == "long" else (entry + 2.0 * av_now)

            tp = None
            if tp_fib:
                hi_sw = roll_hi[i]; lo_sw = roll_lo[i]
                if hi_sw == hi_sw and lo_sw == lo_sw:
                    rng_sw = hi_sw - lo_sw
                    if rng_sw > 0:
                        if signal == "long":
                            cands = [lo_sw + rng_sw * f for f in (0.618, 0.786, 1.0)
                                     if lo_sw + rng_sw * f > entry]
                            tp = min(cands) if cands else None
                        else:
                            cands = [lo_sw + rng_sw * f for f in (0.0, 0.236, 0.382)
                                     if lo_sw + rng_sw * f < entry]
                            tp = max(cands) if cands else None

            position = (signal, entry, stop, tp)

    # ── Métricas ──────────────────────────────────────────────────────────────
    if all_trades == 0:
        return {
            "total_pnl": 0.0, "total_pnl_pct": 0.0, "win_rate": 0.0,
            "total_trades": 0, "max_drawdown": 0.0, "sharpe": 0.0,
            "profit_factor": 0.0, "best_trade": 0.0, "worst_trade": 0.0,
        }

    win_rate      = all_wins / all_trades * 100.0
    total_pnl_pct = all_pnl / INITIAL_EQUITY * 100.0

    if len(eq_curve) > 2:
        eq_arr  = np.array(eq_curve)
        returns = np.diff(eq_arr) / np.where(eq_arr[:-1] > 0, eq_arr[:-1], 1.0)
        std_r   = float(np.std(returns))
        sharpe  = float(np.mean(returns)) / std_r * np.sqrt(252) if std_r > 0 else 0.0
    else:
        sharpe = 0.0

    profit_factor = gross_win / gross_loss if gross_loss > 0 else (2.0 if gross_win > 0 else 1.0)

    return {
        "total_pnl":     round(all_pnl, 2),
        "total_pnl_pct": round(total_pnl_pct, 2),
        "win_rate":      round(win_rate, 2),
        "total_trades":  all_trades,
        "max_drawdown":  round(max_dd, 2),
        "sharpe":        round(sharpe, 3),
        "profit_factor": round(profit_factor, 3),
        "best_trade":    round(all_best, 2),
        "worst_trade":   round(all_worst, 2),
    }


# ─── WORKER (multiprocessing) ─────────────────────────────────────────────────
# Cada worker pre-computa TODOS los indicadores al iniciar.
# Así la evaluación de cada combinación sólo hace el loop de simulación,
# sin recalcular EMA/SMA/RSI/ATR en cada llamada.
_worker_ind: Dict[str, Dict[str, dict]] = {}   # [interval][coin] → {arrays pre-computados}


def _worker_init(cache_pkl: str):
    global _worker_ind
    try:
        with open(cache_pkl, "rb") as f:
            raw: Dict[str, Dict[str, np.ndarray]] = pickle.load(f)
    except Exception:
        _worker_ind = {}
        return

    _worker_ind = {}
    for tf, coins_dict in raw.items():
        _worker_ind[tf] = {}
        for coin, arr in coins_dict.items():
            if arr is None or len(arr) < 250:
                continue
            ts_col  = arr[:, 0]
            highs   = arr[:, 2]
            lows    = arr[:, 3]
            closes  = arr[:, 4]
            volumes = arr[:, 5]

            ind: dict = {
                "ts":      ts_col,
                "closes":  closes,
                "highs":   highs,
                "lows":    lows,
                "volumes": volumes,
                "ema200":  _ema(closes, 200),
                "rsi14":   _rsi(closes, 14),
                "atr14":   _atr(highs, lows, closes, 14),
                "vol_ma":  _sma(volumes, 20),
                # rolling 60-bar max/min para Fibonacci (pre-computado)
                "roll_hi": _roll_max(highs, 60),
                "roll_lo": _roll_min(lows, 60),
                # hour mask por filtro (índice booleano)
                "h_london": _hour_mask(ts_col, _LONDON_NY),
                "h_asia":   _hour_mask(ts_col, _ASIA),
            }
            # Pre-computar las 6 variantes de MA (solo cambia ma_fast/ma_slow)
            for (ma_type, fast, slow) in MA_PAIRS:
                k = f"{ma_type}_{fast}_{slow}"
                if ma_type == "ema":
                    ind[f"maf_{k}"] = _ema(closes, fast)
                    ind[f"mas_{k}"] = _ema(closes, slow)
                else:
                    ind[f"maf_{k}"] = _sma(closes, fast)
                    ind[f"mas_{k}"] = _sma(closes, slow)

            # Referencia ATR% para filtro ATR
            atr_valid = ind["atr14"][~np.isnan(ind["atr14"])]
            price_mean = float(np.mean(closes)) if len(closes) > 0 else 1.0
            ind["atr_pct_ref"] = (float(np.mean(atr_valid)) / price_mean
                                  if len(atr_valid) > 0 and price_mean > 0 else 0.0)
            ind["price_mean"] = price_mean

            _worker_ind[tf][coin] = ind


def _worker_run_segment(task: dict) -> List[dict]:
    """
    Genera `n` combinaciones aleatorias, las simula y devuelve los resultados.
    El worker usa sus propios indicadores pre-computados.
    Sin IPC por combinación — todo el trabajo ocurre dentro del worker.
    """
    n    = task["n"]
    seed = task["seed"]
    rng  = random.Random(seed)
    results: List[dict] = []

    for _ in range(n):
        params = random_params(rng)
        coins_ind = _worker_ind.get(params.interval)
        if not coins_ind:
            continue
        metrics = _simulate_fast(params, coins_ind)
        if metrics["total_trades"] == 0:
            continue
        r = asdict(params)
        r.update(metrics)
        results.append(r)

    return results


def _worker_ping(_) -> bool:
    """Función trivial picklable — se usa para verificar que el Pool arranca."""
    return True


def _worker_run_hc_segment(task: dict) -> List[dict]:
    """Evalúa una lista de params pre-generados (hill climbing)."""
    results: List[dict] = []
    for pd in task.get("params", []):
        try:
            params = OptParams(**pd)
            coins_ind = _worker_ind.get(params.interval)
            if not coins_ind:
                continue
            metrics = _simulate_fast(params, coins_ind)
            if metrics["total_trades"] == 0:
                continue
            r = asdict(params)
            r.update(metrics)
            results.append(r)
        except Exception:
            continue
    return results


# ─── CHECKPOINT ───────────────────────────────────────────────────────────────
def save_checkpoint(path: Path, processed: int,
                    results: List[dict], tf_results: Dict[str, List[dict]]):
    # Guardar los mejores 1000 por score combinado (no por orden de inserción)
    data = {
        "processed":  processed,
        "results":    sorted(results, key=_score, reverse=True)[:1000],
        "tf_results": {
            tf: sorted(r, key=_score, reverse=True)[:200]
            for tf, r in tf_results.items()
        },
        "timestamp":  datetime.now().isoformat(),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f)


def load_checkpoint(path: Path) -> Optional[dict]:
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


# ─── MÉTRICA DE RANKING ───────────────────────────────────────────────────────
def _score(r: dict) -> float:
    """
    Métrica combinada de ranking:
        score = (PnL_porcentual * WinRate) / MaxDrawdown
    - PnL_porcentual: rendimiento sobre capital inicial (%)
    - WinRate: % de trades ganadores (0-100)
    - MaxDrawdown: drawdown máximo en % (mínimo 1 para evitar div/0)
    """
    pnl_pct  = r.get("total_pnl_pct", 0.0)
    win_rate = r.get("win_rate", 0.0)
    max_dd   = max(r.get("max_drawdown", 1.0), 1.0)
    return (pnl_pct * win_rate) / max_dd


# ─── DEDUPLICACIÓN ESTRICTA ───────────────────────────────────────────────────
def deduplicate(results: List[dict], top_n: int = 30) -> List[dict]:
    """Retorna top_n resultados con parámetros estrictamente únicos, ordenados por score descendente."""
    seen  = set()
    unique = []
    for r in sorted(results, key=_score, reverse=True):
        param_dict = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
        key = hashlib.md5(json.dumps(param_dict, sort_keys=True).encode()).hexdigest()
        if key not in seen:
            seen.add(key)
            unique.append(r)
            if len(unique) >= top_n:
                break
    return unique


# ─── EXCEL ────────────────────────────────────────────────────────────────────
def save_excel(out_dir: Path, global_top: List[dict], tf_tops: Dict[str, List[dict]],
               all_results: List[dict]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"[{ts()}] ⚠ openpyxl no disponible — saltando Excel")
        return

    HDR_COLS = [
        "Rank", "Timeframe", "MA_Type", "MA_Fast", "MA_Slow", "Leverage",
        "Trailing%", "SL_Type", "Fib_Mode", "RSI_Filter", "EMA200_Filter",
        "ATR_Filter", "Compound", "TP_Fib", "TimeFilter", "VolProfile",
        "LiqConfirm", "Risk%",
        "PnL$", "PnL%", "WinRate%", "Trades", "MaxDD%",
        "Sharpe", "PF", "BestTrade", "WorstTrade",
    ]
    PARAM_KEYS = [
        "interval", "ma_type", "ma_fast", "ma_slow", "leverage", "trailing_pct",
        "sl_type", "fib_mode", "rsi_filter", "ema200_filter", "atr_filter",
        "compound", "tp_fib", "time_filter", "vol_profile", "liq_confirm", "risk_pct",
        "total_pnl", "total_pnl_pct", "win_rate", "total_trades", "max_drawdown",
        "sharpe", "profit_factor", "best_trade", "worst_trade",
    ]

    def _fv(k, v):
        if k == "trailing_pct": return f"{v*100:.1f}%"
        if k == "risk_pct":     return f"{v*100:.0f}%"
        if isinstance(v, bool): return "Sí" if v else "No"
        if isinstance(v, float): return round(v, 3)
        return v

    H_FILL = PatternFill("solid", fgColor="0D1F2D")
    H_FONT = Font(color="4FC3F7", bold=True, size=9)
    A_FILL = PatternFill("solid", fgColor="090F1A")

    def write_sheet(ws, rows: List[dict], title: str):
        ws.title = title[:31]
        ws.append(HDR_COLS)
        for cell in ws[1]:
            cell.font = H_FONT
            cell.fill = H_FILL
            cell.alignment = Alignment(horizontal="center")
        for i, r in enumerate(rows):
            row_data = [i + 1] + [_fv(k, r.get(k, "")) for k in PARAM_KEYS]
            ws.append(row_data)
            if i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = A_FILL
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 22)

    wb = Workbook()
    ws1 = wb.active
    write_sheet(ws1, global_top, "Top30_Global")

    for tf in TIMEFRAMES:
        write_sheet(wb.create_sheet(), tf_tops.get(tf, []), f"Top30_{tf}")

    # Hoja 7: Análisis por tipo de MA
    ws7 = wb.create_sheet("Análisis_MA")
    ws7.append(["MA_Type", "N", "PnL_Promedio", "PnL_Máximo", "WinRate_Prom"])
    ma_grp: Dict[str, list] = {}
    for r in all_results:
        ma_grp.setdefault(r.get("ma_type", "?"), []).append(r)
    for ma_t, rows in ma_grp.items():
        pnls = [r.get("total_pnl", 0) for r in rows]
        wrs  = [r.get("win_rate", 0) for r in rows]
        ws7.append([ma_t, len(rows), round(float(np.mean(pnls)), 2),
                    round(float(np.max(pnls)), 2), round(float(np.mean(wrs)), 2)])

    # Hoja 8: Análisis por crypto (aproximado — un bot prueba todas)
    ws8 = wb.create_sheet("Análisis_Crypto")
    ws8.append(["Nota"])
    ws8.append(["Cada combinación se evalúa sobre las 10 criptos simultáneamente."])
    ws8.append(["Para análisis por crypto individual, ejecutar optimizer_v3 con --coins COIN."])

    # Hoja 9: Análisis por filtro horario
    ws9 = wb.create_sheet("Análisis_Horario")
    ws9.append(["Filtro_Horario", "N", "PnL_Promedio", "PnL_Máximo", "WinRate_Prom"])
    tf_grp: Dict[str, list] = {}
    for r in all_results:
        tf_grp.setdefault(r.get("time_filter", "none"), []).append(r)
    for tfk, rows in tf_grp.items():
        pnls = [r.get("total_pnl", 0) for r in rows]
        wrs  = [r.get("win_rate", 0) for r in rows]
        ws9.append([tfk, len(rows), round(float(np.mean(pnls)), 2),
                    round(float(np.max(pnls)), 2), round(float(np.mean(wrs)), 2)])

    # Hoja 10: Todas las combinaciones (max 10,000)
    ws10 = wb.create_sheet("Todas_Combinaciones")
    sorted_all = sorted(all_results, key=_score, reverse=True)
    write_sheet(ws10, sorted_all[:10_000], "Todas_Combinaciones")

    out_path = out_dir / "resultados_optimizacion_v3.xlsx"
    try:
        wb.save(out_path)
        print(f"[{ts()}] ✓ Excel: {out_path}")
    except Exception as e:
        print(f"[{ts()}] ⚠ Error Excel: {e}")


# ─── PDF ──────────────────────────────────────────────────────────────────────
def save_pdf(out_dir: Path, global_top: List[dict], tf_tops: Dict[str, List[dict]],
             total_tried: int, duration_sec: float, prev_json: Optional[Path]):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
    except ImportError:
        print(f"[{ts()}] ⚠ matplotlib no disponible — saltando PDF")
        return

    BG = "#07090f"; FG = "#c9d4e0"; ACC = "#4fc3f7"
    plt.rcParams.update({
        "figure.facecolor": BG, "axes.facecolor": BG, "axes.edgecolor": "#162030",
        "text.color": FG, "axes.labelcolor": FG, "xtick.color": FG, "ytick.color": FG,
        "grid.color": "#0d1520", "font.family": "monospace", "font.size": 9,
    })

    out_path = out_dir / "informe_optimizacion_v3.pdf"
    try:
        with PdfPages(out_path) as pdf:

            # ── Portada ──────────────────────────────────────────────────────
            fig, ax = plt.subplots(figsize=(11.69, 8.27))
            ax.axis("off")
            ax.text(0.5, 0.88, "⬡ AlphaChainBots", ha="center", fontsize=30, color=ACC, weight="bold")
            ax.text(0.5, 0.78, "Informe de Optimización v3", ha="center", fontsize=18, color=FG)
            ax.text(0.5, 0.66, f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ha="center", fontsize=11, color="#78909c")
            ax.text(0.5, 0.58, f"Combinaciones probadas: {total_tried:,}", ha="center", fontsize=13, color=FG)
            ax.text(0.5, 0.50, f"Duración: {duration_sec/60:.1f} min", ha="center", fontsize=13, color=FG)
            ax.text(0.5, 0.42, f"Monedas: {', '.join(COINS)}", ha="center", fontsize=9, color="#78909c")
            ax.text(0.5, 0.36, f"Timeframes: {', '.join(TIMEFRAMES)}", ha="center", fontsize=9, color="#78909c")
            if global_top:
                best = global_top[0]
                ax.text(0.5, 0.24, "— Mejor combinación global —", ha="center", fontsize=11, color=ACC)
                summary = (f"TF: {best.get('interval','?')} | "
                           f"{best.get('ma_type','').upper()} {best.get('ma_fast','?')}/{best.get('ma_slow','?')} | "
                           f"Lev: {best.get('leverage','?')}x | "
                           f"PnL: {best.get('total_pnl',0):+.2f}$ ({best.get('total_pnl_pct',0):+.1f}%) | "
                           f"WR: {best.get('win_rate',0):.1f}% | Sharpe: {best.get('sharpe',0):.2f}")
                ax.text(0.5, 0.16, summary, ha="center", fontsize=8.5, color=FG)
            pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()

            # ── Resumen ejecutivo ─────────────────────────────────────────────
            fig, axes = plt.subplots(2, 3, figsize=(11.69, 8.27))
            fig.suptitle("Resumen Ejecutivo", fontsize=13, color=ACC, y=0.98)

            tf_names = TIMEFRAMES
            best_pnls = [tf_tops[tf][0].get("total_pnl", 0) if tf_tops.get(tf) else 0 for tf in tf_names]
            axes[0, 0].bar(tf_names, best_pnls,
                           color=["#00e676" if p >= 0 else "#ff4466" for p in best_pnls], alpha=0.85)
            axes[0, 0].set_title("Mejor PnL por Timeframe", color=ACC, fontsize=9)
            axes[0, 0].set_ylabel("PnL ($)")
            axes[0, 0].tick_params(labelsize=7)

            pnl_vals = [r.get("total_pnl", 0) for r in global_top]
            axes[0, 1].hist(pnl_vals, bins=min(10, len(pnl_vals)), color=ACC, alpha=0.7)
            axes[0, 1].set_title("Distribución PnL — Top 30 Global", color=ACC, fontsize=9)
            axes[0, 1].tick_params(labelsize=7)

            wr_vals = [r.get("win_rate", 0) for r in global_top[:10]]
            axes[0, 2].barh([f"#{i+1}" for i in range(len(wr_vals))], wr_vals, color="#4fc3f7", alpha=0.8)
            axes[0, 2].set_title("Win Rate — Top 10", color=ACC, fontsize=9)
            axes[0, 2].set_xlim(0, 100)
            axes[0, 2].tick_params(labelsize=7)

            ma_cnt: Dict[str, int] = {}
            for r in global_top:
                k = r.get("ma_type", "?")
                ma_cnt[k] = ma_cnt.get(k, 0) + 1
            if ma_cnt:
                axes[1, 0].pie(list(ma_cnt.values()), labels=list(ma_cnt.keys()),
                               colors=[ACC, "#ff6d00", "#69f0ae"],
                               autopct="%1.0f%%", textprops={"color": FG, "size": 8})
            axes[1, 0].set_title("EMA vs SMA (Top 30)", color=ACC, fontsize=9)

            lev_cnt: Dict[str, int] = {}
            for r in global_top:
                k = f"{r.get('leverage','?')}x"
                lev_cnt[k] = lev_cnt.get(k, 0) + 1
            if lev_cnt:
                axes[1, 1].bar(list(lev_cnt.keys()), list(lev_cnt.values()), color="#69f0ae", alpha=0.8)
            axes[1, 1].set_title("Apalancamiento (Top 30)", color=ACC, fontsize=9)
            axes[1, 1].tick_params(labelsize=7)

            sl_cnt: Dict[str, int] = {}
            for r in global_top:
                k = r.get("sl_type", "?")
                sl_cnt[k] = sl_cnt.get(k, 0) + 1
            if sl_cnt:
                axes[1, 2].bar(list(sl_cnt.keys()), list(sl_cnt.values()), color="#ce93d8", alpha=0.8)
            axes[1, 2].set_title("Tipo Stop Loss (Top 30)", color=ACC, fontsize=9)
            axes[1, 2].tick_params(labelsize=7)

            plt.tight_layout()
            pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()

            # ── Top 5 por timeframe ───────────────────────────────────────────
            for tf in TIMEFRAMES:
                tops = tf_tops.get(tf, [])[:5]
                if not tops:
                    continue
                fig, ax = plt.subplots(figsize=(11.69, 8.27))
                ax.axis("off")
                ax.text(0.5, 0.95, f"Top 5 — Timeframe {tf}", ha="center",
                        fontsize=14, color=ACC, weight="bold")
                headers = ["#", "MA", "Lev", "Trail%", "SL", "RSI_F", "EMA200", "ATR_F",
                           "Fib", "Cpnd", "Risk%", "PnL$", "WR%", "DD%", "Sharpe", "PF"]
                ax.text(0.02, 0.87, "  ".join(f"{h:>7}" for h in headers),
                        fontsize=6.5, color=ACC, family="monospace")
                y = 0.79
                for i, r in enumerate(tops):
                    row = [
                        f"#{i+1}",
                        f"{r.get('ma_type','?').upper()}{r.get('ma_fast','')}/{r.get('ma_slow','')}",
                        f"{r.get('leverage','?')}x",
                        f"{r.get('trailing_pct',0)*100:.1f}%",
                        r.get("sl_type","?")[:5],
                        r.get("rsi_filter","none")[:6],
                        r.get("ema200_filter","none")[:6],
                        r.get("atr_filter","none")[:5],
                        r.get("fib_mode","?")[:4],
                        "Y" if r.get("compound") else "N",
                        f"{r.get('risk_pct',0)*100:.0f}%",
                        f"{r.get('total_pnl',0):+.0f}$",
                        f"{r.get('win_rate',0):.1f}%",
                        f"{r.get('max_drawdown',0):.1f}%",
                        f"{r.get('sharpe',0):.2f}",
                        f"{r.get('profit_factor',0):.2f}",
                    ]
                    ax.text(0.02, y, "  ".join(f"{v:>7}" for v in row),
                            fontsize=6.5, color=FG, family="monospace")
                    y -= 0.10
                pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()

            # ── Heatmap parámetros rentables ──────────────────────────────────
            try:
                fig, ax = plt.subplots(figsize=(11.69, 8.27))
                lev_vals = sorted(set(r.get("leverage", 0) for r in global_top))
                tf_vals  = TIMEFRAMES
                heat = np.zeros((len(tf_vals), len(lev_vals)))
                for r in global_top:
                    ti = tf_vals.index(r.get("interval", "1h")) if r.get("interval") in tf_vals else -1
                    li = lev_vals.index(r.get("leverage", 0)) if r.get("leverage") in lev_vals else -1
                    if ti >= 0 and li >= 0:
                        heat[ti, li] += r.get("total_pnl", 0)
                im = ax.imshow(heat, cmap="RdYlGn", aspect="auto")
                ax.set_xticks(range(len(lev_vals))); ax.set_xticklabels([f"{v}x" for v in lev_vals])
                ax.set_yticks(range(len(tf_vals)));  ax.set_yticklabels(tf_vals)
                ax.set_title("Heatmap PnL — Timeframe × Leverage (Top 30)", color=ACC, fontsize=11)
                plt.colorbar(im, ax=ax, label="PnL ($)")
                plt.tight_layout()
                pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()
            except Exception:
                plt.close()

            # ── Comparativa con versión anterior ─────────────────────────────
            if prev_json and prev_json.exists():
                try:
                    with open(prev_json, "r", encoding="utf-8") as f:
                        prev = json.load(f)
                    prev_list = prev if isinstance(prev, list) else prev.get("results", [])
                    prev_best = prev_list[0].get("total_pnl", 0) if prev_list else 0
                    curr_best = global_top[0].get("total_pnl", 0) if global_top else 0
                    fig, ax = plt.subplots(figsize=(11.69, 5))
                    ax.bar(["Optimizer anterior", "Optimizer v3"],
                           [prev_best, curr_best],
                           color=[ACC, "#00e676"], alpha=0.85, width=0.4)
                    ax.set_title("Comparativa — Mejor PnL global", color=ACC, fontsize=12)
                    ax.set_ylabel("PnL ($)", color=FG)
                    plt.tight_layout()
                    pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()
                except Exception:
                    plt.close()

            # ── Conclusiones ──────────────────────────────────────────────────
            fig, ax = plt.subplots(figsize=(11.69, 8.27))
            ax.axis("off")
            ax.text(0.5, 0.93, "Conclusiones y Recomendaciones", ha="center",
                    fontsize=14, color=ACC, weight="bold")
            if global_top:
                best = global_top[0]
                lines = [
                    f"Mejor combinación global encontrada:",
                    f"  Timeframe: {best.get('interval','?')} | MA: {best.get('ma_type','').upper()} {best.get('ma_fast','?')}/{best.get('ma_slow','?')}",
                    f"  Apalancamiento: {best.get('leverage','?')}x | Trailing: {best.get('trailing_pct',0)*100:.1f}% | SL: {best.get('sl_type','?')}",
                    f"  RSI Filter: {best.get('rsi_filter','none')} | EMA200: {best.get('ema200_filter','none')} | ATR: {best.get('atr_filter','none')}",
                    f"  Fibonacci: {best.get('fib_mode','?')} | Compounding: {'Sí' if best.get('compound') else 'No'} | TP Fib: {'Sí' if best.get('tp_fib') else 'No'}",
                    f"  PnL: {best.get('total_pnl',0):+.2f}$ ({best.get('total_pnl_pct',0):+.1f}%)",
                    f"  Win Rate: {best.get('win_rate',0):.1f}% | Max DD: {best.get('max_drawdown',0):.1f}% | Sharpe: {best.get('sharpe',0):.2f}",
                    "",
                    f"Total de {total_tried:,} combinaciones evaluadas en {duration_sec/60:.1f} minutos.",
                    f"Se evaluaron {len(COINS)} criptomonedas en {len(TIMEFRAMES)} timeframes.",
                    "",
                    "IMPORTANTE: Estos resultados son sobre datos históricos (in-sample).",
                    "Se recomienda validar con datos fuera de muestra (out-of-sample)",
                    "antes de aplicar en trading real.",
                    "",
                    "Próximos pasos sugeridos:",
                    "  1. Aplicar top 5 parámetros en paper trading durante 30 días",
                    "  2. Comparar performance con bots actuales en dashboard",
                    "  3. Re-optimizar con --days 180 para detectar cambios de régimen",
                ]
                y = 0.82
                for line in lines:
                    ax.text(0.06, y, line, fontsize=9, color=FG, family="monospace")
                    y -= 0.058
            pdf.savefig(fig, facecolor=BG, bbox_inches="tight"); plt.close()

        print(f"[{ts()}] ✓ PDF: {out_path}")
    except Exception as e:
        print(f"[{ts()}] ⚠ Error PDF: {e}")
        traceback.print_exc()


# ─── JSON ─────────────────────────────────────────────────────────────────────
def save_json(out_dir: Path, global_top: List[dict], tf_tops: Dict[str, List[dict]],
              generated_at: str, total_tried: int):
    data = {
        "version":      VERSION,
        "generated_at": generated_at,
        "total_tried":  total_tried,
        "results":      global_top,
        "by_timeframe": {tf: tf_tops.get(tf, []) for tf in TIMEFRAMES},
    }
    out_path = out_dir / "top_params_v3.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"[{ts()}] ✓ JSON: {out_path}")


# ─── RESUMEN FINAL EN PANTALLA ────────────────────────────────────────────────
def print_summary(global_top: List[dict], tf_tops: Dict[str, List[dict]],
                  total_tried: int, duration_sec: float, out_dir: Path):
    sep = "=" * 72
    print(f"\n{sep}")
    print(f"[{ts()}]  OPTIMIZACIÓN COMPLETADA — AlphaChainBots Optimizer v3")
    print(sep)
    print(f"  Combinaciones probadas : {total_tried:,}")
    print(f"  Duración total         : {duration_sec/60:.1f} min")
    print(f"\n  TOP 5 GLOBAL:")
    print(f"  {'#':>2}  {'TF':>4}  {'MA':>12}  {'Lev':>5}  {'Trail%':>7}  "
          f"{'PnL$':>10}  {'WR%':>6}  {'DD%':>6}  {'Sharpe':>7}")
    print(f"  {'-'*70}")
    for i, r in enumerate(global_top[:5]):
        ma_str = f"{r.get('ma_type','').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
        print(f"  {i+1:>2}  {r.get('interval','?'):>4}  {ma_str:>12}  "
              f"{r.get('leverage','?'):>4}x  "
              f"{r.get('trailing_pct',0)*100:>6.1f}%  "
              f"{r.get('total_pnl',0):>+10.2f}$  "
              f"{r.get('win_rate',0):>5.1f}%  "
              f"{r.get('max_drawdown',0):>5.1f}%  "
              f"{r.get('sharpe',0):>7.3f}")
    print(f"\n  MEJOR POR TIMEFRAME:")
    print(f"  {'TF':>4}  {'MA':>12}  {'Lev':>5}  {'PnL$':>10}  {'WR%':>6}  {'Sharpe':>7}")
    print(f"  {'-'*50}")
    for tf in TIMEFRAMES:
        tops = tf_tops.get(tf, [])
        if tops:
            r = tops[0]
            ma_str = f"{r.get('ma_type','').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
            print(f"  {tf:>4}  {ma_str:>12}  {r.get('leverage','?'):>4}x  "
                  f"{r.get('total_pnl',0):>+10.2f}$  "
                  f"{r.get('win_rate',0):>5.1f}%  "
                  f"{r.get('sharpe',0):>7.3f}")
        else:
            print(f"  {tf:>4}  (sin resultados)")
    print(f"\n  Archivos generados en {out_dir.resolve()}:")
    print(f"    resultados_optimizacion_v3.xlsx")
    print(f"    informe_optimizacion_v3.pdf")
    print(f"    top_params_v3.json")
    print(f"    checkpoint_v3.json")
    print(f"{sep}\n")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def _dl_task(coin: str, tf: str, cache_dir: Path, days: int) -> Tuple[str, str, Optional[np.ndarray]]:
    """Descarga (o carga del caché) las velas para un par coin/TF."""
    arr = load_or_fetch(cache_dir, coin, tf, days)
    return coin, tf, arr


def _add(r: dict, all_results: list, seen_keys: set, tf_results: dict) -> None:
    """Añade un resultado a all_results si no es duplicado."""
    if not r or r.get("total_trades", 0) == 0:
        return
    p   = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
    key = hashlib.md5(json.dumps(p, sort_keys=True).encode()).hexdigest()
    if key in seen_keys:
        return
    seen_keys.add(key)
    all_results.append(r)
    tf = r.get("interval")
    if tf in tf_results:
        tf_results[tf].append(r)


def main():

    parser = argparse.ArgumentParser(description="AlphaChainBots Optimizer v3")
    parser.add_argument("--days",    type=int,  default=365,       help="Días de historial")
    parser.add_argument("--samples", type=int,  default=500_000,   help="Combinaciones a probar")
    parser.add_argument("--out",     type=str,  default="./resultados", help="Directorio de salida")
    parser.add_argument("--workers", type=int,  default=None,      help="Nº de workers paralelos")
    parser.add_argument("--resume",  action="store_true",          help="Retomar desde checkpoint")
    args = parser.parse_args()

    out_dir    = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_dir  = out_dir / "cache_v3"
    cache_dir.mkdir(parents=True, exist_ok=True)
    ckpt_path  = out_dir / "checkpoint_v3.json"
    n_workers  = args.workers or os.cpu_count() or 4

    # ── Banner ────────────────────────────────────────────────────────────────
    sep = "=" * 72
    print(f"\n{sep}")
    print(f"[{ts()}]  AlphaChainBots Optimizer v3 — {VERSION}")
    print(sep)
    print(f"  Cores disponibles      : {os.cpu_count()}")
    print(f"  Combinaciones posibles : {TOTAL_COMBOS:,}")
    print(f"  A probar               : {args.samples:,}")
    print(f"  Días de historial      : {args.days}")
    print(f"  Workers                : {n_workers}")
    print(f"  Método multiprocessing : {mp.get_start_method()}")
    print(f"  Salida                 : {out_dir.resolve()}")
    print(f"{sep}\n")

    # ── Retomar desde checkpoint ──────────────────────────────────────────────
    processed_start = 0
    all_results: List[dict] = []
    tf_results:  Dict[str, List[dict]] = {tf: [] for tf in TIMEFRAMES}
    seen_keys:   set = set()

    if args.resume:
        ckpt = load_checkpoint(ckpt_path)
        if ckpt:
            processed_start = ckpt.get("processed", 0)
            all_results     = ckpt.get("results", [])
            for tf in TIMEFRAMES:
                tf_results[tf] = ckpt.get("tf_results", {}).get(tf, [])
            for r in all_results:
                p = {k: v for k, v in r.items() if k in _PARAM_FIELDS_SET}
                seen_keys.add(hashlib.md5(json.dumps(p, sort_keys=True).encode()).hexdigest())
            print(f"[{ts()}] ▶ Reanudando: {processed_start:,} combinaciones ya procesadas\n")
        else:
            print(f"[{ts()}] ⚠ No se encontró checkpoint — iniciando desde cero\n")

    # ── Descargar velas históricas en PARALELO ────────────────────────────────
    total_tasks = len(COINS) * len(TIMEFRAMES)
    print(f"[{ts()}] Descargando velas: {len(COINS)} monedas × {len(TIMEFRAMES)} TF "
          f"= {total_tasks} tareas en paralelo…")
    candles: Dict[str, Dict[str, np.ndarray]] = {tf: {} for tf in TIMEFRAMES}
    dl_ok = 0
    dl_fail = 0

    with ThreadPoolExecutor(max_workers=min(total_tasks, 20)) as ex:
        futures = {ex.submit(_dl_task, coin, tf, cache_dir, args.days): (coin, tf)
                   for tf in TIMEFRAMES for coin in COINS}
        for fut in as_completed(futures):
            coin, tf_key = futures[fut]
            try:
                _, _, arr = fut.result()
            except Exception as e:
                print(f"  [ERR] {coin}/{tf_key}: {e}")
                dl_fail += 1
                continue
            if arr is not None and len(arr) > 50:
                candles[tf_key][coin] = arr
                print(f"  [OK] {coin:6s}/{tf_key:4s}  {len(arr):,} velas")
                dl_ok += 1
            else:
                print(f"  [--] {coin:6s}/{tf_key:4s}  sin datos")
                dl_fail += 1

    total_ok = sum(len(v) for v in candles.values())
    print(f"\n[{ts()}] Descarga completada: {dl_ok} OK / {dl_fail} fallidos "
          f"({total_ok} pares coin/TF disponibles)\n")

    # ── Cache para workers ────────────────────────────────────────────────────
    worker_pkl = out_dir / "_wcache_v3.pkl"
    with open(worker_pkl, "wb") as f:
        pickle.dump(candles, f)

    # ── Test de arranque del Pool ─────────────────────────────────────────────
    n_sys_cores = os.cpu_count() or 1
    print(f"[{ts()}] CPUs del sistema: {n_sys_cores} — usando {n_workers} workers")
    print(f"[{ts()}] Iniciando Pool de multiprocessing…", end=" ", flush=True)
    with mp.Pool(processes=n_workers, initializer=_worker_init,
                 initargs=(str(worker_pkl),)) as _tp:
        _tp.map(_worker_ping, range(n_workers))
    print(f"✓ {n_workers} workers listos\n")

    # ── Random search ─────────────────────────────────────────────────────────
    rng        = random.Random(42)
    start_time = time.time()
    processed  = processed_start
    # Cada worker recibe un segmento grande y genera sus propios params internamente.
    # Así la IPC es O(1) por ronda, no O(n) por combinación.
    SEG_SIZE   = 4_000   # combos por worker por ronda

    last_top_print = processed_start
    last_ckpt      = processed_start

    print(f"[{ts()}] Random search: {args.samples - processed_start:,} combinaciones restantes…")
    print(f"[{ts()}] Estrategia: {n_workers} workers × {SEG_SIZE:,} combos/seg = "
          f"{n_workers * SEG_SIZE:,} combos/ronda\n")

    initargs = (str(worker_pkl),)
    with mp.Pool(processes=n_workers, initializer=_worker_init,
                 initargs=initargs, maxtasksperchild=500) as pool:
        while processed < args.samples:
            remaining  = args.samples - processed
            n_tasks    = max(1, min(n_workers * 2, remaining // max(SEG_SIZE // 2, 1)))
            actual_seg = max(1, min(SEG_SIZE, remaining // n_tasks))
            tasks = [{"n": actual_seg, "seed": rng.randint(0, 2**32)}
                     for _ in range(n_tasks)]
            try:
                batch_results = pool.map(_worker_run_segment, tasks)
            except Exception as e:
                print(f"\n[{ts()}] ⚠ Error en pool.map: {e} — continuando…")
                processed += n_tasks * actual_seg
                continue
            for seg in batch_results:
                for r in seg:
                    _add(r, all_results, seen_keys, tf_results)
            processed += sum(t["n"] for t in tasks)

            # Progress
            elapsed  = max(time.time() - start_time, 0.001)
            done_now = processed - processed_start
            speed    = done_now / elapsed
            eta_sec  = (args.samples - processed) / speed if speed > 0 else 0
            eta_str  = (f"{eta_sec/3600:.1f}h" if eta_sec >= 3600
                        else f"{eta_sec/60:.1f}min" if eta_sec >= 60
                        else f"{eta_sec:.0f}s")
            pct = processed / args.samples * 100.0
            print(f"\r[{ts()}] {processed:>9,}/{args.samples:,} ({pct:5.1f}%) "
                  f"| {speed:>7,.0f} comb/s | ETA: {eta_str:>8} "
                  f"| únicos: {len(all_results):,}",
                  end="", flush=True)

            # Top 3 cada 25k
            if processed - last_top_print >= PRINT_TOP_EVERY:
                last_top_print = processed
                top3 = sorted(all_results, key=_score, reverse=True)[:3]
                print(f"\n[{ts()}] ─── Top 3 actual ────────────────────────────────────")
                for i, r in enumerate(top3):
                    ma = f"{r.get('ma_type','?').upper()} {r.get('ma_fast','?')}/{r.get('ma_slow','?')}"
                    print(f"  #{i+1}  {r.get('interval','?'):>4}  {ma:>14}  "
                          f"Lev:{r.get('leverage','?'):>2}x  "
                          f"PnL:{r.get('total_pnl',0):>+10.2f}$  "
                          f"WR:{r.get('win_rate',0):>5.1f}%")
                print()

            # Checkpoint cada 20k
            if processed - last_ckpt >= CHECKPOINT_EVERY:
                last_ckpt = processed
                save_checkpoint(ckpt_path, processed, all_results, tf_results)
                print(f"\n[{ts()}] 💾 Checkpoint ({processed:,}) — {ckpt_path.name}\n")

    print(f"\n\n[{ts()}] ✓ Random search completado.")

    # ── Hill climbing sobre top 20% ───────────────────────────────────────────
    if all_results:
        sorted_all = sorted(all_results, key=_score, reverse=True)
        n_seeds    = max(1, len(sorted_all) // 5)
        seeds      = sorted_all[:n_seeds]
        n_hc       = min(n_seeds * 8, 60_000)
        print(f"[{ts()}] Hill climbing: {n_seeds} semillas × 8 perturbaciones = {n_hc:,} intentos…")

        hc_batch: List[dict] = []
        for seed in seeds[: n_hc // 8]:
            sp = OptParams(**{k: seed[k] for k in _PARAM_FIELDS_SET})
            for _ in range(8):
                hc_batch.append(asdict(perturb(sp, rng)))

        # Un chunk por worker — un solo pool.map distribuye todo el HC en paralelo
        chunk_sz = max(1, len(hc_batch) // n_workers)
        hc_tasks = [{"params": hc_batch[i:i + chunk_sz]}
                    for i in range(0, len(hc_batch), chunk_sz)]

        with mp.Pool(processes=n_workers, initializer=_worker_init,
                     initargs=initargs, maxtasksperchild=2000) as pool:
            try:
                hc_results = pool.map(_worker_run_hc_segment, hc_tasks)
                for seg in hc_results:
                    for r in seg:
                        _add(r, all_results, seen_keys, tf_results)
                processed += len(hc_batch)
            except Exception as e:
                print(f"\n[{ts()}] ⚠ Error HC pool.map: {e} — continuando…")

        print(f"[{ts()}] ✓ Hill climbing completado. Total procesadas: {processed:,}")

    # ── Deduplicación y top rankings ──────────────────────────────────────────
    print(f"\n[{ts()}] Deduplicando y construyendo rankings…")
    global_top = deduplicate(all_results, 30)
    tf_tops    = {tf: deduplicate(tf_results.get(tf, []), 30) for tf in TIMEFRAMES}

    duration     = time.time() - start_time
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Guardar salidas ───────────────────────────────────────────────────────
    print(f"[{ts()}] Guardando archivos de salida…")

    prev_json_candidates = [
        out_dir / "top_params.json",
        out_dir.parent / "top_params.json",
    ]
    prev_json = next((p for p in prev_json_candidates if p.exists()), None)

    save_json(out_dir, global_top, tf_tops, generated_at, processed)
    save_excel(out_dir, global_top, tf_tops, all_results)
    save_pdf(out_dir, global_top, tf_tops, processed, duration, prev_json)
    save_checkpoint(ckpt_path, processed, all_results, tf_results)

    # Limpiar pkl temporal
    try:
        worker_pkl.unlink()
    except Exception:
        pass

    # ── Resumen ───────────────────────────────────────────────────────────────
    print_summary(global_top, tf_tops, processed, duration, out_dir)


if __name__ == "__main__":
    # Requerido en Windows para que los subprocesos spawn no relancen main().
    mp.freeze_support()
    # 'fork' evita reinicializar el proceso completo en cada worker (Linux/Mac).
    # Windows no soporta fork — usa spawn por defecto (seguro en todos los SO).
    if platform.system() != "Windows":
        mp.set_start_method("fork", force=True)
    main()
