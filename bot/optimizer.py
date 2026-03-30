#!/usr/bin/env python3
"""
optimizer.py — AlphaChainBots Parameter Optimizer
=====================================================
Descarga velas reales de Hyperliquid, explora ~50 000 combinaciones
con random-search + hill-climbing usando todos los cores disponibles,
y genera tres archivos de informe:
  · resultados_optimizacion.xlsx
  · informe_optimizacion.pdf
  · top_params.json

Uso:
    python optimizer.py [--days 365] [--samples 50000] [--out ./resultados]
"""

import argparse
import json
import math
import multiprocessing as mp
import os
import pickle
import random
import sys
import time
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import requests

# ── Dependencias opcionales (tqdm, openpyxl, matplotlib) ──────────────────────
try:
    from tqdm import tqdm as _tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl no encontrado — install: pip install openpyxl")

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    from matplotlib.backends.backend_pdf import PdfPages
    HAS_MPL = True
except ImportError:
    HAS_MPL = False
    print("[WARN] matplotlib no encontrado — install: pip install matplotlib")

# ── Constantes globales ────────────────────────────────────────────────────────
HL_URL        = "https://api.hyperliquid.xyz/info"
INITIAL_EQ    = 10_000.0
TOP_N_COINS   = 10
FALLBACK_COINS = ["BTC", "ETH", "SOL", "XRP", "DOGE", "AVAX", "BNB", "LINK", "HYPE", "TAO"]

INTERVAL_MS: Dict[str, int] = {
    "15m": 900_000, "30m": 1_800_000, "1h": 3_600_000, "4h": 14_400_000,
}

# Espacio de parámetros a explorar
TIMEFRAMES  = ["15m", "30m", "1h", "4h"]
EMA_PAIRS   = [(8, 21), (13, 34), (21, 55), (20, 50)]
LEVERAGES   = [2, 3, 5, 10]
TRAIL_PCTS  = [0.005, 0.008, 0.010, 0.015, 0.020]
RSI_THRESHS = [(25, 75), (20, 80), (30, 70)]     # (oversold, overbought)
ATR_MULTS   = [1.5, 2.0, 2.5]                    # ATR stop multiplier si use_atr
VOL_RATIOS  = [1.2, 1.5, 2.0]                    # mínimo ratio volumen

# Variable global del worker (cargada una vez por proceso)
_WORKER_CACHE: Optional[Dict] = None


# ── Parámetros de optimización ─────────────────────────────────────────────────
@dataclass
class OptParams:
    timeframe:       str   = "1h"
    ma_fast:         int   = 20
    ma_slow:         int   = 50
    leverage:        int   = 3
    trailing_pct:    float = 0.010
    use_rsi:         bool  = False
    rsi_os:          float = 25.0    # oversold
    rsi_ob:          float = 75.0    # overbought
    use_ema200:      bool  = False
    use_atr_filter:  bool  = False
    atr_mult:        float = 2.0
    use_fib:         bool  = False
    use_sr:          bool  = False
    compounding:     bool  = True
    use_tp_fib:      bool  = False
    time_filter:     bool  = False
    time_start_h:    int   = 8
    time_end_h:      int   = 20
    vol_strict:      bool  = False
    vol_ratio:       float = 1.5
    use_liq_confirm: bool  = False
    stop_type:       str   = "trailing"   # "trailing" | "fixed"
    risk_per_trade:  float = 0.02


def random_params(rng: random.Random) -> OptParams:
    """Genera un OptParams completamente aleatorio."""
    fast, slow = rng.choice(EMA_PAIRS)
    rsi_pair   = rng.choice(RSI_THRESHS)
    return OptParams(
        timeframe       = rng.choice(TIMEFRAMES),
        ma_fast         = fast,
        ma_slow         = slow,
        leverage        = rng.choice(LEVERAGES),
        trailing_pct    = rng.choice(TRAIL_PCTS),
        use_rsi         = rng.random() < 0.4,
        rsi_os          = float(rsi_pair[0]),
        rsi_ob          = float(rsi_pair[1]),
        use_ema200      = rng.random() < 0.4,
        use_atr_filter  = rng.random() < 0.3,
        atr_mult        = rng.choice(ATR_MULTS),
        use_fib         = rng.random() < 0.3,
        use_sr          = rng.random() < 0.3,
        compounding     = rng.random() < 0.6,
        use_tp_fib      = rng.random() < 0.25,
        time_filter     = rng.random() < 0.25,
        time_start_h    = rng.choice([6, 7, 8, 9]),
        time_end_h      = rng.choice([18, 19, 20, 21, 22]),
        vol_strict      = rng.random() < 0.3,
        vol_ratio       = rng.choice(VOL_RATIOS),
        use_liq_confirm = rng.random() < 0.2,
        stop_type       = rng.choice(["trailing", "fixed"]),
        risk_per_trade  = rng.choice([0.01, 0.02, 0.03]),
    )


def perturb(p: OptParams, rng: random.Random) -> OptParams:
    """Modifica un parámetro al azar (hill climbing step)."""
    import copy
    q = copy.deepcopy(p)
    field_name = rng.choice([
        "timeframe", "ma_pair", "leverage", "trailing_pct",
        "use_rsi", "use_ema200", "use_atr_filter", "use_fib",
        "use_sr", "compounding", "use_tp_fib", "time_filter",
        "vol_strict", "use_liq_confirm", "stop_type", "risk_per_trade",
    ])
    if field_name == "timeframe":
        q.timeframe = rng.choice(TIMEFRAMES)
    elif field_name == "ma_pair":
        fast, slow  = rng.choice(EMA_PAIRS)
        q.ma_fast   = fast; q.ma_slow = slow
    elif field_name == "leverage":
        q.leverage  = rng.choice(LEVERAGES)
    elif field_name == "trailing_pct":
        q.trailing_pct = rng.choice(TRAIL_PCTS)
    elif field_name == "use_rsi":
        q.use_rsi   = not q.use_rsi
        if q.use_rsi:
            rp = rng.choice(RSI_THRESHS); q.rsi_os = rp[0]; q.rsi_ob = rp[1]
    elif field_name == "use_ema200":
        q.use_ema200 = not q.use_ema200
    elif field_name == "use_atr_filter":
        q.use_atr_filter = not q.use_atr_filter
        if q.use_atr_filter: q.atr_mult = rng.choice(ATR_MULTS)
    elif field_name == "use_fib":
        q.use_fib = not q.use_fib
    elif field_name == "use_sr":
        q.use_sr  = not q.use_sr
    elif field_name == "compounding":
        q.compounding = not q.compounding
    elif field_name == "use_tp_fib":
        q.use_tp_fib = not q.use_tp_fib
    elif field_name == "time_filter":
        q.time_filter = not q.time_filter
    elif field_name == "vol_strict":
        q.vol_strict = not q.vol_strict
        if q.vol_strict: q.vol_ratio = rng.choice(VOL_RATIOS)
    elif field_name == "use_liq_confirm":
        q.use_liq_confirm = not q.use_liq_confirm
    elif field_name == "stop_type":
        q.stop_type = "fixed" if q.stop_type == "trailing" else "trailing"
    elif field_name == "risk_per_trade":
        q.risk_per_trade = rng.choice([0.01, 0.02, 0.03])
    return q


# ── Descarga de velas ──────────────────────────────────────────────────────────
def fetch_candles(coin: str, interval: str, days: int) -> List[dict]:
    """Descarga velas OHLCV de Hyperliquid."""
    try:
        end_ms   = int(time.time() * 1000)
        start_ms = end_ms - int(days * 86_400_000)
        r = requests.post(HL_URL, json={
            "type": "candleSnapshot",
            "req": {"coin": coin, "interval": interval,
                    "startTime": start_ms, "endTime": end_ms},
        }, timeout=30)
        raw = r.json()
        if not isinstance(raw, list):
            return []
        out = []
        for c in raw:
            try:
                out.append({
                    "t": int(c["t"]),
                    "o": float(c.get("o", 0) or 0),
                    "h": float(c.get("h", 0) or 0),
                    "l": float(c.get("l", 0) or 0),
                    "c": float(c.get("c", 0) or 0),
                    "v": float(c.get("v", 0) or 0),
                })
            except (KeyError, TypeError, ValueError):
                pass
        return sorted(out, key=lambda x: x["t"])
    except Exception as e:
        print(f"  [WARN] fetch_candles {coin}/{interval}: {e}")
        return []


def get_top_coins(n: int = TOP_N_COINS) -> List[str]:
    """Obtiene las top N cryptos por volumen 24h."""
    try:
        r    = requests.post(HL_URL, json={"type": "metaAndAssetCtxs"}, timeout=15)
        d    = r.json()
        meta, ctxs = d[0], d[1]
        coins = []
        for i, asset in enumerate(meta["universe"]):
            if i < len(ctxs):
                vol = float(ctxs[i].get("dayNtlVlm", 0) or 0)
                coins.append((asset["name"], vol))
        coins.sort(key=lambda x: x[1], reverse=True)
        result = [c[0] for c in coins[:n]]
        return result if result else FALLBACK_COINS[:n]
    except Exception:
        return FALLBACK_COINS[:n]


def download_all_candles(coins: List[str], days: int,
                         cache_path: Path) -> Dict[str, List[dict]]:
    """Descarga todas las velas necesarias y las cachea en disco."""
    if cache_path.exists():
        age_h = (time.time() - cache_path.stat().st_mtime) / 3600
        if age_h < 4:
            print(f"[INFO] Usando caché de velas ({age_h:.1f}h de antigüedad): {cache_path}")
            with open(cache_path, "rb") as f:
                return pickle.load(f)
        print(f"[INFO] Caché expirada ({age_h:.1f}h), re-descargando…")

    print(f"[INFO] Descargando velas para {len(coins)} coins × {len(TIMEFRAMES)} timeframes "
          f"({days} días)…")
    cache: Dict[str, List[dict]] = {}
    total = len(coins) * len(TIMEFRAMES)
    done  = 0
    for coin in coins:
        for iv in TIMEFRAMES:
            key = f"{coin}_{iv}"
            cache[key] = fetch_candles(coin, iv, days)
            done += 1
            print(f"  [{done}/{total}] {key}: {len(cache[key])} velas", flush=True)
            time.sleep(0.08)

    with open(cache_path, "wb") as f:
        pickle.dump(cache, f)
    print(f"[INFO] Caché guardada en {cache_path}")
    return cache


# ── Indicadores (vectorizados con numpy) ──────────────────────────────────────
def _ema_np(arr: np.ndarray, period: int) -> np.ndarray:
    if len(arr) < period:
        return np.full(len(arr), np.nan)
    k  = 2.0 / (period + 1)
    out = np.full(len(arr), np.nan)
    out[period - 1] = arr[:period].mean()
    for i in range(period, len(arr)):
        out[i] = arr[i] * k + out[i - 1] * (1 - k)
    return out


def _sma_np(arr: np.ndarray, period: int) -> np.ndarray:
    out = np.full(len(arr), np.nan)
    for i in range(period - 1, len(arr)):
        out[i] = arr[i - period + 1:i + 1].mean()
    return out


def _rsi_np(arr: np.ndarray, period: int = 14) -> np.ndarray:
    out = np.full(len(arr), np.nan)
    if len(arr) < period + 1:
        return out
    diffs  = np.diff(arr)
    gains  = np.where(diffs > 0, diffs, 0.0)
    losses = np.where(diffs < 0, -diffs, 0.0)
    ag     = gains[:period].mean()
    al     = losses[:period].mean()
    for i in range(period, len(diffs)):
        ag = (ag * (period - 1) + gains[i]) / period
        al = (al * (period - 1) + losses[i]) / period
        rs = ag / al if al > 0 else 100.0
        out[i + 1] = 100 - 100 / (1 + rs)
    return out


def _atr_np(high: np.ndarray, low: np.ndarray,
            close: np.ndarray, period: int = 14) -> np.ndarray:
    tr  = np.maximum(high[1:] - low[1:],
          np.maximum(np.abs(high[1:] - close[:-1]),
                     np.abs(low[1:] - close[:-1])))
    atr = np.full(len(close), np.nan)
    if len(tr) < period:
        return atr
    atr[period] = tr[:period].mean()
    for i in range(period, len(tr)):
        atr[i + 1] = (atr[i] * (period - 1) + tr[i]) / period
    return atr


def precompute_indicators(cache: Dict[str, List[dict]]) -> Dict[str, Dict]:
    """Pre-calcula todos los indicadores una vez para todos los (coin, iv)."""
    indicators: Dict[str, Dict] = {}
    for key, candles in cache.items():
        if not candles:
            continue
        c = np.array([x["c"] for x in candles], dtype=np.float64)
        h = np.array([x["h"] for x in candles], dtype=np.float64)
        l = np.array([x["l"] for x in candles], dtype=np.float64)
        v = np.array([x["v"] for x in candles], dtype=np.float64)
        t = np.array([x["t"] for x in candles], dtype=np.int64)
        # Hours UTC (for time filter)
        hours = ((t // 3_600_000) % 24).astype(np.int8)

        ind: Dict = {
            "close": c, "high": h, "low": l, "volume": v,
            "time_ms": t, "hours": hours,
            "ema200": _ema_np(c, 200),
            "rsi14":  _rsi_np(c, 14),
            "atr14":  _atr_np(h, l, c, 14),
        }
        # Pre-compute EMA/SMA pairs
        for fast, slow in EMA_PAIRS:
            ind[f"ema{fast}"] = _ema_np(c, fast)
            ind[f"ema{slow}"] = _ema_np(c, slow)
            ind[f"sma{fast}"] = _sma_np(c, fast)
            ind[f"sma{slow}"] = _sma_np(c, slow)
        indicators[key] = ind
    return indicators


# ── Núcleo de simulación ──────────────────────────────────────────────────────
def _simulate(params: OptParams, indicators: Dict[str, Dict],
              coins: List[str]) -> Dict:
    """
    Ejecuta la simulación de un set de parámetros sobre todos los coins.
    Retorna diccionario de métricas.
    """
    iv        = params.timeframe
    fast      = params.ma_fast
    slow      = params.ma_slow
    lev       = params.leverage
    trail     = params.trailing_pct
    risk      = params.risk_per_trade
    compound  = params.compounding

    equity    = INITIAL_EQ
    all_pnl:  List[float]      = []
    all_ts:   List[int]        = []
    eq_curve: List[Tuple[int, float]] = []

    for coin in coins:
        key = f"{coin}_{iv}"
        ind = indicators.get(key)
        if ind is None or len(ind["close"]) < slow + 30:
            continue

        c    = ind["close"]
        h    = ind["high"]
        l    = ind["low"]
        v    = ind["volume"]
        t    = ind["time_ms"]
        hrs  = ind["hours"]
        rsi  = ind["rsi14"]
        e200 = ind["ema200"]
        atr  = ind["atr14"]
        efst = ind[f"ema{fast}"]
        eslw = ind[f"ema{slow}"]

        avg_vol  = np.nanmean(v)

        in_pos   = False
        dir_     = 0      # 1=long, -1=short
        entry    = 0.0
        best_px  = 0.0
        stop_px  = 0.0
        tp_px    = 0.0
        coin_eq  = equity / len(coins)

        for i in range(slow + 1, len(c)):
            price   = c[i]
            ts_i    = t[i]

            if in_pos:
                # Update trailing stop
                if params.stop_type == "trailing":
                    if dir_ == 1:
                        if price > best_px:
                            best_px = price
                            stop_px = best_px * (1 - trail)
                        hit_stop = price <= stop_px
                    else:
                        if price < best_px:
                            best_px = price
                            stop_px = best_px * (1 + trail)
                        hit_stop = price >= stop_px
                else:
                    # Fixed stop at entry ± ATR * mult
                    atr_v = atr[i] if not np.isnan(atr[i]) else price * 0.01
                    if dir_ == 1:
                        stop_px = entry - atr_v * params.atr_mult
                        hit_stop = price <= stop_px
                    else:
                        stop_px = entry + atr_v * params.atr_mult
                        hit_stop = price >= stop_px

                # TP Fibonacci
                hit_tp = False
                if params.use_tp_fib and tp_px > 0:
                    hit_tp = (dir_ == 1 and price >= tp_px) or \
                             (dir_ == -1 and price <= tp_px)

                if hit_stop or hit_tp:
                    if dir_ == 1:
                        pnl_pct = (price - entry) / entry * lev
                    else:
                        pnl_pct = (entry - price) / entry * lev
                    pnl_usd  = coin_eq * risk * pnl_pct
                    coin_eq += pnl_usd
                    if compound:
                        equity  = max(equity + pnl_usd, 1.0)
                    all_pnl.append(pnl_usd)
                    all_ts.append(ts_i)
                    in_pos  = False
                continue

            # ─── Entry logic ────────────────────────────────────────────────
            if np.isnan(efst[i]) or np.isnan(efst[i-1]) or \
               np.isnan(eslw[i]) or np.isnan(eslw[i-1]):
                continue

            crossed_up   = efst[i-1] <= eslw[i-1] and efst[i] > eslw[i]
            crossed_down = efst[i-1] >= eslw[i-1] and efst[i] < eslw[i]
            if not (crossed_up or crossed_down):
                continue

            sig = 1 if crossed_up else -1

            # Volume filter
            min_vol = avg_vol * (params.vol_ratio if params.vol_strict else 1.2)
            if v[i] < min_vol:
                continue

            # EMA 200 trend filter
            if params.use_ema200 and not np.isnan(e200[i]):
                if sig == 1  and price < e200[i]: continue
                if sig == -1 and price > e200[i]: continue

            # RSI filter
            if params.use_rsi and not np.isnan(rsi[i]):
                if sig == 1  and rsi[i] > params.rsi_ob: continue
                if sig == -1 and rsi[i] < params.rsi_os: continue

            # ATR filter: skip if volatility too low
            if params.use_atr_filter and not np.isnan(atr[i]):
                avg_atr = np.nanmean(atr[max(0, i-20):i])
                if atr[i] < avg_atr * 0.5:
                    continue

            # Time filter
            if params.time_filter:
                if not (params.time_start_h <= int(hrs[i]) < params.time_end_h):
                    continue

            # Liquidation confirm: proxy = volume spike > 3× avg
            if params.use_liq_confirm:
                if v[i] < avg_vol * 3.0:
                    continue

            # Fibonacci confirm: price near 0.5/0.618 retracement
            if params.use_fib:
                lookback = 50
                lo = l[max(0, i-lookback):i].min()
                hi = h[max(0, i-lookback):i].max()
                span = hi - lo
                if span > 0:
                    retrace = (hi - price) / span
                    if not (0.35 <= retrace <= 0.72):
                        continue

            # S/R filter: price near recent high/low
            if params.use_sr:
                lookback = 20
                lo = l[max(0, i-lookback):i].min()
                hi = h[max(0, i-lookback):i].max()
                dist_lo = abs(price - lo) / price
                dist_hi = abs(price - hi) / price
                if min(dist_lo, dist_hi) > 0.015:
                    continue

            # ─── Enter position ─────────────────────────────────────────────
            in_pos  = True
            dir_    = sig
            entry   = price
            best_px = price

            if params.stop_type == "trailing":
                stop_px = entry * (1 - trail) if sig == 1 else entry * (1 + trail)
            else:
                atr_v   = atr[i] if not np.isnan(atr[i]) else price * 0.01
                stop_px = (entry - atr_v * params.atr_mult if sig == 1
                           else entry + atr_v * params.atr_mult)

            # TP at next Fibonacci extension (1.618)
            if params.use_tp_fib:
                lookback = 50
                lo = l[max(0, i-lookback):i].min()
                hi = h[max(0, i-lookback):i].max()
                span = hi - lo
                tp_px = (price + span * 0.618 if sig == 1
                         else price - span * 0.618)
            else:
                tp_px = 0.0

    # ─── Métricas finales ────────────────────────────────────────────────────
    return _metrics(all_pnl, all_ts, equity if compound else INITIAL_EQ, params)


def _metrics(pnls: List[float], tss: List[int],
             final_eq: float, params: OptParams) -> Dict:
    """Calcula métricas a partir de lista de PnL."""
    initial = INITIAL_EQ
    n = len(pnls)
    if n == 0:
        return {
            "total_pnl": 0, "total_pnl_pct": 0, "win_rate": 0,
            "total_trades": 0, "max_drawdown": 0, "sharpe": 0,
            "profit_factor": 0, "best_trade": 0, "worst_trade": 0,
            "final_equity": initial, "equity_curve": [[0, initial]],
        }

    wins   = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p <= 0]
    total_pnl = sum(pnls)
    final_eq  = initial + total_pnl

    # Equity curve
    eq  = initial
    curve: List[Tuple[int, float]] = []
    if tss:
        for ts, pnl in zip(tss, pnls):
            eq += pnl
            curve.append([ts, round(eq, 2)])
    else:
        curve = [[int(time.time() * 1000), initial]]

    # Max drawdown
    peak = initial; max_dd = 0.0
    running = initial
    for p in pnls:
        running += p
        if running > peak: peak = running
        dd = (peak - running) / peak * 100 if peak > 0 else 0
        if dd > max_dd: max_dd = dd

    # Sharpe (simplified)
    if len(pnls) >= 3:
        rets = [p / initial for p in pnls]
        mu   = sum(rets) / len(rets)
        std  = math.sqrt(sum((r - mu)**2 for r in rets) / len(rets))
        sharpe = (mu / std * math.sqrt(len(rets))) if std > 0 else 0.0
    else:
        sharpe = 0.0

    pf = sum(wins) / abs(sum(losses)) if losses and sum(wins) > 0 else 0.0

    # Downsample equity curve
    if len(curve) > 100:
        step = len(curve) // 100
        curve = curve[::step] + [curve[-1]]

    return {
        "total_pnl":     round(total_pnl, 2),
        "total_pnl_pct": round(total_pnl / initial * 100, 2),
        "win_rate":      round(len(wins) / n * 100, 1),
        "total_trades":  n,
        "max_drawdown":  round(max_dd, 2),
        "sharpe":        round(sharpe, 2),
        "profit_factor": round(pf, 2),
        "best_trade":    round(max(pnls), 2),
        "worst_trade":   round(min(pnls), 2),
        "final_equity":  round(final_eq, 2),
        "equity_curve":  curve,
    }


# ── Worker multiprocessing ────────────────────────────────────────────────────
_INDICATORS: Optional[Dict] = None
_COINS_LIST: Optional[List[str]] = None


def _worker_init(cache_path: str, coins: List[str]):
    """Inicializa estado global del worker (llamado una vez por proceso)."""
    global _INDICATORS, _COINS_LIST
    with open(cache_path, "rb") as f:
        raw_cache = pickle.load(f)
    _INDICATORS = precompute_indicators(raw_cache)
    _COINS_LIST = coins


def _worker_eval(params: OptParams) -> Tuple[OptParams, Dict]:
    """Función del worker: evalúa un set de parámetros."""
    try:
        metrics = _simulate(params, _INDICATORS, _COINS_LIST)
        return params, metrics
    except Exception as e:
        empty = _metrics([], [], INITIAL_EQ, params)
        return params, empty


# ── Progress bar ─────────────────────────────────────────────────────────────
class ProgressBar:
    def __init__(self, total: int, desc: str = ""):
        self.total    = total
        self.desc     = desc
        self.done     = 0
        self.start_ts = time.time()

    def update(self, n: int = 1):
        self.done += n
        elapsed = time.time() - self.start_ts
        pct     = self.done / self.total * 100
        rate    = self.done / elapsed if elapsed > 0 else 1
        eta_s   = (self.total - self.done) / rate if rate > 0 else 0
        bar_len = 40
        filled  = int(bar_len * self.done / self.total)
        bar     = "█" * filled + "░" * (bar_len - filled)
        eta_str = f"{int(eta_s//60)}m{int(eta_s%60)}s" if eta_s < 3600 else f"{eta_s/3600:.1f}h"
        print(f"\r{self.desc} [{bar}] {pct:5.1f}% {self.done}/{self.total} "
              f"ETA:{eta_str} ({rate:.0f} eval/s)   ",
              end="", flush=True)

    def close(self):
        print()


# ── Informe Excel ─────────────────────────────────────────────────────────────
def _cell_color(ws, row: int, col: int, hex_color: str):
    ws.cell(row=row, column=col).fill = PatternFill(
        fill_type="solid", fgColor=hex_color)


def _hdr_style(cell, bold: bool = True):
    cell.font      = Font(bold=bold, color="FFFFFF")
    cell.fill      = PatternFill(fill_type="solid", fgColor="0D2030")
    cell.alignment = Alignment(horizontal="center")


def save_excel(top30: List[Dict], all_results: List[Dict],
               out_path: Path, coins: List[str]):
    if not HAS_OPENPYXL:
        print("[SKIP] Excel: openpyxl no disponible")
        return
    wb = openpyxl.Workbook()

    # ── Hoja 1: Top 30 ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Top 30"
    headers = ["Rank","Timeframe","EMA","Leverage","Trailing%","RSI","EMA200",
               "ATR","Fib","S/R","Compound","TP_Fib","TimeFilter","VolStrict",
               "LiqConf","StopType","Risk%","PnL $","PnL %","WinRate%",
               "Trades","MaxDD%","Sharpe","PF","BestTrade","WorstTrade"]
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        _hdr_style(cell)
    for ri, rec in enumerate(top30, 2):
        p = rec["params"]
        m = rec["metrics"]
        row = [
            ri - 1, p.timeframe, f"{p.ma_fast}/{p.ma_slow}", p.leverage,
            f"{p.trailing_pct*100:.1f}%",
            "✓" if p.use_rsi else "✗", "✓" if p.use_ema200 else "✗",
            "✓" if p.use_atr_filter else "✗", "✓" if p.use_fib else "✗",
            "✓" if p.use_sr else "✗", "✓" if p.compounding else "✗",
            "✓" if p.use_tp_fib else "✗", "✓" if p.time_filter else "✗",
            "✓" if p.vol_strict else "✗", "✓" if p.use_liq_confirm else "✗",
            p.stop_type, f"{p.risk_per_trade*100:.0f}%",
            m["total_pnl"], m["total_pnl_pct"], m["win_rate"],
            m["total_trades"], m["max_drawdown"], m["sharpe"],
            m["profit_factor"], m["best_trade"], m["worst_trade"],
        ]
        for ci, val in enumerate(row, 1):
            ws1.cell(row=ri, column=ci, value=val)
        # Green/red PnL
        pnl_col = headers.index("PnL $") + 1
        color = "1B5E20" if m["total_pnl"] >= 0 else "B71C1C"
        _cell_color(ws1, ri, pnl_col, color)
        _cell_color(ws1, ri, pnl_col + 1, color)

    for col in ws1.columns:
        ws1.column_dimensions[get_column_letter(col[0].column)].width = 12

    # ── Hoja 2: Por Timeframe ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Timeframe")
    ws2.append(["Timeframe","Count","Avg PnL $","Avg WinRate%","Avg MaxDD%","Avg Sharpe"])
    for h in ws2[1]: _hdr_style(h)
    for tf in TIMEFRAMES:
        sub = [r for r in all_results if r["params"].timeframe == tf]
        if not sub: continue
        ws2.append([
            tf, len(sub),
            round(sum(r["metrics"]["total_pnl"] for r in sub) / len(sub), 2),
            round(sum(r["metrics"]["win_rate"]   for r in sub) / len(sub), 1),
            round(sum(r["metrics"]["max_drawdown"] for r in sub) / len(sub), 2),
            round(sum(r["metrics"]["sharpe"] for r in sub) / len(sub), 2),
        ])

    # ── Hoja 3: Por Horario ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Horario")
    ws3.append(["Filtro Horario","Count","Avg PnL $","Avg WinRate%"])
    for h in ws3[1]: _hdr_style(h)
    for tf_flag, label in [(True, "Con filtro"), (False, "Sin filtro")]:
        sub = [r for r in all_results if r["params"].time_filter == tf_flag]
        if not sub: continue
        ws3.append([label, len(sub),
                    round(sum(r["metrics"]["total_pnl"] for r in sub) / len(sub), 2),
                    round(sum(r["metrics"]["win_rate"] for r in sub) / len(sub), 1)])

    # ── Hoja 4: Por EMA ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Por EMA")
    ws4.append(["EMA Pair","Count","Avg PnL $","Avg WinRate%","Avg Sharpe"])
    for h in ws4[1]: _hdr_style(h)
    for fast, slow in EMA_PAIRS:
        sub = [r for r in all_results if r["params"].ma_fast == fast]
        if not sub: continue
        ws4.append([f"{fast}/{slow}", len(sub),
                    round(sum(r["metrics"]["total_pnl"] for r in sub) / len(sub), 2),
                    round(sum(r["metrics"]["win_rate"] for r in sub) / len(sub), 1),
                    round(sum(r["metrics"]["sharpe"] for r in sub) / len(sub), 2)])

    # ── Hoja 5: Por Crypto ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Por Crypto")
    ws5.append(["Nota"])
    ws5.append(["El backtest evalúa todos los coins combinados por set de params."])
    ws5.append(["Coins usados:"] + coins)

    # ── Hoja 6: Todas las combinaciones (top 5000) ────────────────────────────
    ws6 = wb.create_sheet("Todas (top 5000)")
    hdr6 = ["Rank","Timeframe","EMA","Lev","Trail%","PnL $","PnL %",
            "WinRate%","Trades","MaxDD%","Sharpe","PF"]
    ws6.append(hdr6)
    for h in ws6[1]: _hdr_style(h)
    for ri, rec in enumerate(all_results[:5000], 2):
        p = rec["params"]; m = rec["metrics"]
        ws6.append([ri-1, p.timeframe, f"{p.ma_fast}/{p.ma_slow}", p.leverage,
                    round(p.trailing_pct*100,1),
                    m["total_pnl"], m["total_pnl_pct"], m["win_rate"],
                    m["total_trades"], m["max_drawdown"], m["sharpe"],
                    m["profit_factor"]])

    wb.save(out_path)
    print(f"[OK] Excel: {out_path}")


# ── Informe PDF ───────────────────────────────────────────────────────────────
def save_pdf(top30: List[Dict], out_path: Path, days: int):
    if not HAS_MPL:
        print("[SKIP] PDF: matplotlib no disponible")
        return

    DARK    = "#07090f"
    ACCENT  = "#4fc3f7"
    GREEN   = "#00e676"
    RED     = "#ff4466"
    FG      = "#c9d4e0"

    plt.rcParams.update({
        "figure.facecolor": DARK, "axes.facecolor": DARK,
        "axes.edgecolor":   "#1e3a4a", "axes.labelcolor": FG,
        "text.color":       FG, "xtick.color": FG, "ytick.color": FG,
        "grid.color":       "#0d1520", "grid.alpha": 0.6,
    })

    with PdfPages(out_path) as pdf:

        # ── Portada ─────────────────────────────────────────────────────────
        fig, ax = plt.subplots(figsize=(11.7, 8.3))
        ax.axis("off")
        ax.text(0.5, 0.82, "AlphaChainBots", ha="center", fontsize=34,
                fontweight="bold", color=ACCENT)
        ax.text(0.5, 0.72, "Informe de Optimización de Parámetros",
                ha="center", fontsize=20, color=FG)
        ax.text(0.5, 0.62, f"Periodo analizado: últimos {days} días",
                ha="center", fontsize=14, color="#78909c")
        ax.text(0.5, 0.55, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                ha="center", fontsize=12, color="#546e7a")
        ax.text(0.5, 0.42, f"Total combinaciones analizadas: {len(top30)}" ,
                ha="center", fontsize=11, color=FG)
        if top30:
            best = top30[0]
            p = best["params"]; m = best["metrics"]
            summary = (f"Mejor configuración: {p.timeframe} · EMA {p.ma_fast}/{p.ma_slow} "
                       f"· x{p.leverage} leverage · trailing {p.trailing_pct*100:.1f}%\n"
                       f"PnL: +${m['total_pnl']:.0f} ({m['total_pnl_pct']:.1f}%) "
                       f"· WinRate: {m['win_rate']}% · Sharpe: {m['sharpe']}")
            ax.text(0.5, 0.30, summary, ha="center", fontsize=12, color=GREEN,
                    multialignment="center",
                    bbox=dict(boxstyle="round,pad=0.5", facecolor="#0d2030",
                              edgecolor=ACCENT, alpha=0.8))
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Top 30 tabla ────────────────────────────────────────────────────
        fig, ax = plt.subplots(figsize=(18, 11))
        ax.axis("off")
        ax.set_title("Top 30 Combinaciones — Ordenadas por PnL",
                     fontsize=14, color=ACCENT, pad=12)
        cols = ["#","TF","EMA","Lev","Trail%","PnL$","PnL%",
                "WR%","Trades","DD%","Sharpe","PF","RSI","E200","Fib"]
        rows = []
        for i, rec in enumerate(top30[:30], 1):
            p = rec["params"]; m = rec["metrics"]
            rows.append([
                str(i), p.timeframe, f"{p.ma_fast}/{p.ma_slow}",
                f"x{p.leverage}", f"{p.trailing_pct*100:.1f}%",
                f"${m['total_pnl']:.0f}", f"{m['total_pnl_pct']:.1f}%",
                f"{m['win_rate']}%", str(m["total_trades"]),
                f"{m['max_drawdown']:.1f}%",
                str(m["sharpe"]), str(m["profit_factor"]),
                "✓" if p.use_rsi else "✗",
                "✓" if p.use_ema200 else "✗",
                "✓" if p.use_fib else "✗",
            ])
        tbl = ax.table(cellText=rows, colLabels=cols,
                       loc="center", cellLoc="center")
        tbl.auto_set_font_size(False); tbl.set_fontsize(8)
        tbl.scale(1, 1.4)
        for (r, c), cell in tbl.get_celld().items():
            cell.set_facecolor(DARK if r > 0 else "#0D2030")
            cell.set_text_props(color=FG if r > 0 else ACCENT)
            cell.set_edgecolor("#1e3a4a")
            if r > 0 and c == 5:
                pnl_v = float(rows[r-1][5].replace("$","").replace(",",""))
                cell.set_facecolor("#1B5E20" if pnl_v >= 0 else "#B71C1C")
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Equity curves — top 12 ──────────────────────────────────────────
        fig, axes = plt.subplots(3, 4, figsize=(18, 10))
        fig.suptitle("Equity Curves — Top 12 Configuraciones",
                     color=ACCENT, fontsize=14)
        for idx, ax in enumerate(axes.flatten()):
            ax.set_facecolor(DARK)
            if idx >= len(top30):
                ax.axis("off"); continue
            rec = top30[idx]
            m   = rec["metrics"]; p = rec["params"]
            curve = m.get("equity_curve", [])
            if len(curve) >= 2:
                xs = [c[0] for c in curve]
                ys = [c[1] for c in curve]
                color = GREEN if m["total_pnl"] >= 0 else RED
                ax.plot(xs, ys, color=color, linewidth=1.2)
                ax.fill_between(xs, INITIAL_EQ, ys,
                                alpha=0.12,
                                color=GREEN if m["total_pnl"] >= 0 else RED)
                ax.axhline(INITIAL_EQ, color="#546e7a", lw=0.5, ls="--")
            title = (f"#{idx+1} {p.timeframe} EMA{p.ma_fast}/{p.ma_slow} "
                     f"x{p.leverage}\nPnL: ${m['total_pnl']:.0f} "
                     f"({m['total_pnl_pct']:+.1f}%) WR:{m['win_rate']}%")
            ax.set_title(title, fontsize=7, color=FG)
            ax.xaxis.set_visible(False)
            ax.tick_params(labelsize=6)
            ax.yaxis.set_major_formatter(mticker.FuncFormatter(
                lambda x, _: f"${x:.0f}"))
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Heatmap: Timeframe × EMA Pair ────────────────────────────────────
        if len(top30) >= 4:
            fig, axes = plt.subplots(1, 2, figsize=(14, 6))
            fig.suptitle("Análisis de Parámetros", color=ACCENT, fontsize=13)

            # PnL por timeframe
            ax = axes[0]
            tf_pnl = {tf: [] for tf in TIMEFRAMES}
            for rec in top30:
                tf_pnl[rec["params"].timeframe].append(rec["metrics"]["total_pnl"])
            tfs   = list(tf_pnl.keys())
            means = [sum(v)/len(v) if v else 0 for v in tf_pnl.values()]
            colors = [GREEN if m >= 0 else RED for m in means]
            ax.bar(tfs, means, color=colors, edgecolor="#1e3a4a")
            ax.set_title("Avg PnL por Timeframe (Top 30)", color=FG, fontsize=10)
            ax.set_ylabel("Avg PnL ($)", color=FG)
            ax.grid(True, axis="y", alpha=0.3)

            # PnL por EMA par
            ax2 = axes[1]
            ema_pnl = {f"{f}/{s}": [] for f, s in EMA_PAIRS}
            for rec in top30:
                p = rec["params"]
                k = f"{p.ma_fast}/{p.ma_slow}"
                if k in ema_pnl:
                    ema_pnl[k].append(rec["metrics"]["total_pnl"])
            ema_lbls  = list(ema_pnl.keys())
            ema_means = [sum(v)/len(v) if v else 0 for v in ema_pnl.values()]
            colors2   = [GREEN if m >= 0 else RED for m in ema_means]
            ax2.bar(ema_lbls, ema_means, color=colors2, edgecolor="#1e3a4a")
            ax2.set_title("Avg PnL por EMA Pair (Top 30)", color=FG, fontsize=10)
            ax2.set_ylabel("Avg PnL ($)", color=FG)
            ax2.grid(True, axis="y", alpha=0.3)

            plt.tight_layout()
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Análisis booleanos ────────────────────────────────────────────────
        fig, axes = plt.subplots(2, 3, figsize=(15, 8))
        fig.suptitle("Impacto de Filtros Booleanos (Top 30)", color=ACCENT, fontsize=13)
        flags = [
            ("use_rsi",     "Filtro RSI"),
            ("use_ema200",  "EMA 200"),
            ("use_fib",     "Fibonacci"),
            ("compounding", "Compounding"),
            ("time_filter", "Filtro Horario"),
            ("use_liq_confirm", "Confirm. Liq"),
        ]
        for ax, (attr, label) in zip(axes.flatten(), flags):
            with_f  = [r["metrics"]["total_pnl"] for r in top30 if getattr(r["params"], attr)]
            without = [r["metrics"]["total_pnl"] for r in top30 if not getattr(r["params"], attr)]
            vals    = [sum(with_f)/len(with_f) if with_f else 0,
                       sum(without)/len(without) if without else 0]
            lbls    = [f"Con\n({len(with_f)})", f"Sin\n({len(without)})"]
            colors  = [GREEN if v >= 0 else RED for v in vals]
            ax.bar(lbls, vals, color=colors, edgecolor="#1e3a4a")
            ax.set_title(label, color=FG, fontsize=9)
            ax.set_ylabel("Avg PnL ($)", color=FG, fontsize=8)
            ax.grid(True, axis="y", alpha=0.3)
            ax.tick_params(labelsize=8)
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Conclusiones ────────────────────────────────────────────────────
        fig, ax = plt.subplots(figsize=(11.7, 8.3))
        ax.axis("off")
        ax.text(0.5, 0.95, "Conclusiones y Recomendaciones",
                ha="center", fontsize=16, color=ACCENT, fontweight="bold",
                transform=ax.transAxes)
        if top30:
            best = top30[0]; p = best["params"]; m = best["metrics"]
            lines = [
                f"",
                f"MEJOR CONFIGURACIÓN ENCONTRADA:",
                f"  • Timeframe: {p.timeframe}",
                f"  • EMA: {p.ma_fast}/{p.ma_slow}",
                f"  • Apalancamiento: x{p.leverage}",
                f"  • Trailing Stop: {p.trailing_pct*100:.1f}%",
                f"  • Stop Type: {p.stop_type}",
                f"  • RSI Filter: {'Sí' if p.use_rsi else 'No'}",
                f"  • EMA 200: {'Sí' if p.use_ema200 else 'No'}",
                f"  • Fibonacci: {'Sí' if p.use_fib else 'No'}",
                f"  • Compounding: {'Sí' if p.compounding else 'No'}",
                f"",
                f"RENDIMIENTO (capital inicial $10,000):",
                f"  • PnL Total: ${m['total_pnl']:,.2f} ({m['total_pnl_pct']:+.2f}%)",
                f"  • Win Rate: {m['win_rate']}%",
                f"  • Trades: {m['total_trades']}",
                f"  • Max Drawdown: {m['max_drawdown']}%",
                f"  • Sharpe Ratio: {m['sharpe']}",
                f"  • Profit Factor: {m['profit_factor']}",
                f"",
                f"ADVERTENCIA: Los resultados históricos no garantizan resultados futuros.",
                f"El trading con apalancamiento conlleva alto riesgo de pérdida de capital.",
            ]
            ax.text(0.08, 0.88, "\n".join(lines), ha="left", va="top",
                    fontsize=11, color=FG, transform=ax.transAxes,
                    family="monospace")
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

    print(f"[OK] PDF: {out_path}")


# ── Informe JSON ──────────────────────────────────────────────────────────────
def save_json(top30: List[Dict], out_path: Path):
    output = []
    for i, rec in enumerate(top30, 1):
        p  = rec["params"]
        m  = rec["metrics"]
        # Omit equity_curve from JSON (too large)
        metrics_clean = {k: v for k, v in m.items() if k != "equity_curve"}
        output.append({
            "rank":    i,
            "params":  asdict(p),
            "metrics": metrics_clean,
        })
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"top_30": output,
                   "generated_at": datetime.now().isoformat()}, f, indent=2)
    print(f"[OK] JSON: {out_path}")


# ── Resumen por consola ───────────────────────────────────────────────────────
def print_summary(top30: List[Dict]):
    print("\n" + "═" * 100)
    print(f"{'AlphaChainBots — TOP 30 COMBINACIONES MÁS RENTABLES':^100}")
    print("═" * 100)
    hdr = f"{'#':>3}  {'TF':<5} {'EMA':<8} {'Lev':>4} {'Trail':>6} {'PnL $':>9} "
    hdr += f"{'PnL %':>7} {'WR%':>6} {'Trades':>7} {'DD%':>6} {'Sharpe':>7} {'PF':>5}"
    print(hdr)
    print("─" * 100)
    for i, rec in enumerate(top30[:30], 1):
        p = rec["params"]; m = rec["metrics"]
        sign  = "+" if m["total_pnl"] >= 0 else ""
        flags = ""
        if p.use_rsi:         flags += "R"
        if p.use_ema200:      flags += "E"
        if p.use_fib:         flags += "F"
        if p.compounding:     flags += "C"
        if p.time_filter:     flags += "T"
        if p.use_liq_confirm: flags += "L"
        print(f"{i:>3}  {p.timeframe:<5} {p.ma_fast}/{p.ma_slow:<5} x{p.leverage:>2} "
              f"{p.trailing_pct*100:>5.1f}% "
              f"{sign}{m['total_pnl']:>8.2f} "
              f"{sign}{m['total_pnl_pct']:>6.1f}% "
              f"{m['win_rate']:>5.1f}% "
              f"{m['total_trades']:>6} "
              f"{m['max_drawdown']:>5.1f}% "
              f"{m['sharpe']:>6.2f} "
              f"{m['profit_factor']:>4.2f} "
              f"[{flags}]")
    print("─" * 100)
    print("Leyenda flags: R=RSI  E=EMA200  F=Fib  C=Compounding  T=TimeFilter  L=LiqConfirm")
    print("═" * 100)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="AlphaChainBots Parameter Optimizer")
    parser.add_argument("--days",    type=int,   default=365,
                        help="Días de historial a descargar (default 365)")
    parser.add_argument("--samples", type=int,   default=50_000,
                        help="Número total de combinaciones (default 50000)")
    parser.add_argument("--out",     type=str,   default=".",
                        help="Carpeta de salida para los informes")
    parser.add_argument("--seed",    type=int,   default=42)
    parser.add_argument("--workers", type=int,   default=0,
                        help="Número de procesos (0 = todos los cores)")
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_path = out_dir / "candles_cache.pkl"

    n_workers  = args.workers or max(1, mp.cpu_count() - 1)
    n_random   = int(args.samples * 0.80)
    n_hillclimb = args.samples - n_random
    rng        = random.Random(args.seed)

    print(f"\n{'='*60}")
    print(f"  AlphaChainBots — Optimizador de Parámetros")
    print(f"{'='*60}")
    print(f"  Días historial  : {args.days}")
    print(f"  Combinaciones   : {args.samples:,} ({n_random:,} random + {n_hillclimb:,} hill-climbing)")
    print(f"  Workers         : {n_workers}")
    print(f"  Salida          : {out_dir.resolve()}")
    print(f"{'='*60}\n")

    # ── 1. Descargar velas ───────────────────────────────────────────────────
    print("[1/5] Descargando velas históricas…")
    coins = get_top_coins(TOP_N_COINS)
    print(f"  Coins: {', '.join(coins)}")
    candles_cache = download_all_candles(coins, args.days, cache_path)

    # ── 2. Generar combinaciones aleatorias ──────────────────────────────────
    print(f"\n[2/5] Generando {n_random:,} combinaciones aleatorias…")
    random_combos = [random_params(rng) for _ in range(n_random)]

    # ── 3. Fase random search ────────────────────────────────────────────────
    print(f"\n[3/5] Random search con {n_workers} workers…")
    t0       = time.time()
    results: List[Dict] = []

    if HAS_TQDM:
        pbar = _tqdm(total=n_random, desc="  Random search",
                     ncols=90, unit="eval",
                     bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining} {rate_fmt}]")
    else:
        pbar = ProgressBar(n_random, "  Random search")

    with mp.Pool(
        processes=n_workers,
        initializer=_worker_init,
        initargs=(str(cache_path), coins),
    ) as pool:
        for params, metrics in pool.imap_unordered(
                _worker_eval, random_combos, chunksize=50):
            results.append({"params": params, "metrics": metrics})
            if HAS_TQDM: pbar.update(1)
            else:        pbar.update(1)

    if HAS_TQDM: pbar.close()
    else:        pbar.close()
    print(f"  Random search: {len(results):,} eval en {time.time()-t0:.1f}s")

    # ── 4. Hill climbing ────────────────────────────────────────────────────
    print(f"\n[4/5] Hill climbing ({n_hillclimb:,} iteraciones)…")
    results.sort(key=lambda r: r["metrics"]["total_pnl"], reverse=True)
    seeds = [r["params"] for r in results[:20]]   # top 20 como puntos de arranque
    hc_results: List[Dict] = []

    if HAS_TQDM:
        hpbar = _tqdm(total=n_hillclimb, desc="  Hill climbing",
                      ncols=90, unit="eval",
                      bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining} {rate_fmt}]")
    else:
        hpbar = ProgressBar(n_hillclimb, "  Hill climbing")

    # Generamos combos de hill climbing
    hc_combos = []
    for _ in range(n_hillclimb):
        seed_p = rng.choice(seeds)
        hc_combos.append(perturb(seed_p, rng))

    with mp.Pool(
        processes=n_workers,
        initializer=_worker_init,
        initargs=(str(cache_path), coins),
    ) as pool:
        for params, metrics in pool.imap_unordered(
                _worker_eval, hc_combos, chunksize=50):
            hc_results.append({"params": params, "metrics": metrics})
            hpbar.update(1)
            # Actualizar seeds si encontramos algo mejor
            if metrics["total_pnl"] > results[0]["metrics"]["total_pnl"]:
                seeds.insert(0, params)
                seeds = seeds[:20]

    if HAS_TQDM: hpbar.close()
    else:        hpbar.close()

    results.extend(hc_results)
    results.sort(key=lambda r: r["metrics"]["total_pnl"], reverse=True)
    top30 = results[:30]

    print(f"  Total evaluaciones: {len(results):,}")
    print(f"  Mejor PnL encontrado: ${top30[0]['metrics']['total_pnl']:,.2f}")
    print(f"  Tiempo total optimización: {time.time()-t0:.1f}s")

    # ── 5. Generar informes ──────────────────────────────────────────────────
    print(f"\n[5/5] Generando informes en {out_dir.resolve()}…")

    xl_path   = out_dir / "resultados_optimizacion.xlsx"
    pdf_path  = out_dir / "informe_optimizacion.pdf"
    json_path = out_dir / "top_params.json"

    save_excel(top30, results, xl_path, coins)
    save_pdf(top30, pdf_path, args.days)
    save_json(top30, json_path)

    # ── Resumen por consola ──────────────────────────────────────────────────
    print_summary(top30)

    print(f"\n✓ Archivos generados:")
    print(f"  {xl_path.resolve()}")
    print(f"  {pdf_path.resolve()}")
    print(f"  {json_path.resolve()}")
    print()


if __name__ == "__main__":
    # IMPORTANTE: guard necesario en Windows para multiprocessing
    mp.freeze_support()
    main()
